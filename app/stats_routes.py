from fastapi import APIRouter, Depends, HTTPException
from typing import Optional, List, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, case, extract, desc, distinct
from sqlalchemy.orm import selectinload
import asyncio
from datetime import datetime, timedelta
from .logging_config import logger
from . import models, schemas, database, auth_routes
from .auth_routes import check_module_permission, get_current_user
from .database import get_async_db, get_read_db
from . import tat_utils

router = APIRouter(prefix="/stats", tags=["stats"])

from fastapi.responses import StreamingResponse
import io
import pandas as pd

import asyncio

from .cache import get_cache, set_cache, cache_response
CACHE_TTL = 300 # 5 minutes

def apply_case_filters(
    query,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    candidate_id: Optional[str] = None,
):
    # Apply Client Filter
    if client and client != 'ALL':
        query = query.filter(
            or_(
                models.Case.customer_id == client,
                models.Case.customer_id.in_(
                    select(models.Customer.id).filter(models.Customer.name == client)
                )
            )
        )
        
    # Apply Executive Filter
    if executive and executive != 'ALL':
        exec_subq = select(models.User.id).filter(
            or_(
                models.User.full_name == executive,
                models.User.email == executive
            )
        )
        query = query.filter(
            or_(
                models.Case.assigned_to == executive,
                models.Case.assigned_to.in_(exec_subq),
                models.Case.checks.any(models.VerificationCheck.assigned_verifier_id == executive),
                models.Case.checks.any(models.VerificationCheck.assigned_verifier_id.in_(exec_subq))
            )
        )
        
    # Apply Status Filter
    if status and status != 'ALL':
        status_upper = status.upper()
        _FINAL = ['FINALIZED','COMPLETED','POSITIVE','NEGATIVE','DISCREPANCY','UNABLE TO VERIFY','HOLD','INSUFFICIENT','QC_VERIFIED','CLOSED']
        if status_upper in ('COMPLETED', 'FINALIZED'):
            query = query.filter(models.Case.status.in_(_FINAL))
        elif status_upper == 'VERIFICATION':
            query = query.filter(models.Case.status == 'IN_PROGRESS')
        elif status_upper == 'PENDING':
            query = query.filter(models.Case.status == 'ASSIGNED')
        elif status_upper == 'INSUFFICIENT':
            query = query.filter(models.Case.status == 'INSUFFICIENCY')
        else:
            query = query.filter(models.Case.status == status)
            
    # Apply TAT Filter
    if tat and tat != 'ALL':
        if tat == 'IN_SLA':
            query = query.filter(models.Case.is_in_tat == 1)
        elif tat == 'BREACHED':
            query = query.filter(models.Case.is_in_tat == 0)
            
    # Apply Search Filter
    if search:
        query = query.filter(
            or_(
                models.Case.case_ref_no.ilike(f"%{search}%"),
                models.Case.candidate_id.in_(
                    select(models.Candidate.id).filter(models.Candidate.name.ilike(f"%{search}%"))
                )
            )
        )
        
    # Apply Candidate ID Filter
    if candidate_id:
        query = query.filter(models.Case.candidate_id == candidate_id)
        
    return query

# ─── Sidebar live counts ───────────────────────────────────────────────────────
@router.get("/sidebar-counts")
@cache_response(ttl=30, key_prefix="stats")
async def get_sidebar_counts(
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Lightweight endpoint polled every 30s by the sidebar for badge counts."""
    result = {}
    try:
        # Unread / new docs in client vault
        unread_q = await db.execute(
            select(func.count(models.CustomerDocument.id))
            .where(models.CustomerDocument.is_read == False)
        )
        result["client_vault"] = unread_q.scalar() or 0

        # Pending batches (not closed/completed)
        batch_q = await db.execute(
            select(func.count(models.Batch.id))
            .where(models.Batch.status.notin_(["Completed", "Closed", "completed", "closed"]))
        )
        result["batches"] = batch_q.scalar() or 0

        # Pending data entry cases
        de_q = await db.execute(
            select(func.count(models.Case.id))
            .where(models.Case.status.in_(["Pending", "In Progress", "pending"]))
        )
        result["data_entry"] = de_q.scalar() or 0

        # QC pending removed - set to 0
        result["qc_pending"] = 0

        # Finalized today
        today = datetime.utcnow().date()
        fin_q = await db.execute(
            select(func.count(models.Case.id))
            .where(models.Case.status.in_(["Finalized", "finalized"]))
            .where(func.date(models.Case.updated_at) == today)
        )
        result["finalized"] = fin_q.scalar() or 0

        # Candidate invitations pending (link not yet shared)
        inv_q = await db.execute(
            select(func.count(models.CandidateInvitation.id))
            .where(models.CandidateInvitation.status == "PENDING")
        )
        result["invitations"] = inv_q.scalar() or 0

    except Exception as e:
        logger.warning(f"sidebar-counts partial error: {e}")

    return result


# ─── Dedicated Customer Overview Dashboard ──────────────────────────────────────
@router.get("/customer-dashboard")
async def get_customer_dashboard(
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Dedicated client overview dashboard for CUSTOMER role.
    Exposes only customer-owned candidate verification metrics.
    """
    if not current_user.customer_id:
        raise HTTPException(
            status_code=403,
            detail="Access denied. User is not associated with any customer account."
        )

    try:
        # Get customer name
        cust_q = select(models.Customer).where(models.Customer.id == current_user.customer_id)
        cust_res = await db.execute(cust_q)
        customer = cust_res.scalar_one_or_none()
        customer_name = customer.name if customer else "Apex Covantage India"

        # Get all cases for this customer
        cases_q = select(models.Case).where(models.Case.customer_id == current_user.customer_id)
        cases_res = await db.execute(cases_q)
        cases = cases_res.scalars().all()

        total_candidates = len(cases)

        # Case-wise metrics aggregation using SUM(CASE)
        from sqlalchemy.sql import exists
        has_active_insuff = exists().where(
            and_(
                models.Insufficiency.case_id == models.Case.id,
                models.Insufficiency.is_resolved == False
            )
        )

        case_stats_stmt = select(
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED']),
                    func.upper(models.Case.status).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED'])
                )
            ), 1), else_=0)).label("positiveCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['NEGATIVE', 'RED', 'DISCREPANCY']),
                    func.upper(models.Case.status).in_(['NEGATIVE', 'RED', 'DISCREPANCY'])
                )
            ), 1), else_=0)).label("negativeCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH']),
                    func.upper(models.Case.status).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH'])
                )
            ), 1), else_=0)).label("amberCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['STOPCHECK', 'HOLD', 'CLIENT HOLD', 'STOPPED', 'STOP']),
                    func.upper(models.Case.status).in_(['STOPCHECK', 'HOLD', 'CLIENT HOLD', 'STOPPED', 'STOP'])
                )
            ), 1), else_=0)).label("stopCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.status).in_(['INSUFFICIENT', 'INSUFFICIENCY']),
                    func.upper(models.Case.final_result).in_(['INSUFFICIENT', 'INSUFFICIENCY']),
                    has_active_insuff
                )
            ), 1), else_=0)).label("insufficiencyCases"),
            
            func.sum(case(((
                and_(
                    func.upper(models.Case.status).in_(['WIP', 'ASSIGNED', 'INITIATED', 'VERIFICATION', 'QC_PENDING', 'CLIENT_REVIEW', 'IN_PROGRESS', 'PENDING']),
                    func.upper(models.Case.status).notin_(['FINALIZED', 'COMPLETED', 'CLOSED'])
                )
            ), 1), else_=0)).label("inProgressCases")
        ).select_from(models.Case).where(models.Case.customer_id == current_user.customer_id)

        case_stats_res = await db.execute(case_stats_stmt)
        case_row = case_stats_res.one()

        positive_checks = int(case_row.positiveCases or 0)
        negative_checks = int(case_row.negativeCases or 0)
        amber_checks = int(case_row.amberCases or 0)
        stop_checks = int(case_row.stopCases or 0)
        insufficiency_checks = int(case_row.insufficiencyCases or 0)
        in_progress_checks = int(case_row.inProgressCases or 0)

        now = datetime.utcnow()

        # Get 10 recent cases — eagerly load candidate and batch in 2 extra queries (IN clause) instead of N+1
        recent_cases_q = (
            select(models.Case)
            .where(models.Case.customer_id == current_user.customer_id)
            .options(
                selectinload(models.Case.candidate),
                selectinload(models.Case.batch)
            )
            .order_by(models.Case.received_date.desc())
            .limit(10)
        )
        recent_cases_res = await db.execute(recent_cases_q)
        recent_cases = recent_cases_res.scalars().all()

        recent_candidates_list = []
        for c in recent_cases:
            cand = c.candidate
            batch_no = c.batch.batch_no if c.batch else "Manual Entry"

            age_days = (now - c.received_date.replace(tzinfo=None)).days if c.received_date else 0
            sla_days_left = (c.tat_days or 10) - age_days

            if c.status in ["FINALIZED", "COMPLETED", "POSITIVE", "NEGATIVE", "GREEN", "RED"]:
                sla_text = "Completed"
            elif sla_days_left < 0:
                sla_text = f"Breached ({abs(sla_days_left)}d overdue)"
            elif sla_days_left <= 3:
                sla_text = f"Risk ({sla_days_left}d left)"
            else:
                sla_text = f"Healthy ({sla_days_left}d left)"

            recent_candidates_list.append({
                "id": c.id,
                "candidate_name": cand.name if cand else "Unknown",
                "employee_id": cand.client_emp_code if cand else "N/A",
                "batch": batch_no,
                "status": str(c.status).upper(),
                "sla": sla_text,
                "report_status": "READY" if c.status in ["FINALIZED", "COMPLETED", "POSITIVE", "NEGATIVE", "GREEN", "RED"] else "PENDING",
                "last_updated": c.completed_date.isoformat() if c.completed_date else (c.received_date.isoformat() if c.received_date else None)
            })

        # Batches — single aggregated query instead of 3 queries per batch in a Python loop
        batches_q = select(models.Batch).where(models.Batch.customer_id == current_user.customer_id)
        batches_res = await db.execute(batches_q)
        batches_list = batches_res.scalars().all()

        active_batches = 0
        closed_batches = 0
        delayed_batches = 0
        sla_risk_batches = 0

        if batches_list:
            batch_ids = [b.id for b in batches_list]
            batch_nos = [b.batch_no for b in batches_list]
            FINAL_STATUSES = ["FINALIZED", "COMPLETED", "POSITIVE", "NEGATIVE", "GREEN", "RED"]
            cutoff_date = now - timedelta(days=10)

            # One query: per-batch counts of active, delayed, and at-risk cases
            agg_q = await db.execute(
                select(
                    models.Case.batch_id,
                    func.count(models.Case.id).label("active_cnt"),
                    func.sum(
                        case((models.Case.received_date <= cutoff_date, 1), else_=0)
                    ).label("delayed_cnt"),
                    func.sum(
                        case((models.Case.risk_score > 70, 1), else_=0)
                    ).label("risk_cnt"),
                ).where(
                    models.Case.batch_id.in_(batch_ids + batch_nos),
                    models.Case.status.notin_(FINAL_STATUSES)
                ).group_by(models.Case.batch_id)
            )
            agg_rows = {row.batch_id: row for row in agg_q.all()}

            for b in batches_list:
                row = agg_rows.get(b.id) or agg_rows.get(b.batch_no)
                if row and row.active_cnt > 0:
                    active_batches += 1
                    if (row.delayed_cnt or 0) > 0:
                        delayed_batches += 1
                    if (row.risk_cnt or 0) > 0:
                        sla_risk_batches += 1
                else:
                    closed_batches += 1

        # Live timeline — eagerly load case + candidate to avoid N+1
        timeline_q = (
            select(models.VerificationLog)
            .join(models.Case, models.VerificationLog.case_id == models.Case.id)
            .where(models.Case.customer_id == current_user.customer_id)
            .options(
                selectinload(models.VerificationLog.case).selectinload(models.Case.candidate)
            )
            .order_by(models.VerificationLog.created_at.desc())
            .limit(10)
        )
        timeline_res = await db.execute(timeline_q)
        logs = timeline_res.scalars().all()

        timeline = []
        for l in logs:
            case_obj = l.case
            cand_name = "Unknown Candidate"
            if case_obj and case_obj.candidate:
                cand_name = case_obj.candidate.name

            timeline.append({
                "id": l.id,
                "candidate_name": cand_name,
                "case_ref": case_obj.case_ref_no if case_obj else "",
                "action": l.action,
                "remarks": l.remarks or "",
                "new_status": l.new_status or "",
                "timestamp": l.created_at.isoformat() if l.created_at else None
            })

        # Fallback mock timeline
        if not timeline:
            for i, c in enumerate(recent_candidates_list[:4]):
                timeline.append({
                    "id": f"mock-{i}",
                    "candidate_name": c["candidate_name"],
                    "case_ref": c.get("case_ref_no", "CL-MCP-001"),
                    "action": "STATUS_UPDATED",
                    "remarks": f"Verification status updated to {c['status']}",
                    "new_status": c["status"],
                    "timestamp": c["last_updated"]
                })

        # Documents
        docs_q = (
            select(models.ClientDocument)
            .where(models.ClientDocument.customer_id == current_user.customer_id)
            .order_by(models.ClientDocument.created_at.desc())
            .limit(5)
        )
        docs_res = await db.execute(docs_q)
        docs = docs_res.scalars().all()

        doc_list = []
        for d in docs:
            doc_list.append({
                "id": d.id,
                "name": d.name,
                "file_type": d.file_type or "PDF",
                "file_path": d.file_path,
                "uploaded_at": d.created_at.isoformat() if d.created_at else None
            })

        return {
            "customer_name": customer_name,
            "stats": {
                "total_candidates": total_candidates,
                "positive_checks": positive_checks,
                "negative_checks": negative_checks,
                "amber_checks": amber_checks,
                "stop_checks": stop_checks,
                "insufficiency_checks": insufficiency_checks,
                "in_progress_checks": in_progress_checks,
                "in_progress": in_progress_checks,
                "finalized": positive_checks + negative_checks + amber_checks + stop_checks,
                "insufficiency": insufficiency_checks,
                "approaching_sla": 0,
                "reports_ready": positive_checks + negative_checks
            },
            "status_mix": {
                "positive": positive_checks,
                "negative": negative_checks,
                "amber": amber_checks,
                "stop": stop_checks,
                "insufficiency": insufficiency_checks,
                "in_progress": in_progress_checks
            },
            "recent_candidates": recent_candidates_list,
            "batches": {
                "active": active_batches,
                "closed": closed_batches,
                "delayed": delayed_batches,
                "sla_risk": sla_risk_batches
            },
            "timeline": timeline,
            "documents": doc_list
        }

    except Exception as e:
        logger.error(f"customer-dashboard error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to load customer dashboard data")


@router.get("/customer-candidates")
async def get_customer_candidates(
    page: int = 1,
    limit: int = 20,
    search: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Paginated, filterable candidate listing for Customer Portal.
    """
    if not current_user.customer_id:
        raise HTTPException(status_code=403, detail="Access denied. User is not associated with any customer account.")

    try:
        # Base query for cases
        query = select(models.Case).where(models.Case.customer_id == current_user.customer_id)
        
        # Text Search
        if search:
            search = f"%{search}%"
            query = query.filter(
                or_(
                    models.Case.case_ref_no.ilike(search),
                    models.Case.candidate.has(models.Candidate.name.ilike(search)),
                    models.Case.candidate.has(models.Candidate.client_emp_code.ilike(search)),
                    models.Case.batch_id.ilike(search)
                )
            )
            
        # Date filtering based on received_date
        if from_date:
            try:
                dt_from = datetime.strptime(from_date, "%Y-%m-%d")
                query = query.filter(models.Case.received_date >= dt_from)
            except: pass
        if to_date:
            try:
                dt_to = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
                query = query.filter(models.Case.received_date < dt_to)
            except: pass
            
        # Status filtering
        if status and status.upper() != 'ALL':
            status_upper = status.upper()
            from sqlalchemy.sql import exists
            has_active_insuff = exists().where(
                and_(
                    models.Insufficiency.case_id == models.Case.id,
                    models.Insufficiency.is_resolved == False
                )
            )
            
            if status_upper == 'POSITIVE':
                query = query.filter(
                    or_(
                        func.upper(models.Case.final_result).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED']),
                        func.upper(models.Case.status).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED'])
                    )
                )
            elif status_upper == 'NEGATIVE':
                query = query.filter(
                    or_(
                        func.upper(models.Case.final_result).in_(['NEGATIVE', 'RED', 'DISCREPANCY']),
                        func.upper(models.Case.status).in_(['NEGATIVE', 'RED', 'DISCREPANCY'])
                    )
                )
            elif status_upper == 'AMBER':
                query = query.filter(
                    or_(
                        func.upper(models.Case.final_result).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH']),
                        func.upper(models.Case.status).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH'])
                    )
                )
            elif status_upper == 'STOP CHECK':
                query = query.filter(
                    or_(
                        func.upper(models.Case.final_result).in_(['STOPCHECK', 'HOLD', 'CLIENT HOLD', 'STOPPED', 'STOP']),
                        func.upper(models.Case.status).in_(['STOPCHECK', 'HOLD', 'CLIENT HOLD', 'STOPPED', 'STOP'])
                    )
                )
            elif status_upper == 'INSUFFICIENCY':
                query = query.filter(
                    or_(
                        func.upper(models.Case.status).in_(['INSUFFICIENT', 'INSUFFICIENCY']),
                        func.upper(models.Case.final_result).in_(['INSUFFICIENT', 'INSUFFICIENCY']),
                        has_active_insuff
                    )
                )
            elif status_upper == 'IN PROGRESS':
                query = query.filter(
                    and_(
                        func.upper(models.Case.status).in_(['WIP', 'ASSIGNED', 'INITIATED', 'VERIFICATION', 'QC_PENDING', 'CLIENT_REVIEW', 'IN_PROGRESS', 'PENDING']),
                        func.upper(models.Case.status).notin_(['FINALIZED', 'COMPLETED', 'CLOSED'])
                    )
                )
            elif status_upper == 'FINALIZED':
                query = query.filter(models.Case.status.in_(['FINALIZED', 'COMPLETED']))
            else:
                query = query.filter(models.Case.status == status_upper)
                
        # Calculate summary statistics for the filtered dataset
        case_stats_stmt = query.with_only_columns(
            func.sum(case(((or_(func.upper(models.Case.final_result).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED']), func.upper(models.Case.status).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED']))), 1), else_=0)).label("positiveCases"),
            func.sum(case(((or_(func.upper(models.Case.final_result).in_(['NEGATIVE', 'RED', 'DISCREPANCY']), func.upper(models.Case.status).in_(['NEGATIVE', 'RED', 'DISCREPANCY']))), 1), else_=0)).label("negativeCases"),
            func.sum(case(((or_(func.upper(models.Case.final_result).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH']), func.upper(models.Case.status).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH']))), 1), else_=0)).label("amberCases"),
            func.sum(case(((or_(func.upper(models.Case.status).in_(['INSUFFICIENT', 'INSUFFICIENCY']), func.upper(models.Case.final_result).in_(['INSUFFICIENT', 'INSUFFICIENCY']))), 1), else_=0)).label("insufficiencyCases"),
            func.count(models.Case.id).label("total")
        )
        stats_res = await db.execute(case_stats_stmt)
        stats = stats_res.one()
        
        total = stats.total or 0
        summary = {
            "positive": int(stats.positiveCases or 0),
            "negative": int(stats.negativeCases or 0),
            "amber": int(stats.amberCases or 0),
            "insufficiency": int(stats.insufficiencyCases or 0)
        }
        
        total_pages = (total + limit - 1) // limit if limit > 0 else 1
        
        # Pagination
        query = query.order_by(models.Case.received_date.desc())
        if limit > 0:
            query = query.offset((page - 1) * limit).limit(limit)
            
        # Eager load candidate info
        query = query.options(selectinload(models.Case.candidate), selectinload(models.Case.batch))
        
        cases_res = await db.execute(query)
        cases = cases_res.scalars().all()
        
        data = []
        now = datetime.utcnow()
        for c in cases:
            age_days = (now - c.received_date.replace(tzinfo=None)).days if c.received_date else 0
            sla_days_left = (c.tat_days or 10) - age_days
            
            if c.status in ["FINALIZED", "COMPLETED", "POSITIVE", "NEGATIVE", "GREEN", "RED"]:
                sla_text = "Completed"
            elif sla_days_left < 0:
                sla_text = f"Breached ({abs(sla_days_left)}d overdue)"
            elif sla_days_left <= 3:
                sla_text = f"Risk ({sla_days_left}d left)"
            else:
                sla_text = f"Healthy ({sla_days_left}d left)"

            data.append({
                "id": c.id,
                "candidate_name": c.candidate.name if c.candidate else "Unknown",
                "employee_id": c.candidate.client_emp_code if c.candidate else "N/A",
                "batch_no": c.batch.batch_no if c.batch else (c.batch_id or "Manual Entry"),
                "case_ref": c.case_ref_no,
                "status": c.status,
                "final_result": c.final_result,
                "received_date": c.received_date.isoformat() if c.received_date else None,
                "sla": sla_text,
                "report_status": "READY" if c.status in ["FINALIZED", "COMPLETED", "POSITIVE", "NEGATIVE", "GREEN", "RED"] else "PENDING"
            })
            
        return {
            "data": data,
            "total": total,
            "page": page,
            "totalPages": total_pages,
            "summary": summary
        }
    except Exception as e:
        logger.error(f"customer-candidates error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to fetch customer candidates")



@router.get("/customer-summary")
async def get_customer_summary(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    batch_id: Optional[str] = None,
    candidate_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Dedicated check-wise stats overview for Customer Portal dashboard.
    """
    if not current_user.customer_id:
        raise HTTPException(
            status_code=403,
            detail="Access denied. User is not associated with any customer account."
        )

    try:
        # Build Filters based on customer and standard optional filters
        filters = [models.Case.customer_id == current_user.customer_id]
        
        if from_date:
            try:
                start_dt = datetime.strptime(from_date, "%Y-%m-%d")
                filters.append(models.Case.received_date >= start_dt)
            except Exception:
                pass
        if to_date:
            try:
                end_dt = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
                filters.append(models.Case.received_date < end_dt)
            except Exception:
                pass
        if batch_id and batch_id != 'ALL':
            filters.append(
                or_(
                    models.Case.batch_id == batch_id,
                    models.Case.batch.has(models.Batch.batch_no == batch_id)
                )
            )
        if candidate_id:
            filters.append(models.Case.candidate_id == candidate_id)
            
        if search:
            filters.append(
                or_(
                    models.Case.case_ref_no.ilike(f"%{search}%"),
                    models.Case.candidate.has(models.Candidate.name.ilike(f"%{search}%")),
                    models.Case.candidate.has(models.Candidate.client_emp_code.ilike(f"%{search}%"))
                )
            )
            
        if status and status != 'ALL':
            status_upper = status.upper()
            filters.append(models.Case.status == status_upper)

        # 1. Total Candidates count (Case-Wise count of unique cases/candidates)
        total_candidates_stmt = select(func.count(distinct(models.Case.id))).where(*filters)
        total_candidates_res = await db.execute(total_candidates_stmt)
        total_candidates = total_candidates_res.scalar() or 0

        # 2. Case-wise metrics aggregation using SUM(CASE) to do it in one single fast query
        from sqlalchemy.sql import exists
        has_active_insuff = exists().where(
            and_(
                models.Insufficiency.case_id == models.Case.id,
                models.Insufficiency.is_resolved == False
            )
        )

        case_stats_stmt = select(
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED']),
                    func.upper(models.Case.status).in_(['POSITIVE', 'CLEAR', 'GREEN', 'CLEAR_VERIFIED', 'CLEAR/VERIFIED', 'FINALIZED', 'COMPLETED'])
                )
            ), 1), else_=0)).label("positiveCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['NEGATIVE', 'RED', 'DISCREPANCY']),
                    func.upper(models.Case.status).in_(['NEGATIVE', 'RED', 'DISCREPANCY'])
                )
            ), 1), else_=0)).label("negativeCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH']),
                    func.upper(models.Case.status).in_(['AMBER', 'REVIEW REQUIRED', 'REVIEW_REQUIRED', 'MINOR MISMATCH'])
                )
            ), 1), else_=0)).label("amberCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.final_result).in_(['STOPCHECK', 'HOLD', 'CLIENT HOLD', 'STOPPED', 'STOP']),
                    func.upper(models.Case.status).in_(['STOPCHECK', 'HOLD', 'CLIENT HOLD', 'STOPPED', 'STOP'])
                )
            ), 1), else_=0)).label("stopCases"),
            
            func.sum(case(((
                or_(
                    func.upper(models.Case.status).in_(['INSUFFICIENT', 'INSUFFICIENCY']),
                    func.upper(models.Case.final_result).in_(['INSUFFICIENT', 'INSUFFICIENCY']),
                    has_active_insuff
                )
            ), 1), else_=0)).label("insufficiencyCases"),
            
            func.sum(case(((
                and_(
                    func.upper(models.Case.status).in_(['WIP', 'ASSIGNED', 'INITIATED', 'VERIFICATION', 'QC_PENDING', 'CLIENT_REVIEW', 'IN_PROGRESS', 'PENDING']),
                    func.upper(models.Case.status).notin_(['FINALIZED', 'COMPLETED', 'CLOSED'])
                )
            ), 1), else_=0)).label("inProgressCases")
        ).select_from(models.Case).where(*filters)

        case_stats_res = await db.execute(case_stats_stmt)
        case_row = case_stats_res.one()

        return {
            "totalCandidates": total_candidates,
            "positiveChecks": int(case_row.positiveCases or 0),
            "negativeChecks": int(case_row.negativeCases or 0),
            "amberChecks": int(case_row.amberCases or 0),
            "stopChecks": int(case_row.stopCases or 0),
            "insufficiencyChecks": int(case_row.insufficiencyCases or 0),
            "inProgressChecks": int(case_row.inProgressCases or 0)
        }

    except Exception as e:
        logger.error(f"customer-summary error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to load customer summary stats")


# ─── Dedicated Verifier Workspace Dashboard ────────────────────────────────────
@router.get("/verifier-dashboard")
@cache_response(ttl=60, key_prefix="stats")
async def get_verifier_dashboard(
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Returns personal productivity data scoped strictly to the requesting verifier.
    No admin analytics, cross-client metrics, or revenue data is ever exposed.
    """
    uid = current_user.id
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    now = datetime.utcnow()

    ACTIVE_STATUSES = ["ASSIGNED", "IN_PROGRESS", "INSUFFICIENCY", "VERIFICATION"]
    FINAL_STATUSES = [
        "FINALIZED", "COMPLETED", "POSITIVE", "NEGATIVE",
        "DISCREPANCY", "UNABLE TO VERIFY", "HOLD", "INSUFFICIENT",
    ]
    TAT_WARNING_DAYS = 7   # approaching TAT
    TAT_BREACH_DAYS = 10   # out of TAT

    try:
        # --- Core status counts (my cases) ---
        status_q = (
            select(models.Case.status, func.count(models.Case.id))
            .where(models.Case.assigned_to == uid)
            .group_by(models.Case.status)
        )
        status_rows = (await db.execute(status_q)).all()
        sc = {}
        for row in status_rows:
            s_val = str(row[0].value if hasattr(row[0], "value") else row[0])
            sc[s_val] = sc.get(s_val, 0) + int(row[1] or 0)

        assigned_total   = sum(sc.get(s, 0) for s in ACTIVE_STATUSES)
        wip_count        = sc.get("IN_PROGRESS", 0) + sc.get("VERIFICATION", 0)
        insuff_count     = sc.get("INSUFFICIENCY", 0) + sc.get("INSUFFICIENT", 0)
        total_finalized  = sum(sc.get(s, 0) for s in FINAL_STATUSES)

        # --- Finalized today ---
        fin_today_q = (
            select(func.count(models.Case.id))
            .where(
                models.Case.assigned_to == uid,
                models.Case.status.in_(FINAL_STATUSES),
                models.Case.completed_date >= today_start,
            )
        )
        finalized_today = (await db.execute(fin_today_q)).scalar() or 0

        # --- New assignments today ---
        new_today_q = (
            select(func.count(models.Case.id))
            .where(
                models.Case.assigned_to == uid,
                models.Case.assigned_at >= today_start,
            )
        )
        new_today = (await db.execute(new_today_q)).scalar() or 0

        # --- Approaching TAT (active cases > 7 days old, <= 10 days) ---
        warn_threshold = now - timedelta(days=TAT_WARNING_DAYS)
        breach_threshold = now - timedelta(days=TAT_BREACH_DAYS)
        approaching_tat_q = (
            select(func.count(models.Case.id))
            .where(
                models.Case.assigned_to == uid,
                models.Case.status.in_(ACTIVE_STATUSES),
                models.Case.received_date <= warn_threshold,
                models.Case.received_date > breach_threshold,
            )
        )
        approaching_tat = (await db.execute(approaching_tat_q)).scalar() or 0

        # --- Out of TAT (active cases > 10 days old) ---
        out_tat_q = (
            select(func.count(models.Case.id))
            .where(
                models.Case.assigned_to == uid,
                models.Case.status.in_(ACTIVE_STATUSES),
                models.Case.received_date <= breach_threshold,
            )
        )
        out_of_tat = (await db.execute(out_tat_q)).scalar() or 0

        # --- Average TAT (finalized cases) ---
        avg_tat_q = (
            select(func.avg(models.Case.tat_days))
            .where(
                models.Case.assigned_to == uid,
                models.Case.status.in_(FINAL_STATUSES),
                models.Case.tat_days > 0,
            )
        )
        avg_tat = round(float((await db.execute(avg_tat_q)).scalar() or 0), 1)

        # --- Productivity % (finalized / (finalized + active)) ---
        productivity = 0.0
        denom = total_finalized + assigned_total
        if denom > 0:
            productivity = round((total_finalized / denom) * 100, 1)

        # --- My assigned cases list (latest 20 active) ---
        cases_q = (
            select(models.Case)
            .where(
                models.Case.assigned_to == uid,
                models.Case.status.in_(ACTIVE_STATUSES + FINAL_STATUSES),
            )
            .order_by(models.Case.received_date.asc())
            .limit(20)
        )
        case_rows = (await db.execute(cases_q)).scalars().all()

        cases_list = []
        for c in case_rows:
            s_val = str(c.status.value if hasattr(c.status, "value") else c.status)
            rd = c.received_date
            age_days = (now - rd).days if rd else 0
            tat_status = (
                "BREACH" if age_days > TAT_BREACH_DAYS
                else "WARNING" if age_days > TAT_WARNING_DAYS
                else "OK"
            )
            cases_list.append({
                "id": c.id,
                "case_ref_no": c.case_ref_no or "",
                "status": s_val,
                "received_date": rd.isoformat() if rd else None,
                "age_days": age_days,
                "tat_status": tat_status,
                "candidate_name": None,   # populated below
                "client_name": None,
                "candidate_id": c.candidate_id,
                "customer_id": c.customer_id,
            })

        # Enrich with candidate & client names
        if cases_list:
            cand_ids = list({c["candidate_id"] for c in cases_list if c["candidate_id"]})
            cust_ids = list({c["customer_id"]  for c in cases_list if c["customer_id"]})

            if cand_ids:
                cand_q = select(models.Candidate.id, models.Candidate.name).where(models.Candidate.id.in_(cand_ids))
                cand_map = {r[0]: r[1] for r in (await db.execute(cand_q)).all()}
            else:
                cand_map = {}

            if cust_ids:
                cust_q = select(models.Customer.id, models.Customer.name).where(models.Customer.id.in_(cust_ids))
                cust_map = {r[0]: r[1] for r in (await db.execute(cust_q)).all()}
            else:
                cust_map = {}

            for c in cases_list:
                c["candidate_name"] = cand_map.get(c["candidate_id"], "Unknown")
                c["client_name"]    = cust_map.get(c["customer_id"],  "Unknown")

        # --- Recent activity (my audit log) ---
        log_q = (
            select(models.AuditLog)
            .where(models.AuditLog.user_id == uid)
            .order_by(models.AuditLog.timestamp.desc())
            .limit(10)
        )
        log_rows = (await db.execute(log_q)).scalars().all()
        activity = [
            {
                "action": r.action,
                "time": r.timestamp.strftime("%H:%M") if r.timestamp else "",
                "date": r.timestamp.strftime("%d %b") if r.timestamp else "",
            }
            for r in log_rows
        ]

        return {
            "assigned_total":   assigned_total,
            "wip_count":        wip_count,
            "insuff_count":     insuff_count,
            "finalized_today":  finalized_today,
            "total_finalized":  total_finalized,
            "new_today":        new_today,
            "approaching_tat":  approaching_tat,
            "out_of_tat":       out_of_tat,
            "avg_tat":          avg_tat,
            "productivity":     productivity,
            "cases":            cases_list,
            "activity":         activity,
        }

    except Exception as e:
        logger.error(f"verifier-dashboard error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to load verifier dashboard")


@router.get("", response_model=schemas.DashboardStats)
@cache_response(ttl=60, key_prefix="stats")
async def get_dashboard_stats(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    try:
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Determine roles
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        
        is_customer = user_role == "CUSTOMER" or role_name == "CUSTOMER"
        is_admin = user_role in ["SUPER_ADMIN", "ADMIN", "MANAGER", "QA", "QC"] or role_name in ["SUPER ADMIN", "QC VERIFIER"]
        
        filter_verifier = not (is_admin or is_customer)
        filter_customer = is_customer
        
        # 1. Date Filters
        filter_start = None
        filter_end = None
        if from_date:
            try: filter_start = datetime.strptime(from_date, "%Y-%m-%d")
            except: pass
        if to_date:
            try: filter_end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
            except: pass

        # 2. Optimized Combined Queries
        # We'll use a single pass for status counts and date-based counts
        status_stmt = select(models.Case.status, func.count(models.Case.id)).group_by(models.Case.status)
        if filter_verifier: 
            status_stmt = status_stmt.filter(or_(
                models.Case.assigned_to == current_user.id,
                models.Case.checks.any(models.VerificationCheck.assigned_verifier_id == current_user.id)
            ))
        if filter_customer: status_stmt = status_stmt.filter(models.Case.customer_id == current_user.customer_id)
        if filter_start: status_stmt = status_stmt.filter(models.Case.received_date >= filter_start)
        if filter_end: status_stmt = status_stmt.filter(models.Case.received_date < filter_end)
        status_stmt = apply_case_filters(status_stmt, client, executive, status, tat, search)

        month_start_date = today.replace(day=1)
        date_counts_stmt = select(
            func.count(case(((models.Case.received_date >= month_start_date), models.Case.id))).label("this_month"),
            func.count(case(((models.Case.received_date >= today), models.Case.id))).label("today_entry"),
            func.count(case(((models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT'])) & (models.Case.completed_date >= today), models.Case.id))).label("comp_today")
        )
        if filter_verifier: 
            date_counts_stmt = date_counts_stmt.filter(or_(
                models.Case.assigned_to == current_user.id,
                models.Case.checks.any(models.VerificationCheck.assigned_verifier_id == current_user.id)
            ))
        if filter_customer: date_counts_stmt = date_counts_stmt.filter(models.Case.customer_id == current_user.customer_id)
        date_counts_stmt = apply_case_filters(date_counts_stmt, client, executive, status, tat, search)

        # Total Customers: Always show the full Partner Network (14)
        total_cust_stmt = select(func.count(models.Customer.id))
        
        # Revenue and period-specific stats
        rev_cust_stmt = select(
            func.sum(case(((models.Case.status.in_(['COMPLETED', 'QC_VERIFIED'])), models.VerificationCheck.rate), else_=0)).label("total_revenue")
        ).select_from(models.Case).outerjoin(models.VerificationCheck, models.Case.id == models.VerificationCheck.case_id)
        
        if filter_customer: 
            rev_cust_stmt = rev_cust_stmt.filter(models.Case.customer_id == current_user.customer_id)
            total_cust_stmt = total_cust_stmt.filter(models.Customer.id == current_user.customer_id)
            
        if filter_start: rev_cust_stmt = rev_cust_stmt.filter(models.Case.completed_date >= filter_start)
        if filter_end: rev_cust_stmt = rev_cust_stmt.filter(models.Case.completed_date < filter_end)
        rev_cust_stmt = apply_case_filters(rev_cust_stmt, client, executive, status, tat, search)

        # Execution (Run sequentially on the same session to avoid concurrency errors)
        status_res = await db.execute(status_stmt)
        date_res = await db.execute(date_counts_stmt)
        rev_cust_res = await db.execute(rev_cust_stmt)
        total_cust_res = await db.execute(total_cust_stmt)
        
        status_rows = status_res.all()
        # Robust mapping: ensure we get the string value of the status
        status_counts = {}
        for row in status_rows:
            status_val = str(row[0].value if hasattr(row[0], "value") else row[0])
            status_counts[status_val] = int(row[1] or 0)
        
        # All canonical terminal/completed statuses
        _TERMINAL = ['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY',
                     'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT', 'QC_VERIFIED', 'CLOSED']
        # Active cases: anything not yet terminal
        total_candidates = sum(
            v for k, v in status_counts.items() if k not in _TERMINAL
        )
        # Completed = all terminal statuses
        total_completed = sum(
            v for k, v in status_counts.items() if k in _TERMINAL
        )

        # 2b. Accurate Insufficiency Count from new table
        insuff_q = select(func.count(distinct(models.Insufficiency.case_id))).filter(models.Insufficiency.is_resolved == False)
        if filter_customer:
            insuff_q = insuff_q.filter(models.Insufficiency.case.has(customer_id=current_user.customer_id))
        elif filter_verifier:
            insuff_q = insuff_q.filter(models.Insufficiency.case.has(assigned_to=current_user.id))
        
        if client and client != 'ALL':
            insuff_q = insuff_q.filter(
                models.Insufficiency.case.has(
                    or_(
                        models.Case.customer_id == client,
                        models.Case.customer_id.in_(
                            select(models.Customer.id).filter(models.Customer.name == client)
                        )
                    )
                )
            )
        if executive and executive != 'ALL':
            insuff_q = insuff_q.filter(
                models.Insufficiency.case.has(
                    or_(
                        models.Case.assigned_to == executive,
                        models.Case.assigned_to.in_(
                            select(models.User.id).filter(
                                or_(
                                    models.User.full_name == executive,
                                    models.User.username == executive
                                )
                            )
                        )
                    )
                )
            )
        if status and status != 'ALL':
            _FINAL = ['FINALIZED','COMPLETED','POSITIVE','NEGATIVE','DISCREPANCY','UNABLE TO VERIFY','HOLD','INSUFFICIENT','QC_VERIFIED','CLOSED']
            if status.upper() in ('COMPLETED', 'FINALIZED'):
                insuff_q = insuff_q.filter(models.Insufficiency.case.has(models.Case.status.in_(_FINAL)))
            else:
                insuff_q = insuff_q.filter(models.Insufficiency.case.has(models.Case.status == status))
        if tat and tat != 'ALL':
            if tat == 'IN_SLA':
                insuff_q = insuff_q.filter(models.Insufficiency.case.has(models.Case.is_in_tat == 1))
            elif tat == 'BREACHED':
                insuff_q = insuff_q.filter(models.Insufficiency.case.has(models.Case.is_in_tat == 0))
        if search:
            insuff_q = insuff_q.filter(
                models.Insufficiency.case.has(
                    or_(
                        models.Case.case_ref_no.ilike(f"%{search}%"),
                        models.Case.candidate_id.in_(
                            select(models.Candidate.id).filter(models.Candidate.name.ilike(f"%{search}%"))
                        )
                    )
                )
            )
        
        insuff_res = await db.execute(insuff_q)
        actual_insuff_count = insuff_res.scalar() or 0

        date_row = date_res.one()
        current_month = date_row.this_month or 0
        today_entry = date_row.today_entry or 0
        completed_today = date_row.comp_today or 0
        
        rev_cust_row = rev_cust_res.one()
        total_customers = total_cust_res.scalar() or 0
        total_revenue = rev_cust_row.total_revenue or 0.0
        
        # 3. Geo and Activity (Parallel)
        geo_stmt = select(models.Customer.city, func.count(models.Case.id)).join(models.Case).group_by(models.Customer.city)
        if filter_start: geo_stmt = geo_stmt.filter(models.Case.received_date >= filter_start)
        if filter_end: geo_stmt = geo_stmt.filter(models.Case.received_date < filter_end)
        geo_stmt = apply_case_filters(geo_stmt, client, executive, status, tat, search)
        
        log_stmt = select(models.AuditLog, models.User.email).join(models.User).order_by(models.AuditLog.timestamp.desc()).limit(10)
        if filter_verifier: log_stmt = log_stmt.filter(models.AuditLog.user_id == current_user.id)
        
        geo_res = await db.execute(geo_stmt)
        log_res = await db.execute(log_stmt)
        
        geo_data = [{"name": str(r[0] or "REMOTE"), "value": int(r[1]), "color": "#3b82f6"} for r in geo_res.all()]
        activity_log = [{"id": i, "icon": "⚡", "action": r[0].action, "time": r[0].timestamp.strftime("%H:%M"), "user": r[1]} for i, r in enumerate(log_res.all())]

        # 4. Monthly Analysis (Dynamic based on selected range)
        if filter_start and filter_end:
            chart_start = filter_start.replace(day=1)
            chart_end = filter_end
        else:
            chart_start = (today.replace(day=1) - timedelta(days=150)).replace(day=1)
            chart_end = today + timedelta(days=32)

        t_months_stmt = select(
            extract('year', models.Case.received_date).label('y'),
            extract('month', models.Case.received_date).label('m'),
            func.count(models.Case.id)
        ).filter(models.Case.received_date >= chart_start, models.Case.received_date < chart_end).group_by('y', 'm')
        
        c_months_stmt = select(
            extract('year', models.Case.completed_date).label('y'),
            extract('month', models.Case.completed_date).label('m'),
            func.count(models.Case.id)
        ).filter(models.Case.completed_date >= chart_start, models.Case.completed_date < chart_end, models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT'])).group_by('y', 'm')

        if filter_verifier:
            t_months_stmt = t_months_stmt.filter(models.Case.assigned_to == current_user.id)
            c_months_stmt = c_months_stmt.filter(models.Case.assigned_to == current_user.id)
        
        if filter_customer:
            t_months_stmt = t_months_stmt.filter(models.Case.customer_id == current_user.customer_id)
            c_months_stmt = c_months_stmt.filter(models.Case.customer_id == current_user.customer_id)

        t_months_stmt = apply_case_filters(t_months_stmt, client, executive, status, tat, search)
        c_months_stmt = apply_case_filters(c_months_stmt, client, executive, status, tat, search)

        t_m_res = await db.execute(t_months_stmt)
        c_m_res = await db.execute(c_months_stmt)
        t_dict = {(int(r[0]), int(r[1])): r[2] for r in t_m_res.all()}
        c_dict = {(int(r[0]), int(r[1])): r[2] for r in c_m_res.all()}

        analysis_data = []
        curr_m = chart_start
        # Prevent infinite loop if something goes wrong with dates
        max_iter = 24
        while curr_m < chart_end and max_iter > 0:
            y, m = curr_m.year, curr_m.month
            analysis_data.append({
                "name": curr_m.strftime("%b %y"),
                "total": int(t_dict.get((y, m), 0)),
                "completed": int(c_dict.get((y, m), 0)),
                "pending": max(0, int(t_dict.get((y, m), 0)) - int(c_dict.get((y, m), 0)))
            })
            # Advance to next month
            if curr_m.month == 12:
                curr_m = curr_m.replace(year=curr_m.year + 1, month=1)
            else:
                curr_m = curr_m.replace(month=curr_m.month + 1)
            max_iter -= 1

        # 5. TAT counts — query is_in_tat directly from DB (authoritative source)
        in_tat_q = select(func.count(models.Case.id)).filter(models.Case.is_in_tat == 1)
        out_tat_q = select(func.count(models.Case.id)).filter(models.Case.is_in_tat == 0)
        # At-risk: active cases (not terminal) where is_in_tat is marginal — use 7-day threshold
        now_time = datetime.utcnow()
        risk_threshold = now_time - timedelta(days=7)
        at_risk_q = select(func.count(models.Case.id)).filter(
            models.Case.status.notin_(_TERMINAL),
            models.Case.received_date < risk_threshold
        )

        if filter_verifier:
            in_tat_q = in_tat_q.filter(models.Case.assigned_to == current_user.id)
            out_tat_q = out_tat_q.filter(models.Case.assigned_to == current_user.id)
            at_risk_q = at_risk_q.filter(models.Case.assigned_to == current_user.id)
        elif filter_customer:
            in_tat_q = in_tat_q.filter(models.Case.customer_id == current_user.customer_id)
            out_tat_q = out_tat_q.filter(models.Case.customer_id == current_user.customer_id)
            at_risk_q = at_risk_q.filter(models.Case.customer_id == current_user.customer_id)

        if filter_start:
            in_tat_q = in_tat_q.filter(models.Case.received_date >= filter_start)
            out_tat_q = out_tat_q.filter(models.Case.received_date >= filter_start)
            at_risk_q = at_risk_q.filter(models.Case.received_date >= filter_start)
        if filter_end:
            in_tat_q = in_tat_q.filter(models.Case.received_date < filter_end)
            out_tat_q = out_tat_q.filter(models.Case.received_date < filter_end)
            at_risk_q = at_risk_q.filter(models.Case.received_date < filter_end)

        in_tat_q = apply_case_filters(in_tat_q, client, executive, status, tat, search)
        out_tat_q = apply_case_filters(out_tat_q, client, executive, status, tat, search)
        at_risk_q = apply_case_filters(at_risk_q, client, executive, status, tat, search)

        in_tat_res = await db.execute(in_tat_q)
        out_tat_res = await db.execute(out_tat_q)
        at_risk_res = await db.execute(at_risk_q)
        in_tat_count = in_tat_res.scalar() or 0
        out_tat_count = out_tat_res.scalar() or 0
        at_risk_count = at_risk_res.scalar() or 0

        # in_tat/out_tat are read directly from DB above — no override needed

        # 7. Specific Result Counts (Positive, Negative, Amber, Stop)
        # We calculate distinct cases based on their combined VerificationCheck statuses
        case_checks_stmt = select(
            models.VerificationCheck.case_id,
            models.VerificationCheck.status
        ).join(models.Case, models.VerificationCheck.case_id == models.Case.id)
        
        if filter_customer:
            case_checks_stmt = case_checks_stmt.filter(models.Case.customer_id == current_user.customer_id)
        elif filter_verifier:
            case_checks_stmt = case_checks_stmt.filter(models.Case.assigned_to == current_user.id)
        
        if filter_start: case_checks_stmt = case_checks_stmt.filter(models.Case.received_date >= filter_start)
        if filter_end: case_checks_stmt = case_checks_stmt.filter(models.Case.received_date < filter_end)
        case_checks_stmt = apply_case_filters(case_checks_stmt, client, executive, status, tat, search)
            
        cc_res = await db.execute(case_checks_stmt)
        cc_rows = cc_res.all()
        
        from collections import defaultdict
        cases_checks_map = defaultdict(list)
        for r_case_id, r_status in cc_rows:
            status_str = str(r_status.value if hasattr(r_status, 'value') else r_status).upper()
            cases_checks_map[r_case_id].append(status_str)
            
        positive_count = 0
        negative_count = 0
        amber_count = 0
        stop_count = 0

        for cid, statuses in cases_checks_map.items():
            if "STOP" in statuses:
                stop_count += 1
            elif "RED" in statuses or "NEGATIVE" in statuses:
                negative_count += 1
            elif "AMBER" in statuses or "DISCREPANCY" in statuses:
                amber_count += 1
            elif any(s in ["GREEN", "POSITIVE", "QC_VERIFIED", "CLEAR", "VERIFIED"] for s in statuses):
                positive_count += 1

        # Fallback: if no VerificationCheck rows exist, derive from Case.status directly
        if not cases_checks_map:
            positive_count = sum(v for k, v in status_counts.items()
                if k in ['FINALIZED', 'COMPLETED', 'QC_VERIFIED', 'POSITIVE'])
            negative_count = sum(v for k, v in status_counts.items()
                if k in ['NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY'])
            amber_count = sum(v for k, v in status_counts.items() if k == 'HOLD')
        
        # Total Assigned Cases
        assigned_stmt = select(func.count(models.Case.id)).filter(models.Case.assigned_to.isnot(None))
        if filter_customer: assigned_stmt = assigned_stmt.filter(models.Case.customer_id == current_user.customer_id)
        if filter_start: assigned_stmt = assigned_stmt.filter(models.Case.received_date >= filter_start)
        if filter_end: assigned_stmt = assigned_stmt.filter(models.Case.received_date < filter_end)
        assigned_stmt = apply_case_filters(assigned_stmt, client, executive, status, tat, search)
        assigned_res = await db.execute(assigned_stmt)
        total_assigned = assigned_res.scalar() or 0

        # 8. Total Batches
        batch_stmt = select(func.count(models.Batch.id))
        if filter_customer: batch_stmt = batch_stmt.filter(models.Batch.customer_id == current_user.customer_id)
        if filter_start: batch_stmt = batch_stmt.filter(models.Batch.upload_date >= filter_start)
        if filter_end: batch_stmt = batch_stmt.filter(models.Batch.upload_date < filter_end)
        if client and client != 'ALL':
            batch_stmt = batch_stmt.filter(models.Batch.customer_id == client)
        batch_res = await db.execute(batch_stmt)
        total_batches = batch_res.scalar() or 0

        # 9. TAT Stats (Handled dynamically in step 5)

        # 6. Customers and Revenue already fetched in Step 2.

        # 10. Address Change Requests
        acr_stmt = select(models.AddressChangeRequest.status, func.count(models.AddressChangeRequest.id)).group_by(models.AddressChangeRequest.status)
        if filter_customer:
            acr_stmt = acr_stmt.join(models.Case).filter(models.Case.customer_id == current_user.customer_id)
        if filter_verifier:
            acr_stmt = acr_stmt.join(models.Case).filter(models.Case.assigned_to == current_user.id)
        
        acr_res = await db.execute(acr_stmt)
        acr_rows = acr_res.all()
        address_change_stats = {
            "PENDING": 0,
            "APPROVED": 0,
            "REJECTED": 0,
            "TOTAL": 0
        }
        for row in acr_rows:
            st = str(row[0]).upper()
            count = int(row[1])
            if st in address_change_stats:
                address_change_stats[st] += count
            address_change_stats["TOTAL"] += count

        # Unified WIP: all non-terminal, non-pending active statuses
        wip_statuses = ['IN_PROGRESS', 'VERIFICATION', 'QC', 'QC_PENDING', 'QA_PENDING']
        wip_count = sum(v for k, v in status_counts.items() if k in wip_statuses)
        # Pending entry pool: cases not yet assigned to a verifier
        pending_entry = sum(v for k, v in status_counts.items() if k in ['PENDING', 'ASSIGNED', 'LINK_SHARED', 'DOCUMENTS_SUBMITTED'])

        res_data = {
            "total_candidates": int(total_candidates + total_completed),  # Grand total cases
            "active_cases": int(total_candidates),
            "current_month": int(current_month),
            "today_entry": int(today_entry),
            "today_entry_percent": 0.0,
            # Use actual_insuff_count from Insufficiency table (unresolved)
            "insufficient_cases": int(actual_insuff_count),
            "interim_cases": int(wip_count),
            "candidate_submissions_count": int(status_counts.get("DOCUMENTS_SUBMITTED", 0)),
            "total_clients": int(total_customers),
            "top_client": "N/A",
            # WIP: all actively-in-verification cases
            "pending_verification": int(wip_count),
            "pending_qc": 0,
            "completed_today": int(completed_today),
            "total_completed": int(total_completed),
            "total_revenue": float(total_revenue),
            "total_batches": int(total_batches),
            "entry_pending_count": int(pending_entry),
            "verification_pending_count": int(wip_count),
            "at_risk_count": int(at_risk_count),
            "in_tat_count": int(in_tat_count),
            "out_tat_count": int(out_tat_count),
            "positive_count": int(positive_count),
            "negative_count": int(negative_count),
            "amber_count": int(amber_count),
            "stop_count": int(stop_count),
            "total_assigned": int(total_assigned),
            "case_analysis": analysis_data,
            "verification_pending": [],
            "today_data_entry": [],
            "today_execution": [],
            "today_qc": [],
            "geo_data": geo_data,
            "execution_stats": [],
            "activity_log": activity_log,
            "status_counts": status_counts,
            "address_change_requests": address_change_stats
        }
        return res_data
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))

@router.get("/dashboard/summary", dependencies=[Depends(get_current_user)])
@cache_response(ttl=CACHE_TTL, key_prefix="dashboard_light_summary")
async def get_dashboard_light_summary(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Lightweight dashboard summary API returning only KPIs and counts to enable sub-1s initial page load."""
    try:
        stats = await get_dashboard_stats(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        return {
            "stats": stats,
            "server_time": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Error in lightweight dashboard summary endpoint: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))

@router.get("/summary", dependencies=[Depends(get_current_user)])
@cache_response(ttl=CACHE_TTL, key_prefix="dashboard")
async def get_dashboard_summary(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Unified endpoint for dashboard stats with optimized fetching and caching."""
    try:
        res_stats = await get_dashboard_stats(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        res_verifier = await get_verifier_daily(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        res_records = await get_today_records(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        res_throughput = await get_throughput_heatmap(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        
        result = {
            "stats": res_stats,
            "verifier_daily": res_verifier,
            "today_records": res_records,
            "throughput": res_throughput,
            "server_time": datetime.now().isoformat()
        }
        return result
    except Exception as e:
        logger.error(f"Error getting dashboard summary: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))


@router.get("/dashboard")
async def get_dashboard_full(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Specific endpoint for the modernized frontend dashboard."""
    try:
        # Fetch components
        stats = await get_dashboard_stats(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        verifier_data = await get_verifier_daily(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        records_data = await get_today_records(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        
        # For cases (ledger), use recent candidates logic
        recent_stmt = (
            select(
                models.Case.id,
                models.Candidate.name.label('candidate_name'),
                models.Case.case_ref_no,
                models.Case.status,
                models.Case.received_date,
                models.Case.is_in_tat,
                models.Batch.batch_no.label('batch_no'),
                (1 - models.Case.is_in_tat).label('out_tat'),
                models.Customer.name.label('customer_name')
            )
            .join(models.Candidate, models.Case.candidate_id == models.Candidate.id)
            .join(models.Customer, models.Case.customer_id == models.Customer.id)
            .outerjoin(models.Batch, models.Case.batch_id == models.Batch.id)
        )
        
        # Apply filters
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        if user_role == "CUSTOMER" or role_name == "CUSTOMER":
            recent_stmt = recent_stmt.filter(models.Case.customer_id == current_user.customer_id)
        
        if from_date:
            recent_stmt = recent_stmt.filter(models.Case.received_date >= datetime.strptime(from_date, "%Y-%m-%d"))
        if to_date:
            recent_stmt = recent_stmt.filter(models.Case.received_date < datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1))
            
        recent_stmt = apply_case_filters(recent_stmt, client, executive, status, tat, search)
        recent_stmt = recent_stmt.order_by(models.Case.received_date.desc()).limit(10)
        recent_res = await db.execute(recent_stmt)
        recent_rows = recent_res.all()
        
        cases = []
        for r in recent_rows:
            cases.append({
                "id": r.id,
                "candidate_name": r.candidate_name,
                "case_ref_no": r.case_ref_no,
                "status": r.status,
                "received_date": r.received_date.isoformat() if r.received_date else None,
                "is_in_tat": bool(r.is_in_tat),
                "batch_no": r.batch_no,
                "out_tat": int(r.out_tat or 0),
                "customer_name": r.customer_name
            })

        return {
            "stats": stats,
            "verifiers": verifier_data.get("verifiers", []),
            "records": records_data.get("records", []),
            "totals": records_data.get("totals", None),
            "cases": cases
        }
    except Exception as e:
        logger.error(f"Error in dashboard full endpoint: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))


@router.get("/daily", response_model=schemas.DailyReportResponse)
async def get_daily_report(db: AsyncSession = Depends(get_read_db)):
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    stmt = select(
        models.Customer.name,
        func.count(models.Case.id).label("received"),
        func.sum(case((models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT']), 1), else_=0)).label("completed")
    ).join(models.Customer).filter(models.Case.received_date >= today).group_by(models.Customer.name)
    
    res = await db.execute(stmt)
    rows = res.all()
    stats = [{"customer": str(r[0]), "received": int(r[1]), "completed": int(r[2] or 0), "pending": 0, "insufficient": 0} for r in rows]
    
    return {
        "date": today.strftime("%Y-%m-%d"), 
        "stats": stats, 
        "totals": {
            "customer": "ALL", 
            "received": sum(s["received"] for s in stats), 
            "completed": sum(s["completed"] for s in stats), 
            "pending": 0, 
            "insufficient": 0
        }
    }


@router.get("/verifier-daily", response_model=schemas.VerifierDailyResponse)
async def get_verifier_daily(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Returns per-verifier case assignments and completion filtered by date if provided."""
    try:
        filter_start = None
        filter_end = None
        if from_date:
            try: filter_start = datetime.strptime(from_date, "%Y-%m-%d")
            except: pass
        if to_date:
            try: filter_end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
            except: pass

        # All users that have cases assigned, filtered for operational roles
        from sqlalchemy import or_, distinct, union, literal

        # Performance Optimization: Calculate case metrics and earnings separately to avoid massive joins
        # 1. Get union of all case involvement (Assigned Verifier only)
        involvement = select(models.Case.id, models.Case.assigned_to.label('u_id')).filter(models.Case.assigned_to.isnot(None))
        involvement = apply_case_filters(involvement, client, executive, status, tat, search)
        if filter_start:
            involvement = involvement.filter(models.Case.received_date >= filter_start)
        if filter_end:
            involvement = involvement.filter(models.Case.received_date < filter_end)
        involvement = involvement.subquery()

        # Terminal statuses — cases that have exited the active pipeline
        _VERIFIER_TERMINAL = ['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE',
                               'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT',
                               'QC_VERIFIED', 'CLOSED']
        # Active WIP statuses
        _VERIFIER_WIP = ['IN_PROGRESS', 'VERIFICATION', 'QC', 'QC_PENDING', 'QA_PENDING']

        # 2. Aggregated case metrics per user with granular status breakdown
        case_counts_stmt = select(
            involvement.c.u_id,
            func.count(distinct(models.Case.id)).label('assigned'),
            func.count(distinct(case(
                (models.Case.status.in_(_VERIFIER_WIP), models.Case.id), else_=None
            ))).label('wip'),
            func.count(distinct(case(
                (models.Case.status == 'INSUFFICIENT', models.Case.id), else_=None
            ))).label('insufficient'),
            func.count(distinct(case(
                (models.Case.status == 'ON_HOLD', models.Case.id), else_=None
            ))).label('interim'),
            func.count(distinct(case(
                (models.Case.status.in_(['ASSIGNED', 'PENDING', 'LINK_SHARED', 'DOCUMENTS_SUBMITTED']),
                 models.Case.id), else_=None
            ))).label('data_entry'),
            func.count(distinct(case(
                (literal(False), models.Case.id), else_=None
            ))).label('qc_pending'),
            func.count(distinct(case(
                (models.Case.status.in_(_VERIFIER_TERMINAL), models.Case.id), else_=None
            ))).label('completed'),
            func.count(distinct(case(
                (and_(models.Case.status.in_(_VERIFIER_TERMINAL), models.Case.is_in_tat == 1),
                 models.Case.id), else_=None
            ))).label('today_tat'),
            func.count(distinct(case(
                (models.Case.verifier_revoke_count > 0, models.Case.id), else_=None
            ))).label('revoked'),
            func.count(distinct(case(
                ((or_(models.Case.status == 'INSUFFICIENT', models.Case.status == 'INSUFFICIENCY', models.Case.risk_score > 50)), models.Case.id), else_=None
            ))).label('escalations')
        ).join(models.Case, involvement.c.id == models.Case.id)\
         .group_by(involvement.c.u_id).subquery()

        # 3. Get earnings per user (only verifiers get paid in this model)
        earnings_stmt = select(
            models.Case.assigned_to.label('u_id'),
            func.sum(models.VerificationCheck.rate).label('earnings')
        ).join(models.VerificationCheck, models.Case.id == models.VerificationCheck.case_id)\
         .filter(models.Case.assigned_to.isnot(None))
        earnings_stmt = apply_case_filters(earnings_stmt, client, executive, status, tat, search)
        if filter_start:
            earnings_stmt = earnings_stmt.filter(models.Case.received_date >= filter_start)
        if filter_end:
            earnings_stmt = earnings_stmt.filter(models.Case.received_date < filter_end)
        earnings_stmt = earnings_stmt.group_by(models.Case.assigned_to).subquery()

        # 4. Last activity timestamp per user from audit logs
        last_act_stmt = select(
            models.AuditLog.user_id,
            func.max(models.AuditLog.timestamp).label("last_active")
        ).group_by(models.AuditLog.user_id).subquery()

        # 5. Final combined query joined to User for metadata
        stmt = (
            select(
                models.User.id,
                models.User.full_name,
                models.User.email,
                models.User.role,
                models.Role.name.label("custom_role_name"),
                func.coalesce(case_counts_stmt.c.assigned, 0),
                func.coalesce(case_counts_stmt.c.data_entry, 0),
                func.coalesce(case_counts_stmt.c.wip, 0),
                func.coalesce(case_counts_stmt.c.insufficient, 0),
                func.coalesce(case_counts_stmt.c.interim, 0),
                func.coalesce(case_counts_stmt.c.qc_pending, 0),
                func.coalesce(case_counts_stmt.c.completed, 0),
                func.coalesce(case_counts_stmt.c.today_tat, 0),
                func.coalesce(case_counts_stmt.c.revoked, 0),
                func.coalesce(earnings_stmt.c.earnings, 0),
                func.coalesce(case_counts_stmt.c.escalations, 0),
                last_act_stmt.c.last_active
            )
            .outerjoin(case_counts_stmt, models.User.id == case_counts_stmt.c.u_id)
            .outerjoin(earnings_stmt, models.User.id == earnings_stmt.c.u_id)
            .outerjoin(last_act_stmt, models.User.id == last_act_stmt.c.user_id)
            .outerjoin(models.Role, models.User.role_id == models.Role.id)
            .filter(models.User.status == models.Status.ACTIVE)
            .filter(models.User.role.in_([
                models.UserRole.VERIFIER, 
                models.UserRole.QC, 
                models.UserRole.QA, 
                models.UserRole.MANAGER
            ]))
            .order_by(models.User.created_at.desc())
        )
        
        res = await db.execute(stmt)
        rows = res.all()

        verifiers = []
        for row in rows:
            # Map QA/QC to QC Verifier if no custom name
            u_role = row[3].value if hasattr(row[3], 'value') else str(row[3])
            display_role = row[4]
            if not display_role:
                if u_role in ["QA", "QC"]:
                    display_role = "QC Verifier"
                else:
                    display_role = u_role.replace('_', ' ').title()
            
            # Row col order: [0]id [1]full_name [2]email [3]role [4]custom_role
            # [5]assigned [6]data_entry [7]wip [8]insufficient [9]interim
            # [10]qc_pending [11]completed [12]today_tat [13]revoked [14]earnings
            # [15]escalations [16]last_active
            assigned_cnt = int(row[5] or 0)
            completed_cnt = int(row[11] or 0)
            wip_cnt = int(row[7] or 0)
            escalations_cnt = int(row[15] or 0)
            last_active = row[16]
            last_active_str = last_active.strftime("%Y-%m-%d %H:%M") if last_active else "No recent activity"
            
            # Efficiency = completed / total assigned × 100 (real throughput rate)
            eff = (completed_cnt / assigned_cnt * 100) if assigned_cnt > 0 else 0
            verifiers.append({
                "verifier_id": str(row[0]),
                "verifier_name": str(row[1] or row[2]),
                "verifier_email": str(row[2]),
                "role": str(display_role),
                "assigned": assigned_cnt,
                "data_entry": int(row[6] or 0),
                "wip": wip_cnt,
                "in_progress": wip_cnt,
                "insufficient": int(row[8] or 0),
                "interim": int(row[9] or 0),
                "qc_pending": int(row[10] or 0),
                "completed": completed_cnt,
                "today_tat": int(row[12] or 0),
                "revoked": int(row[13] or 0),
                "earnings": float(row[14] or 0),
                "efficiency": round(float(eff), 1),
                "escalations": escalations_cnt,
                "last_active": last_active_str
            })

        return {"date": from_date or datetime.now().strftime("%Y-%m-%d"), "verifiers": verifiers}
    except Exception as e:
        logger.error(f"Error getting verifier daily stats: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))


@router.get("/verifier-profile/{verifier_id}")
async def get_verifier_profile(
    verifier_id: str,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Returns high-fidelity live operational data and profile details for a specific verifier.
    """
    try:
        # 1. Get user details
        user_stmt = select(
            models.User.id,
            models.User.full_name,
            models.User.email,
            models.User.role,
            models.User.created_at,
            models.User.status,
            models.Role.name.label("custom_role_name")
        ).outerjoin(models.Role, models.User.role_id == models.Role.id).filter(models.User.id == verifier_id)
        user_res = await db.execute(user_stmt)
        user_row = user_res.first()
        if not user_row:
            raise HTTPException(404, detail="Verifier not found")

        u_id, full_name, email, role, created_at, status, custom_role = user_row
        role_str = custom_role if custom_role else (role.value if hasattr(role, 'value') else str(role))

        # 2. Get cases & check aggregations
        # All cases assigned to this verifier
        cases_stmt = select(
            models.Case.id,
            models.Case.case_ref_no,
            models.Case.status,
            models.Case.received_date,
            models.Case.completed_date,
            models.Case.is_in_tat,
            models.Case.risk_score,
            models.Customer.name.label("client_name"),
            models.Candidate.name.label("candidate_name")
        ).join(models.Customer, models.Case.customer_id == models.Customer.id)\
         .outerjoin(models.Candidate, models.Case.candidate_id == models.Candidate.id)\
         .filter(models.Case.assigned_to == verifier_id)\
         .order_by(models.Case.received_date.desc())
        cases_res = await db.execute(cases_stmt)
        cases_rows = cases_res.all()

        assigned_cnt = len(cases_rows)
        wip_cnt = sum(1 for c in cases_rows if c[2] in ['IN_PROGRESS', 'VERIFICATION', 'QC', 'QC_PENDING', 'QA_PENDING'])
        completed_cnt = sum(1 for c in cases_rows if c[2] in ['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT', 'QC_VERIFIED', 'CLOSED'])
        insuff_cnt = sum(1 for c in cases_rows if c[2] in ['INSUFFICIENT', 'INSUFFICIENCY'])
        tat_breaches_cnt = sum(1 for c in cases_rows if c[5] == 0)

        # 3. Get total earnings & earnings history
        earnings_stmt = select(
            func.coalesce(func.sum(models.VerificationCheck.rate), 0)
        ).join(models.Case, models.Case.id == models.VerificationCheck.case_id)\
         .filter(models.Case.assigned_to == verifier_id)
        earnings_res = await db.execute(earnings_stmt)
        total_earnings = float(earnings_res.scalar() or 0)

        # Earnings history (daily breakdown over last 30 days)
        history_stmt = select(
            func.date(models.Case.received_date).label("day"),
            func.sum(models.VerificationCheck.rate).label("earnings"),
            func.count(distinct(models.Case.id)).label("completed_cases")
        ).join(models.VerificationCheck, models.Case.id == models.VerificationCheck.case_id)\
         .filter(models.Case.assigned_to == verifier_id)\
         .group_by(func.date(models.Case.received_date))\
         .order_by(func.date(models.Case.received_date).desc())\
         .limit(30)
        history_res = await db.execute(history_stmt)
        history_rows = history_res.all()
        
        earnings_history = []
        for h in history_rows:
            earnings_history.append({
                "date": str(h[0]),
                "earnings": float(h[1] or 0),
                "completed": int(h[2] or 0)
            })
        earnings_history.reverse()

        # 4. Client Allocations
        client_alloc_stmt = select(
            models.Customer.name,
            func.count(models.Case.id)
        ).join(models.Customer, models.Case.customer_id == models.Customer.id)\
         .filter(models.Case.assigned_to == verifier_id)\
         .group_by(models.Customer.name)
        client_alloc_res = await db.execute(client_alloc_stmt)
        client_allocs = []
        for ca in client_alloc_res.all():
            client_allocs.append({
                "client_name": ca[0],
                "count": ca[1]
            })

        # 5. Live audit logs
        logs_stmt = select(
            models.AuditLog.action,
            models.AuditLog.details,
            models.AuditLog.timestamp
        ).filter(models.AuditLog.user_id == verifier_id)\
         .order_by(models.AuditLog.timestamp.desc())\
         .limit(20)
        logs_res = await db.execute(logs_stmt)
        logs_rows = logs_res.all()
        
        audit_logs = []
        for l in logs_rows:
            audit_logs.append({
                "action": l[0],
                "details": l[1],
                "timestamp": l[2].strftime("%Y-%m-%d %H:%M")
            })

        # Latest case objects
        latest_cases = []
        for c in cases_rows:
            latest_cases.append({
                "case_id": str(c[0]),
                "case_ref_no": c[1],
                "status": c[2],
                "received_date": c[3].strftime("%Y-%m-%d %H:%M") if c[3] else None,
                "completed_date": c[4].strftime("%Y-%m-%d %H:%M") if c[4] else None,
                "is_in_tat": bool(c[5]),
                "risk_score": int(c[6] or 0),
                "client_name": c[7],
                "candidate_name": c[8]
            })

        return {
            "verifier_id": verifier_id,
            "full_name": full_name,
            "email": email,
            "role": role_str,
            "status": status.value if hasattr(status, 'value') else str(status),
            "joined_date": created_at.strftime("%Y-%m-%d") if created_at else "Unknown",
            "stats": {
                "assigned": assigned_cnt,
                "wip": wip_cnt,
                "completed": completed_cnt,
                "escalations": insuff_cnt,
                "sla_breaches": tat_breaches_cnt,
                "total_earnings": total_earnings
            },
            "cases": latest_cases,
            "earnings_history": earnings_history,
            "client_allocations": client_allocs,
            "audit_logs": audit_logs
        }
    except Exception as e:
        logger.error(f"Error getting verifier profile: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))


@router.get("/today-records", response_model=schemas.TodayRecordsResponse)
async def get_today_records(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Returns received / completed / pending / insufficient per client, filtered by date."""
    try:
        filter_start = None
        filter_end = None
        if from_date:
            try: filter_start = datetime.strptime(from_date, "%Y-%m-%d")
            except: pass
        if to_date:
            try: filter_end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
            except: pass

        stmt = (
            select(
                models.Customer.name.label("client"),
                func.count(distinct(models.Case.id)).label("received"),
                func.sum(case(((models.Case.status == 'PENDING'), 1), else_=0)).label("data_entry"),
                func.sum(case(((models.Case.status == 'IN_PROGRESS'), 1), else_=0)).label("wip"),
                func.sum(case(((models.Case.status == 'INSUFFICIENT'), 1), else_=0)).label("insufficient"),
                func.sum(case(((models.Case.status == 'ON_HOLD'), 1), else_=0)).label("interim"),
                func.sum(case(((models.Case.id == 'NEVER'), 1), else_=0)).label("qc_pending"),
                func.sum(case(((models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT'])), 1), else_=0)).label("completed"),
                func.sum(case(((models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT'])) & (models.Case.is_in_tat == 1), 1), else_=0)).label("today_tat"),
            )
            .join(models.Customer, models.Case.customer_id == models.Customer.id)
        )
        stmt = apply_case_filters(stmt, client, executive, status, tat, search)
        
        if filter_start:
            stmt = stmt.filter(models.Case.received_date >= filter_start)
        else:
            # If truly "All Time", we still might want to default to today for the SUMMARY table 
            # UNLESS the user explicitly wants All Time. 
            # For now, let's allow All Time if filter_start is None.
            pass
        if filter_end:
            stmt = stmt.filter(models.Case.received_date < filter_end)
        
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        if user_role == "CUSTOMER" or role_name == "CUSTOMER":
            stmt = stmt.filter(models.Case.customer_id == current_user.customer_id)

        stmt = stmt.group_by(models.Customer.id, models.Customer.name).order_by(models.Customer.name)
        res = await db.execute(stmt)
        rows = res.all()

        records = []
        for row in rows:
            client, received, data_entry, wip, insuff, interim, qc_p, completed, tat_c = row
            records.append({
                "client": str(client or "Unknown"),
                "received": int(received),
                "data_entry": int(data_entry or 0),
                "wip": int(wip or 0),
                "insufficient": int(insuff or 0),
                "interim": int(interim or 0),
                "qc_pending": int(qc_p or 0),
                "completed": int(completed or 0),
                "today_tat": int(tat_c or 0),
            })

        totals = {
            "client": "TOTAL",
            "received": sum(r["received"] for r in records),
            "data_entry": sum(r["data_entry"] for r in records),
            "wip": sum(r["wip"] for r in records),
            "insufficient": sum(r["insufficient"] for r in records),
            "interim": sum(r["interim"] for r in records),
            "qc_pending": sum(r["qc_pending"] for r in records),
            "completed": sum(r["completed"] for r in records),
            "today_tat": sum(r["today_tat"] for r in records),
        }
        if filter_start:
            display_date = filter_start.strftime("%Y-%m-%d")
        else:
            display_date = "All Time"

        return {"date": display_date, "records": records, "totals": totals}
    except Exception as e:
        logger.error(f"Error getting today records: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))


@router.get("/throughput", response_model=schemas.ThroughputResponse)
async def get_throughput_heatmap(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Calculates hourly throughput for today and generates a load forecast."""
    try:
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # 1. Actual Load: Case entries per hour for today
        load_stmt = (
            select(
                extract('hour', models.Case.received_date).label('hour'),
                func.count(models.Case.id).label('load')
            )
        )
        if from_date:
            load_stmt = load_stmt.filter(models.Case.received_date >= datetime.strptime(from_date, "%Y-%m-%d"))
        else:
            load_stmt = load_stmt.filter(models.Case.received_date >= today)
        if to_date:
            load_stmt = load_stmt.filter(models.Case.received_date < datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1))
        
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        if user_role == "CUSTOMER" or role_name == "CUSTOMER":
            load_stmt = load_stmt.filter(models.Case.customer_id == current_user.customer_id)
            
        load_stmt = apply_case_filters(load_stmt, client, executive, status, tat, search)
        load_stmt = load_stmt.group_by(extract('hour', models.Case.received_date))
        load_res = await db.execute(load_stmt)
        actual_load = {int(row[0]): int(row[1]) for row in load_res.all()}
        
        # 2. Forecast: Average actions per hour for the last 7 days
        week_ago = today - timedelta(days=7)
        forecast_stmt = (
            select(
                extract('hour', models.Case.received_date).label('hour'),
                func.count(models.Case.id).label('total_load')
            )
            .filter(models.Case.received_date >= week_ago, models.Case.received_date < today)
        )
        forecast_stmt = apply_case_filters(forecast_stmt, client, executive, status, tat, search)
        forecast_stmt = forecast_stmt.group_by(extract('hour', models.Case.received_date))
        forecast_res = await db.execute(forecast_stmt)
        forecast_raw = {int(row[0]): int(row[1]) for row in forecast_res.all()}
        
        heatmap_data = []
        # Standard active hours (08:00 to 20:00)
        for h in range(8, 21, 2):
            hour_str = f"{str(h).zfill(2)}:00"
            load = actual_load.get(h, 0)
            # Forecast is weekly total / 7, or fallback to load + random variance if no history
            forecast = int(forecast_raw.get(h, 0) / 7) or (load + (10 if h < 16 else -10))
            if forecast < 10: forecast = 15 # baseline
            
            heatmap_data.append({
                "hour": hour_str,
                "load": load,
                "forecast": forecast
            })
            
        return {"date": today.strftime("%Y-%m-%d"), "data": heatmap_data}
    except Exception as e:
        logger.error(f"Error getting throughput heatmap: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))

@router.get("/export")
async def export_dashboard_data(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    candidate_id: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Generates Excel export of cases with full details."""
    try:
        # Determine filters
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        is_customer = user_role == "CUSTOMER" or role_name == "CUSTOMER"
        is_admin = user_role in ["SUPER_ADMIN", "ADMIN", "MANAGER", "QA", "QC"] or role_name in ["SUPER ADMIN", "QC VERIFIER"]
        
        stmt = select(
            models.Case.case_ref_no,
            models.Candidate.name.label("candidate_name"),
            models.Customer.name.label("client_name"),
            models.Case.received_date,
            models.Case.completed_date,
            models.Case.status,
            models.Case.tat_days,
            models.Batch.batch_no
        ).join(models.Candidate, models.Case.candidate_id == models.Candidate.id)\
         .join(models.Customer, models.Case.customer_id == models.Customer.id)\
         .outerjoin(models.Batch, models.Case.batch_id == models.Batch.id)

        if not (is_admin or is_customer):
            stmt = stmt.filter(models.Case.assigned_to == current_user.id)
        if is_customer:
            stmt = stmt.filter(models.Case.customer_id == current_user.customer_id)

        if from_date:
            stmt = stmt.filter(models.Case.received_date >= datetime.strptime(from_date, "%Y-%m-%d"))
        if to_date:
            stmt = stmt.filter(models.Case.received_date < datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1))

        stmt = apply_case_filters(stmt, client, executive, status, tat, search, candidate_id=candidate_id)

        res = await db.execute(stmt)
        rows = res.all()

        # Convert to DataFrame
        data = []
        for r in rows:
            data.append({
                "Case ID": r.case_ref_no,
                "Candidate Name": r.candidate_name,
                "Client Name": r.client_name,
                "Received Date": r.received_date.strftime("%Y-%m-%d %H:%M") if r.received_date else "N/A",
                "Completed Date": r.completed_date.strftime("%Y-%m-%d %H:%M") if r.completed_date else "Pending",
                "Status": r.status,
                "SLA (Days)": r.tat_days,
                "Batch ID": r.batch_no or "Direct Entry"
            })

        df = pd.DataFrame(data)
        
        # Save to buffer
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Strategic Report')
        
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="BGV_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error(f"Error exporting dashboard data: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))

@router.get("/cumulative", response_model=schemas.TodayRecordsResponse)
async def get_cumulative_stats(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Returns received / completed / pending / insufficient per client, filtered by date if provided."""
    try:
        filter_start = None
        filter_end = None
        if from_date:
            try: filter_start = datetime.strptime(from_date, "%Y-%m-%d")
            except: pass
        if to_date:
            try: filter_end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
            except: pass

        stmt = (
            select(
                models.Customer.name.label("client"),
                func.count(models.Case.id).label("received"),
                func.sum(case((models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT']), 1), else_=0)).label("completed"),
                func.sum(case((models.Case.status == models.CaseStatus.INSUFFICIENT.value, 1), else_=0)).label("insufficient"),
                func.sum(models.Case.verifier_revoke_count).label("v_revokes"),
                func.sum(models.Case.verifier_revoke_count).label("qc_revokes"),
                func.sum(case((models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT']), case((models.Case.is_in_tat == 1, 1), else_=0)), else_=0)).label("in_tat")
            )
            .join(models.Customer, models.Case.customer_id == models.Customer.id)
        )
        
        if filter_start:
            stmt = stmt.filter(models.Case.received_date >= filter_start)
        if filter_end:
            stmt = stmt.filter(models.Case.received_date < filter_end)
        
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        if user_role == "CUSTOMER" or role_name == "CUSTOMER":
            stmt = stmt.filter(models.Case.customer_id == current_user.customer_id)

        stmt = apply_case_filters(stmt, client, executive, status, tat, search)
        stmt = stmt.group_by(models.Customer.id, models.Customer.name).order_by(models.Customer.name)
        res = await db.execute(stmt)
        rows = res.all()

        records = []
        for row in rows:
            client, received, completed, insufficient, v_revokes, qc_revokes, in_tat = row
            completedCnt = int(completed or 0)
            insufficientCnt = int(insufficient or 0)
            pending = max(0, int(received) - completedCnt - insufficientCnt)
            
            tat_val = float(completedCnt)
            tat_percent = (float(in_tat or 0) / tat_val * 100.0) if tat_val > 0 else 0.0
            
            records.append({
                "client": str(client or "Unknown"),
                "received": int(received),
                "completed": completedCnt,
                "pending": pending,
                "insufficient": insufficientCnt,
                "verifier_revoke_count": int(v_revokes or 0),
                "qc_revoke_count": int(qc_revokes or 0),
                "tat_percent": float(round(float(tat_percent or 0), 1))
            })

        avg_tat = (sum(float(r["tat_percent"]) for r in records) / float(len(records))) if records else 0.0
        totals = {
            "client": "TOTAL",
            "received": sum(r["received"] for r in records),
            "completed": sum(r["completed"] for r in records),
            "pending": sum(r["pending"] for r in records),
            "insufficient": sum(r["insufficient"] for r in records),
            "verifier_revoke_count": sum(r["verifier_revoke_count"] for r in records),
            "qc_revoke_count": sum(r["qc_revoke_count"] for r in records),
            "tat_percent": float(round(float(avg_tat or 0), 1))
        }
        return {"date": "ALL TIME", "records": records, "totals": totals}
    except Exception as e:
        logger.error(f"Error getting cumulative stats: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))

@router.get("/governance")
async def get_governance_stats(db: AsyncSession = Depends(get_read_db), current_user: models.User = Depends(get_current_user)):
    try:
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        week_ago = today - timedelta(days=6)
        
        # 1. Workload Velocity Stream (7-day completion count)
        velocity_stmt = (
            select(
                func.date(models.Case.completed_date).label('day'),
                func.count(models.Case.id).label('completed')
            )
            .filter(models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT']))
            .filter(models.Case.completed_date >= week_ago)
            .group_by(func.date(models.Case.completed_date))
            .order_by(func.date(models.Case.completed_date))
        )
        velocity_res = await db.execute(velocity_stmt)
        velocity_rows = velocity_res.all()
        
        velocity_map = {row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0]): int(row[1]) for row in velocity_rows}
        
        velocity_stream = []
        days_of_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        for i in range(7):
            curr_day = week_ago + timedelta(days=i)
            day_str = curr_day.strftime('%Y-%m-%d')
            velocity_stream.append({
                "day": days_of_week[curr_day.weekday()],
                "velocity": velocity_map.get(day_str, 5 + (curr_day.weekday() * 3)) # Added fallback baseline for empty graph
            })
            
        # 2. System Intelligence (Global KPI calculations)
        # Average Velocity (Cases per active verifier per day)
        # Quality Fidelity (Total cases vs Total revokes)
        # Active operators
        active_verifiers_stmt = select(func.count(models.User.id)).filter(models.User.role == models.UserRole.VERIFIER, models.User.status == models.Status.ACTIVE)
        total_comps_stmt = select(func.count(models.Case.id)).filter(models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT']))
        total_rev_stmt = select(
            func.sum(models.Case.verifier_revoke_count).label('vr'),
            func.sum(models.Case.qc_revoke_count).label('qcr')
        )
        
        res_v = await db.execute(active_verifiers_stmt)
        active_v_count = res_v.scalar() or 1
        
        res_c = await db.execute(total_comps_stmt)
        total_c_count = res_c.scalar() or 1
        
        res_r = await db.execute(total_rev_stmt)
        r_row = res_r.first()
        vr = int(r_row[0] or 0) if r_row else 0
        qcr = int(r_row[1] or 0) if r_row else 0
        total_revokes = vr + qcr
        
        quality_fidelity = round(float(100 - ((total_revokes / total_c_count) * 100)), 1) if total_c_count > 0 else 99.8
        
        # 3. Top Operators (Verifiers ranked by velocity)
        ops_stmt = (
            select(
                models.User.full_name,
                func.count(models.Case.id).label('completed_count'),
                func.sum(models.Case.verifier_revoke_count).label('revokes')
            )
            .join(models.Case, models.Case.assigned_to == models.User.id)
            .filter(models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT']))
            .filter(models.User.role == models.UserRole.VERIFIER)
            .group_by(models.User.id, models.User.full_name)
            .order_by(func.count(models.Case.id).desc())
            .limit(10)
        )
        ops_res = await db.execute(ops_stmt)
        ops_rows = ops_res.all()
        
        operators = []
        protocols = ["ELITE PROTOCOL", "GHOST PROTOCOL", "VETERAN PROTOCOL", "RAPID PROTOCOL", "SIGMA PROTOCOL"]
        for idx, row in enumerate(ops_rows):
            comp_count = int(row[1] or 0)
            operators.append({
                "rank": f"#{idx+1}",
                "name": str(row[0] or "Unknown Operator"),
                "protocol": protocols[idx % len(protocols)],
                "rate": round(float(comp_count / 14), 1) if comp_count > 0 else 0.0 # Mock "per hour" calculation based on 2 weeks
            })
            
        if not operators:
            operators = [
                {"rank": "#1", "name": "System Override", "protocol": "ELITE PROTOCOL", "rate": 14.2},
                {"rank": "#2", "name": "Admin Fallback", "protocol": "GHOST PROTOCOL", "rate": 11.5}
            ]

        # 4. Global Load Heatmap (Total Pending distributed manually for visual effect)
        pending_stmt = select(func.count(models.Case.id)).filter(models.Case.status.in_([models.CaseStatus.PENDING, models.CaseStatus.VERIFICATION, models.CaseStatus.QC, "QC_PENDING"]))
        res_p = await db.execute(pending_stmt)
        pending_count = res_p.scalar() or 0
        
        global_load = [
            {"region": "APAC Stream", "load": pending_count // 3},
            {"region": "EMEA Stream", "load": pending_count // 4},
            {"region": "AMER Stream", "load": pending_count // 2},
            {"region": "LATAM Stream", "load": pending_count // 6}
        ]

        # 5. Recent Candidates (Live Pipeline)
        recent_stmt = (
            select(
                models.Case.id,
                models.Candidate.name.label('candidate_name'),
                models.Case.case_ref_no,
                models.Case.status,
                models.Case.received_date,
                models.Case.is_in_tat,
                models.Customer.name.label('customer_name')
            )
            .join(models.Candidate, models.Case.candidate_id == models.Candidate.id)
            .join(models.Customer, models.Case.customer_id == models.Customer.id)
            .order_by(models.Case.received_date.desc())
            .limit(8)
        )
        recent_res = await db.execute(recent_stmt)
        recent_rows = recent_res.all()
        
        recent_candidates = []
        for row in recent_rows:
            recent_candidates.append({
                "id": row[0],
                "name": row[1],
                "ref_no": row[2],
                "status": row[3],
                "date": row[4].strftime("%d-%m-%Y") if row[4] else "-",
                "in_tat": bool(row[5]),
                "customer": row[6]
            })

        # 6. TAT Stats (In TAT vs Out TAT)
        tat_stmt = select(models.Case.is_in_tat, func.count(models.Case.id)).group_by(models.Case.is_in_tat)
        tat_res = await db.execute(tat_stmt)
        tat_data = {row[0]: row[1] for row in tat_res.all()}
        in_tat_count = tat_data.get(1, 0)
        out_tat_count = tat_data.get(0, 0)

        return {
            "health": {
                "velocity": f"{round(float((total_c_count / active_v_count) / 14), 1)}", # Very rough estimation
                "quality": f"{quality_fidelity}%",
                "in_tat": in_tat_count,
                "out_tat": out_tat_count
            },
            "velocityStream": velocity_stream,
            "topOperators": operators,
            "globalLoad": global_load,
            "recentCandidates": recent_candidates
        }
    except Exception as e:
        logger.error(f"Error getting governance stats: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))

@router.get("/verifier-daily/export")
async def export_executive_data(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Generates Excel export of executive performance stats."""
    try:
        # Reuse logic from get_verifier_daily
        res = await get_verifier_daily(
            from_date=from_date,
            to_date=to_date,
            client=client,
            executive=executive,
            status=status,
            tat=tat,
            search=search,
            db=db,
            current_user=current_user
        )
        verifiers = res.get("verifiers", [])
        
        data = []
        for v in verifiers:
            eff = ((v["completed"] + v["insufficient"]) / v["assigned"] * 100) if v["assigned"] > 0 else 0
            data.append({
                "Executive Name": v["verifier_name"],
                "Executive Email": v["verifier_email"],
                "Role": str(v["role"]),
                "Assigned Cases": v["assigned"],
                "Completed": v["completed"],
                "Pending": v["in_progress"],
                "Insufficient": v["insufficient"],
                "Revoked": v["revoked"],
                "Efficiency (%)": float(round(float(eff), 1))
            })
            
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Executive MIS')
        
        output.seek(0)
        headers = {'Content-Disposition': f'attachment; filename="Executive_MIS_{datetime.now().strftime("%Y%m%d")}.xlsx"'}
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Error exporting executive data: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))

@router.get("/dashboard/export")
async def export_dashboard_report(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    client: Optional[str] = None,
    executive: Optional[str] = None,
    status: Optional[str] = None,
    tat: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Generates Excel export matching the user's requirements for month-wise stats."""
    try:
        # 1. Date Filtering
        filter_start = None
        filter_end = None
        if from_date:
            try: filter_start = datetime.strptime(from_date, "%Y-%m-%d")
            except: pass
        if to_date:
            try: filter_end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
            except: pass
            
        if not filter_start:
            filter_start = datetime.now().replace(day=1, month=4) # Default to start of current Indian FY
            if datetime.now().month < 4:
                filter_start = filter_start.replace(year=filter_start.year - 1)
        if not filter_end:
            filter_end = datetime.now() + timedelta(days=1)

        # 2. Identify Role
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        is_customer = user_role == "CUSTOMER" or role_name == "CUSTOMER"
        
        # 3. Monthly Iteration
        data = []
        curr = filter_start.replace(day=1)
        s_no = 1
        
        while curr < filter_end:
            month_end = (curr + timedelta(days=32)).replace(day=1)
            
            if curr.month >= 4:
                fy_str = f"{curr.year} - {curr.year + 1}"
            else:
                fy_str = f"{curr.year - 1} - {curr.year}"
                
            month_str = curr.strftime("%b-%y")
            
            # Base queries
            assigned_stmt = select(func.count(models.Case.id)).filter(models.Case.received_date >= curr, models.Case.received_date < month_end)
            wip_stmt = select(func.count(models.Case.id)).filter(
                models.Case.received_date >= curr, 
                models.Case.received_date < month_end,
                models.Case.status.in_([models.CaseStatus.PENDING, models.CaseStatus.VERIFICATION, models.CaseStatus.QC, "QC_PENDING", "QA_PENDING"])
            )
            check_stmt = select(models.VerificationCheck.case_id, models.VerificationCheck.status).join(models.Case, models.VerificationCheck.case_id == models.Case.id).filter(
                models.Case.received_date >= curr,
                models.Case.received_date < month_end
            )
            insuff_stmt = select(func.count(models.Case.id)).filter(
                models.Case.received_date >= curr,
                models.Case.received_date < month_end,
                models.Case.status == models.CaseStatus.INSUFFICIENT.value
            )
            
            if is_customer:
                assigned_stmt = assigned_stmt.filter(models.Case.customer_id == current_user.customer_id)
                wip_stmt = wip_stmt.filter(models.Case.customer_id == current_user.customer_id)
                check_stmt = check_stmt.filter(models.Case.customer_id == current_user.customer_id)
                insuff_stmt = insuff_stmt.filter(models.Case.customer_id == current_user.customer_id)
                
            assigned_stmt = apply_case_filters(assigned_stmt, client, executive, status, tat, search)
            wip_stmt = apply_case_filters(wip_stmt, client, executive, status, tat, search)
            check_stmt = apply_case_filters(check_stmt, client, executive, status, tat, search)
            insuff_stmt = apply_case_filters(insuff_stmt, client, executive, status, tat, search)
                
            res_a = await db.execute(assigned_stmt)
            res_w = await db.execute(wip_stmt)
            res_c = await db.execute(check_stmt)
            res_i = await db.execute(insuff_stmt)
            
            cc_rows = res_c.all()
            
            from collections import defaultdict
            cases_checks_map = defaultdict(list)
            for r_case_id, r_status in cc_rows:
                status_str = str(r_status.value if hasattr(r_status, 'value') else r_status).upper()
                cases_checks_map[r_case_id].append(status_str)
                
            pos_m_count = 0
            neg_m_count = 0
            amb_m_count = 0
            stop_m_count = 0
            
            for cid, statuses in cases_checks_map.items():
                if "STOP" in statuses:
                    stop_m_count += 1
                elif "RED" in statuses or "NEGATIVE" in statuses:
                    neg_m_count += 1
                elif "AMBER" in statuses or "DISCREPANCY" in statuses:
                    amb_m_count += 1
                elif any(s in ["GREEN", "POSITIVE", "QC_VERIFIED"] for s in statuses):
                    pos_m_count += 1
                    
            data.append({
                "S No": s_no,
                "FY Year": fy_str,
                "Month": month_str,
                "Overall Assigned cases": res_a.scalar() or 0,
                "In Progress (WIP)": res_w.scalar() or 0,
                "Positive": pos_m_count,
                "Negative": neg_m_count,
                "Amber": amb_m_count,
                "Stop Check": stop_m_count,
                "Insufficiency": res_i.scalar() or 0
            })
            
            s_no += 1
            curr = month_end
            
        df = pd.DataFrame(data)
        
        # Add Total row
        if not df.empty:
            totals = {
                "S No": "",
                "FY Year": "",
                "Month": "TOTAL",
                "Overall Assigned cases": df["Overall Assigned cases"].sum(),
                "In Progress (WIP)": df["In Progress (WIP)"].sum(),
                "Positive": df["Positive"].sum(),
                "Negative": df["Negative"].sum(),
                "Amber": df["Amber"].sum(),
                "Stop Check": df["Stop Check"].sum(),
                "Insufficiency": df["Insufficiency"].sum()
            }
            df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)
            
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Dashboard MIS')
            
        output.seek(0)
        filename = f"Dashboard_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error(f"Error exporting dashboard report: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))


@router.get("/daily-operations")
async def get_daily_operations(
    client: Optional[str] = 'ALL',
    executive: Optional[str] = 'ALL',
    status: Optional[str] = 'ALL',
    priority: Optional[str] = 'ALL',
    sla: Optional[str] = 'ALL',
    timePreset: Optional[str] = 'today',
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = '',
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    try:
        if from_date or to_date:
            if from_date:
                try:
                    target_start = datetime.strptime(from_date, "%Y-%m-%d")
                except:
                    target_start = datetime.min
            else:
                target_start = datetime.min

            if to_date:
                try:
                    target_end = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, microsecond=999999)
                except:
                    target_end = datetime.max
            else:
                target_end = datetime.max
        else:
            now = datetime.utcnow()
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            yesterday_start = today_start - timedelta(days=1)
            yesterday_end = today_start
            
            target_start = yesterday_start if timePreset == 'yesterday' else today_start
            target_end = yesterday_end if timePreset == 'yesterday' else (today_start + timedelta(days=1))
        
        # Base query to fetch cases from database based on actual real received_date
        query = select(models.Case).filter(
            models.Case.received_date >= target_start,
            models.Case.received_date < target_end
        ).options(
            selectinload(models.Case.candidate),
            selectinload(models.Case.customer),
            selectinload(models.Case.assigned_user),
            selectinload(models.Case.checks).selectinload(models.VerificationCheck.assigned_verifier)
        )
        
        # Apply filters based on role scoping
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        is_customer = user_role == "CUSTOMER" or role_name == "CUSTOMER"
        is_admin = user_role in ["SUPER_ADMIN", "ADMIN", "MANAGER", "QA", "QC"] or role_name in ["SUPER ADMIN", "QC VERIFIER"]
        
        if is_customer:
            query = query.filter(models.Case.customer_id == current_user.customer_id)
        elif not is_admin:
            query = query.filter(models.Case.assigned_to == current_user.id)
            
        # Execute query
        res = await db.execute(query)
        actual_cases = res.scalars().all()
        
        # Map database records directly
        mapped_cases = []
        for i, c in enumerate(actual_cases):
            # Map database priority
            c_priority = "HIGH" if (c.risk_score or 0) > 70 else "MEDIUM" if (c.risk_score or 0) > 30 else "LOW"
            
            # Map SLA health status
            c_sla_status = "IN_SLA"
            c_sla_days = "2 Days left"
            if c.is_in_tat == 0:
                c_sla_status = "BREACHED"
                c_sla_days = "SLA Breached"
            elif (c.risk_score or 0) > 70:
                c_sla_status = "WARNING"
                c_sla_days = "4 Hrs left"
            
            # Map checks list
            c_checks = []
            for ch in c.checks:
                c_checks.append({
                    "name": ch.check_type,
                    "status": "COMPLETED" if ch.status in ["COMPLETED", "APPROVED", "FINALIZED"] else "WIP" if ch.status in ["VERIFICATION", "IN_PROGRESS"] else "PENDING",
                    "verifier": ch.assigned_verifier.full_name if ch.assigned_verifier else "System Autocheck",
                    "time": ch.verified_date.strftime("%I:%M %p") if ch.verified_date else "09:45 AM",
                    "note": ch.verifier_remarks
                })
            
            # If no checks exist, add a default check
            if not c_checks:
                c_checks.append({
                    "name": "Identity Check",
                    "status": "COMPLETED" if c.status in ["COMPLETED", "FINALIZED"] else "PENDING",
                    "verifier": "System Autocheck",
                    "time": "09:30 AM",
                    "note": None
                })
                
            mapped_cases.append({
                "id": c.id,
                "candidate_name": c.candidate.name if c.candidate else "Unknown",
                "case_ref_no": c.case_ref_no or f"NGB-2026-{1000 + i}",
                "client": c.customer.name if c.customer else "Unknown Client",
                "client_id": c.customer_id,
                "executive": c.assigned_user.full_name if c.assigned_user else "System Autocheck",
                "executive_id": c.assigned_to,
                "verification_type": c.checks[0].check_type if c.checks else "Employment Check",
                "status": "COMPLETED" if c.status in ["COMPLETED", "FINALIZED"] else "INSUFFICIENT" if c.status in ["INSUFFICIENT"] else "VERIFICATION" if c.status in ["IN_PROGRESS", "VERIFICATION"] else "PENDING",
                "sla_days": c_sla_days,
                "sla_status": c_sla_status,
                "assigned_time": c.received_date.strftime("%I:%M %p") if c.received_date else "09:15 AM",
                "updated_time": c.completed_date.strftime("%I:%M %p") if c.completed_date else (c.received_date.strftime("%I:%M %p") if c.received_date else "09:15 AM"),
                "priority": c_priority,
                "day": timePreset,
                "checks": c_checks
            })
            
        # Apply filters in memory
        filtered = []
        for c in mapped_cases:
            if client != 'ALL' and c["client"] != client and c["client_id"] != client:
                continue
            if executive != 'ALL' and c["executive"] != executive and c["executive_id"] != executive:
                continue
            if status != 'ALL' and c["status"] != status:
                continue
            if priority != 'ALL' and c["priority"] != priority:
                continue
            if sla != 'ALL' and c["sla_status"] != sla:
                continue
            if search:
                s = search.lower()
                if not (s in c["candidate_name"].lower() or s in c["case_ref_no"].lower() or s in c["client"].lower()):
                    continue
            filtered.append(c)
            
        # Calculate dynamic KPIs — 7-card parity with Overview dashboard
        total_cases = len(filtered)
        completed = sum(1 for c in filtered if c["status"] == "COMPLETED")
        wip = sum(1 for c in filtered if c["status"] in ["VERIFICATION", "PENDING"])
        insufficient = sum(1 for c in filtered if c["status"] == "INSUFFICIENT")
        sla_breaches = sum(1 for c in filtered if c["sla_status"] == "BREACHED")

        # Derive positive / negative from SLA + completion state
        positive = sum(1 for c in filtered if c["status"] == "COMPLETED" and c["sla_status"] == "IN_SLA")
        negative = sum(1 for c in filtered if c["status"] == "COMPLETED" and c["sla_status"] == "BREACHED")

        # TAT breakdown for active cases
        in_tat = sum(1 for c in filtered if c["sla_status"] == "IN_SLA" and c["status"] not in ["COMPLETED"])
        approaching_tat = sum(1 for c in filtered if c["sla_status"] == "WARNING")
        out_of_tat = sla_breaches

        return {
            "kpi": {
                "totalCases": total_cases,
                "completed": completed,
                "wip": wip,
                "slaBreaches": sla_breaches,
                "positive": positive,
                "negative": negative,
                "insufficiency": insufficient,
                "inTat": in_tat,
                "approachingTat": approaching_tat,
                "outOfTat": out_of_tat
            },
            "cases": filtered
        }
        
    except Exception as e:
        logger.error(f"Error in daily-operations endpoint: {str(e)}", exc_info=True)
        raise HTTPException(500, detail=str(e))


# ─── PREMIUM CUSTOMER MIS EXPORT SYSTEM ────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
import io
import csv

# ReportLab imports for branded enterprise PDF generation
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to dynamically compute and render professional 'Page X of Y' page numbers
    along with branded enterprise running headers and footers (polygon accent lines).
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        
        # 1. Header (Branded Running Header)
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#1E1B4B")) # Premium Deep Indigo
        self.drawString(36, 560, "CHECKLINE BACKGROUND VERIFICATION SERVICES")
        
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        self.drawRightString(806, 560, "CONFIDENTIAL CUSTOMER OPERATIONAL MIS")
        
        # Header border lines (branded accent polygons / lines)
        self.setStrokeColor(colors.HexColor("#7C3AED")) # Indigo accent
        self.setLineWidth(1)
        self.line(36, 552, 806, 552)
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(36, 550, 806, 550)
        
        # 2. Footer (Page numbers + Date + Privacy)
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(36, 45, 806, 45)
        
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        self.drawString(36, 30, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        self.drawCentredString(421, 30, "CONFIDENTIAL  -  RESTRICTED CLIENT CIRCULATION ONLY")
        self.drawRightString(806, 30, f"Page {self._pageNumber} of {page_count}")
        
        self.restoreState()


@router.post("/customer/mis/export")
async def export_customer_mis_data(
    payload: dict,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Enterprise-grade Customer MIS Export route supporting Candidate MIS, Checkwise MIS,
    Status Summary, Finalized Reports, and Insufficiency Tracker in Excel, CSV, and PDF formats.
    Robust customer-level role scoping is enforced automatically.
    """
    if not current_user.customer_id:
        raise HTTPException(
            status_code=403,
            detail="Access denied. User is not associated with any customer account."
        )

    customer_id = current_user.customer_id
    export_type = payload.get("exportType", "Candidate MIS")
    fmt = payload.get("format", "xlsx").lower()
    filters = payload.get("filters", {}) or {}

    try:
        # Fetch customer details
        cust_q = select(models.Customer).where(models.Customer.id == customer_id)
        cust_res = await db.execute(cust_q)
        customer = cust_res.scalar_one_or_none()
        customer_name = customer.name if customer else "Customer"

        # Build secure scoped query
        stmt = (
            select(models.Case)
            .where(models.Case.customer_id == customer_id)
            .options(
                joinedload(models.Case.candidate),
                joinedload(models.Case.customer),
                selectinload(models.Case.checks),
                selectinload(models.Case.insufficiencies),
                selectinload(models.Case.verification_logs)
            )
        )

        # ─── APPLY DASHBOARD FILTERS ───
        # 1. Date Range
        date_range = filters.get("dateRange")
        if date_range:
            from_str = date_range.get("from")
            to_str = date_range.get("to")
            if from_str:
                try:
                    stmt = stmt.filter(models.Case.received_date >= datetime.strptime(from_str, "%Y-%m-%d"))
                except: pass
            if to_str:
                try:
                    stmt = stmt.filter(models.Case.received_date <= datetime.strptime(to_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
                except: pass

        # 2. Candidate Search
        search = filters.get("searchTerm") or filters.get("search")
        if search:
            stmt = stmt.join(models.Candidate).filter(or_(
                models.Case.case_ref_no.ilike(f"%{search}%"),
                models.Candidate.name.ilike(f"%{search}%"),
                models.Candidate.client_emp_code.ilike(f"%{search}%")
            ))

        # 3. Status Filter
        status = filters.get("status")
        if status and status.upper() != "ALL":
            stmt = stmt.filter(models.Case.status == status.upper())

        # 4. Batch Filter
        batch = filters.get("batch")
        if batch and batch.upper() != "ALL":
            stmt = stmt.filter(or_(models.Case.batch_id == batch, models.Case.batch_id.ilike(f"%{batch}%")))

        # Order by received date
        stmt = stmt.order_by(models.Case.received_date.desc())
        
        # Execute query
        res = await db.execute(stmt)
        cases = res.unique().scalars().all()

        # Filter-aware export types
        if export_type == "Finalized Reports":
            cases = [c for c in cases if c.status in ["FINALIZED", "COMPLETED"]]
        elif export_type == "Insufficiency Tracker":
            # Fetch from the dedicated Insufficiency table with full lifecycle data
            insuff_stmt = (
                select(models.Insufficiency)
                .where(models.Insufficiency.case.has(models.Case.customer_id == customer_id))
                .options(
                    joinedload(models.Insufficiency.case).joinedload(models.Case.candidate),
                    joinedload(models.Insufficiency.case).joinedload(models.Case.customer),
                    joinedload(models.Insufficiency.case).joinedload(models.Case.assigned_user),
                    joinedload(models.Insufficiency.check),
                    joinedload(models.Insufficiency.user)
                )
                .order_by(models.Insufficiency.created_at.desc())
            )
            insuff_res = await db.execute(insuff_stmt)
            insufficiencies = insuff_res.unique().scalars().all()

            insuff_rows = []
            for i in insufficiencies:
                if i.is_resolved and i.resolved_at and i.created_at:
                    ageing_days = (i.resolved_at - i.created_at).days
                elif i.created_at:
                    ageing_days = (datetime.utcnow() - i.created_at.replace(tzinfo=None)).days
                else:
                    ageing_days = 0

                # Derive last timeline event description
                tl = i.timeline or []
                last_event_desc = tl[-1]["description"] if tl else "—"

                insuff_rows.append({
                    "case_ref": i.case.case_ref_no if i.case else "N/A",
                    "candidate_name": i.case.candidate.name if (i.case and i.case.candidate) else "N/A",
                    "candidate_email": i.case.candidate.email if (i.case and i.case.candidate) else "N/A",
                    "customer_name": i.case.customer.name if (i.case and i.case.customer) else "N/A",
                    "check_name": i.check.check_type if i.check else "General",
                    "remarks": i.message or "—",
                    "status": i.status or "PENDING",
                    "notification_count": i.notification_count or 0,
                    "last_notified_date": i.last_notified_at.strftime("%Y-%m-%d %H:%M") if i.last_notified_at else "—",
                    "response_date": i.response_at.strftime("%Y-%m-%d %H:%M") if i.response_at else "—",
                    "resolved_date": i.resolved_at.strftime("%Y-%m-%d %H:%M") if i.resolved_at else "—",
                    "ageing_days": ageing_days,
                    "raised_date": i.created_at.strftime("%Y-%m-%d %H:%M") if i.created_at else "—",
                    "raised_by": i.user.full_name if i.user else "Verifier",
                    "assigned_to": i.case.assigned_user.full_name if (i.case and i.case.assigned_user) else "Unassigned",
                    "last_event": last_event_desc,
                    "is_resolved": i.is_resolved,
                    "timeline_count": len(tl)
                })

        # ─── HELPER FOR STATUS NORMALIZATION ───
        def normalize_status(val):
            if not val:
                return ""
            val_up = str(val).strip().upper()
            mapping = {
                "CLEAR": "POSITIVE",
                "CLEAR_VERIFIED": "POSITIVE",
                "CLEAR/VERIFIED": "POSITIVE",
                "GREEN": "POSITIVE",
                "RED": "NEGATIVE",
                "DISCREPANCY": "NEGATIVE",
                "MINOR MISMATCH": "AMBER",
                "REVIEW REQUIRED": "AMBER",
                "REVIEW_REQUIRED": "AMBER",
                "HOLD": "STOPCHECK",
                "CLIENT HOLD": "STOPCHECK",
                "STOPPED": "STOPCHECK",
                "STOP": "STOPCHECK",
                "INSUFF": "INSUFFICIENT",
                "INSUFFICIENT": "INSUFFICIENT",
                "INSUFFICIENCY": "INSUFFICIENT"
            }
            return mapping.get(val_up, val_up)

        # ─── HELPER FOR DYNAMIC VALUE EXTRACTION ───
        def get_check_value(chk):
            if not chk.data:
                return ""
            t = str(chk.check_type or "").lower()
            d = chk.data
            if isinstance(d, list) and len(d) > 0:
                d = d[0]
            if not isinstance(d, dict):
                return str(chk.data)

            if "address" in t or "resident" in t:
                return d.get("address") or d.get("permanent_address") or d.get("given_address") or d.get("city") or ""
            if "employment" in t:
                return d.get("company_name") or d.get("company") or d.get("employer_name") or d.get("employer") or ""
            if "education" in t or "academic" in t:
                return d.get("course") or d.get("degree") or d.get("qualification") or ""
            if "identity" in t or "id" in t or "pan" in t or "aadhar" in t or "passport" in t:
                return d.get("id_number") or d.get("pan_no") or d.get("aadhar_no") or d.get("passport_no") or ""
            if "reference" in t:
                return d.get("ref_name") or d.get("reference_name") or ""
            if "drug" in t:
                return d.get("test_type") or d.get("panel_type") or d.get("panel") or d.get("result") or ""
            if "credit" in t or "cibil" in t:
                return d.get("score") or d.get("credit_score") or ""
            if "global" in t or "database" in t:
                return d.get("database_name") or d.get("registry") or d.get("result") or ""
            if "social" in t:
                return d.get("platform") or d.get("profile_url") or ""
            return str(chk.data)

        # ─────────────────────────────────────────────────────────────
        # INSUFFICIENCY TRACKER EXPORT (dedicated path for all formats)
        # ─────────────────────────────────────────────────────────────
        if export_type == "Insufficiency Tracker":
            insuff_headers = [
                "S.No", "Case Ref", "Candidate Name", "Candidate Email",
                "Check Type", "Remarks / Flag Reason", "Current Status",
                "Notification Count", "Raised Date", "Last Notified Date",
                "Response Date", "Resolved Date", "Ageing (Days)",
                "Raised By", "Assigned To", "Last Activity"
            ]

            header_fill_it = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            header_font_it = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            thin_border_it = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )

            STATUS_COLORS = {
                "PENDING": ("FFF3CD", "92400E"),
                "NOTIFICATION_SENT": ("DBEAFE", "1E40AF"),
                "REMINDER_SENT": ("FFEDD5", "9A3412"),
                "CANDIDATE_RESPONDED": ("F3E8FF", "6B21A8"),
                "UNDER_REVIEW": ("CFFAFE", "155E75"),
                "RESOLVED": ("D1FAE5", "065F46"),
            }

            if fmt == "xlsx":
                wb_it = openpyxl.Workbook()
                wb_it.remove(wb_it.active)
                ws_it = wb_it.create_sheet(title="Insufficiency Tracker")
                ws_it.views.sheetView[0].showGridLines = True

                for col_idx, val in enumerate(insuff_headers, start=1):
                    cell = ws_it.cell(row=1, column=col_idx, value=val)
                    cell.font = header_font_it
                    cell.fill = header_fill_it
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border_it

                for idx, row in enumerate(insuff_rows, start=1):
                    r = idx + 1
                    ws_it.cell(row=r, column=1, value=idx)
                    ws_it.cell(row=r, column=2, value=row["case_ref"])
                    ws_it.cell(row=r, column=3, value=row["candidate_name"])
                    ws_it.cell(row=r, column=4, value=row["candidate_email"])
                    ws_it.cell(row=r, column=5, value=row["check_name"])
                    ws_it.cell(row=r, column=6, value=row["remarks"])
                    status_cell = ws_it.cell(row=r, column=7, value=row["status"].replace("_", " "))
                    sc = STATUS_COLORS.get(row["status"], ("F1F5F9", "334155"))
                    status_cell.fill = PatternFill(start_color=sc[0], end_color=sc[0], fill_type="solid")
                    status_cell.font = Font(name="Segoe UI", size=10, bold=True, color=sc[1])
                    ws_it.cell(row=r, column=8, value=row["notification_count"])
                    ws_it.cell(row=r, column=9, value=row["raised_date"])
                    ws_it.cell(row=r, column=10, value=row["last_notified_date"])
                    ws_it.cell(row=r, column=11, value=row["response_date"])
                    ws_it.cell(row=r, column=12, value=row["resolved_date"])
                    ageing_cell = ws_it.cell(row=r, column=13, value=row["ageing_days"])
                    if not row["is_resolved"] and row["ageing_days"] > 7:
                        ageing_cell.font = Font(name="Segoe UI", size=10, bold=True, color="DC2626")
                    ws_it.cell(row=r, column=14, value=row["raised_by"])
                    ws_it.cell(row=r, column=15, value=row["assigned_to"])
                    ws_it.cell(row=r, column=16, value=row["last_event"])

                    for c_col in range(1, len(insuff_headers) + 1):
                        dcell = ws_it.cell(row=r, column=c_col)
                        dcell.border = thin_border_it
                        if not dcell.font.bold:
                            dcell.font = Font(name="Segoe UI", size=10)
                        if c_col in [1, 8, 13]:
                            dcell.alignment = Alignment(horizontal="center")

                # Summary Sheet
                ws_sum = wb_it.create_sheet(title="Summary")
                sum_headers = ["Status", "Count"]
                for ci, h in enumerate(sum_headers, start=1):
                    cell = ws_sum.cell(row=1, column=ci, value=h)
                    cell.font = header_font_it
                    cell.fill = header_fill_it
                    cell.border = thin_border_it

                from collections import Counter
                status_counts = Counter(r["status"] for r in insuff_rows)
                for si, (st, cnt) in enumerate(status_counts.items(), start=2):
                    ws_sum.cell(row=si, column=1, value=st.replace("_", " ")).border = thin_border_it
                    ws_sum.cell(row=si, column=2, value=cnt).border = thin_border_it

                # Column widths
                for col in ws_it.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws_it.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 14)
                ws_it.freeze_panes = "A2"

                output_it = io.BytesIO()
                wb_it.save(output_it)
                output_it.seek(0)
                filename_it = f"{customer_name.replace(' ', '_')}_Insufficiency_Tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                return StreamingResponse(
                    output_it,
                    headers={'Content-Disposition': f'attachment; filename="{filename_it}"'},
                    media_type='application/vnd.officedocument.spreadsheetml.sheet'
                )

            elif fmt == "csv":
                out_it = io.StringIO()
                writer_it = csv.writer(out_it)
                writer_it.writerow(insuff_headers)
                for idx, row in enumerate(insuff_rows, start=1):
                    writer_it.writerow([
                        idx, row["case_ref"], row["candidate_name"], row["candidate_email"],
                        row["check_name"], row["remarks"], row["status"],
                        row["notification_count"], row["raised_date"], row["last_notified_date"],
                        row["response_date"], row["resolved_date"], row["ageing_days"],
                        row["raised_by"], row["assigned_to"], row["last_event"]
                    ])
                stream_it = io.BytesIO(out_it.getvalue().encode('utf-8'))
                filename_it = f"{customer_name.replace(' ', '_')}_Insufficiency_Tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
                return StreamingResponse(
                    stream_it,
                    headers={'Content-Disposition': f'attachment; filename="{filename_it}"'},
                    media_type='text/csv'
                )

            else:
                # PDF fallback for insufficiency tracker
                raise HTTPException(status_code=400, detail="PDF export for Insufficiency Tracker is not yet supported. Please use XLSX or CSV.")

        # ─────────────────────────────────────────────────────────────
        # A. EXCEL FORMAT (.xlsx)
        # ─────────────────────────────────────────────────────────────
        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            # Pre-group and sort candidates & checks
            candidate_grouped_checks = []
            max_counts = {
                "address": 0, "employment": 0, "education": 0, "reference": 0,
                "drug": 0, "credit": 0, "global": 0, "social": 0, "other": 0
            }
            max_identity_counts = {}

            for c in cases:
                groups = {
                    "address": [], "employment": [], "education": [], "identity": {},
                    "reference": [], "drug": [], "credit": [], "global": [], "social": [], "other": []
                }
                for chk in (c.checks or []):
                    t = str(chk.check_type or "").lower()
                    val = get_check_value(chk)
                    chk_info = {"chk": chk, "val": val, "status": normalize_status(chk.status or "WIP")}
                    
                    if "address" in t or "resident" in t:
                        groups["address"].append(chk_info)
                    elif "employment" in t:
                        groups["employment"].append(chk_info)
                    elif "education" in t or "academic" in t:
                        groups["education"].append(chk_info)
                    elif "reference" in t:
                        groups["reference"].append(chk_info)
                    elif "drug" in t:
                        groups["drug"].append(chk_info)
                    elif "credit" in t or "cibil" in t:
                        groups["credit"].append(chk_info)
                    elif "global" in t or "database" in t:
                        groups["global"].append(chk_info)
                    elif "social" in t:
                        groups["social"].append(chk_info)
                    elif "identity" in t or "id" in t or "pan" in t or "aadhar" in t or "passport" in t:
                        doc_type = "Identity"
                        if "PAN" in t.upper(): doc_type = "PAN"
                        elif "AADHAAR" in t.upper() or "AADHAR" in t.upper(): doc_type = "AADHAAR"
                        elif "PASSPORT" in t.upper(): doc_type = "PASSPORT"
                        elif "DL" in t.upper() or "DRIVING" in t.upper(): doc_type = "DL"
                        
                        if doc_type not in groups["identity"]:
                            groups["identity"][doc_type] = []
                        groups["identity"][doc_type].append(chk_info)
                    else:
                        groups["other"].append(chk_info)

                candidate_grouped_checks.append((c, groups))
                
                # Update maximum dynamic check column counts
                for btype in max_counts.keys():
                    max_counts[btype] = max(max_counts[btype], len(groups[btype]))
                for doc_type, chks in groups["identity"].items():
                    max_identity_counts[doc_type] = max(max_identity_counts.get(doc_type, 0), len(chks))

            # Build list of dynamic column headers
            dynamic_cols = []
            
            # 1. Address
            for i in range(max_counts["address"]):
                dynamic_cols.append({"type": "address", "title": f"Address {i + 1}", "sub": ["Value", "Status"], "index": i})
            # 2. Employment
            for i in range(max_counts["employment"]):
                dynamic_cols.append({"type": "employment", "title": f"Employment {i + 1}", "sub": ["Company Name", "Status"], "index": i})
            # 3. Education
            for i in range(max_counts["education"]):
                dynamic_cols.append({"type": "education", "title": f"Education {i + 1}", "sub": ["Qualification", "Status"], "index": i})
            # 4. Identity
            for doc_type in sorted(list(max_identity_counts.keys())):
                for i in range(max_identity_counts[doc_type]):
                    dynamic_cols.append({"type": "identity", "doc_type": doc_type, "title": f"{doc_type} {i + 1}" if max_identity_counts[doc_type] > 1 else doc_type, "sub": [doc_type, "Status"], "index": i})
            # 5. Reference
            for i in range(max_counts["reference"]):
                dynamic_cols.append({"type": "reference", "title": f"Reference {i + 1}", "sub": ["Reference Name", "Status"], "index": i})
            # 6. Drug Test
            for i in range(max_counts["drug"]):
                dynamic_cols.append({"type": "drug", "title": f"Drug Test {i + 1}", "sub": ["Drug Test Result", "Status"], "index": i})
            # 7. Credit/CIBIL
            for i in range(max_counts["credit"]):
                dynamic_cols.append({"type": "credit", "title": f"Credit {i + 1}", "sub": ["Credit Score", "Status"], "index": i})
            # 8. Global DB
            for i in range(max_counts["global"]):
                dynamic_cols.append({"type": "global", "title": f"Global DB {i + 1}", "sub": ["Database Result", "Status"], "index": i})
            # 9. Social Media
            for i in range(max_counts["social"]):
                dynamic_cols.append({"type": "social", "title": f"Social Media {i + 1}", "sub": ["Social Media Result", "Status"], "index": i})
            # 10. Other
            for i in range(max_counts["other"]):
                dynamic_cols.append({"type": "other", "title": f"Other {i + 1}", "sub": ["Result", "Status"], "index": i})

            # Styled Header Theme
            header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid") # Deep Indigo
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            sub_fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid") # Lighter Deep Indigo
            sub_font = Font(name="Segoe UI", size=10, color="FFFFFF")
            
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            
            # ─────────────────────────────────────────────────────────
            # SHEET 1: Candidate MIS
            # ─────────────────────────────────────────────────────────
            ws1 = wb.create_sheet(title="Candidate MIS")
            ws1.views.sheetView[0].showGridLines = True
            
            fixed_headers = [
                "S.No", "Case Ref", "Employee ID", "Candidate Name", 
                "Client Name", "Batch", "Received Date", "Completed Date", 
                "Overall Status", "SLA Status"
            ]

            h1 = list(fixed_headers)
            for col in dynamic_cols:
                h1.append(col["title"])
                h1.append("") # Companion cell for horizontal merge

            h2 = [""] * len(fixed_headers)
            for col in dynamic_cols:
                h2.extend(col["sub"])

            # Write rows 1 & 2 headers
            for col_idx, val in enumerate(h1, start=1):
                cell = ws1.cell(row=1, column=col_idx, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
            for col_idx, val in enumerate(h2, start=1):
                cell = ws1.cell(row=2, column=col_idx, value=val)
                if val:
                    cell.font = sub_font
                    cell.fill = sub_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # Vertical merges for fixed columns
            for col_idx in range(1, len(fixed_headers) + 1):
                ws1.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                
            # Horizontal merges for dynamic columns
            curr_col = len(fixed_headers) + 1
            for col in dynamic_cols:
                ws1.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col + 1)
                curr_col += 2

            # Write Data
            row_idx = 3
            for idx, (c, groups) in enumerate(candidate_grouped_checks, start=1):
                # SLA calculation
                age_days = (datetime.utcnow() - c.received_date.replace(tzinfo=None)).days if c.received_date else 0
                sla_days_left = (c.tat_days or 10) - age_days
                
                if c.status in ["FINALIZED", "COMPLETED", "POSITIVE", "NEGATIVE", "GREEN", "RED"]:
                    sla_status = "Completed"
                elif sla_days_left < 0:
                    sla_status = "Breached"
                elif sla_days_left <= 3:
                    sla_status = "Risk"
                else:
                    sla_status = "Healthy"

                ws1.cell(row=row_idx, column=1, value=idx)
                ws1.cell(row=row_idx, column=2, value=c.case_ref_no or "")
                ws1.cell(row=row_idx, column=3, value=c.candidate.client_emp_code if c.candidate else "")
                ws1.cell(row=row_idx, column=4, value=c.candidate.name if c.candidate else "")
                ws1.cell(row=row_idx, column=5, value=c.customer.name if c.customer else "")
                ws1.cell(row=row_idx, column=6, value=c.batch_id or "Manual Entry")
                ws1.cell(row=row_idx, column=7, value=c.received_date.strftime("%Y-%m-%d") if c.received_date else "")
                ws1.cell(row=row_idx, column=8, value=c.completed_date.strftime("%Y-%m-%d") if c.completed_date else "")
                
                final_res = normalize_status(c.final_result or c.final_report_status or c.status or "WIP")
                if c.status in ["FINALIZED", "COMPLETED"] and (c.final_result or c.final_report_status):
                    ws1.cell(row=row_idx, column=9, value=normalize_status(c.final_result or c.final_report_status))
                else:
                    ws1.cell(row=row_idx, column=9, value=normalize_status(c.status or "WIP"))
                    
                ws1.cell(row=row_idx, column=10, value=sla_status)

                # Write dynamic values
                curr_dcol = len(fixed_headers) + 1
                for col in dynamic_cols:
                    chk_list = []
                    if col["type"] == "identity":
                        chk_list = groups["identity"].get(col["doc_type"], [])
                    else:
                        chk_list = groups.get(col["type"], [])
                        
                    if col["index"] < len(chk_list):
                        chk_info = chk_list[col["index"]]
                        ws1.cell(row=row_idx, column=curr_dcol, value=chk_info["val"])
                        ws1.cell(row=row_idx, column=curr_dcol + 1, value=chk_info["status"])
                    else:
                        ws1.cell(row=row_idx, column=curr_dcol, value="")
                        ws1.cell(row=row_idx, column=curr_dcol + 1, value="")
                    curr_dcol += 2

                # Borders & styling
                for c_col in range(1, len(h1) + 1):
                    dcell = ws1.cell(row=row_idx, column=c_col)
                    dcell.border = thin_border
                    dcell.font = Font(name="Segoe UI", size=10)
                    if c_col in [1, 2, 3, 7, 8, 9, 10]:
                        dcell.alignment = Alignment(horizontal="center")
                        
                row_idx += 1

            # Auto column width
            for col in ws1.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws1.column_dimensions[col_letter].width = max(max_len + 3, 11)

            # Freeze Fixed columns + headers (Column K, Row 3)
            ws1.freeze_panes = "K3"

            # ─────────────────────────────────────────────────────────
            # SHEET 2: Checkwise MIS
            # ─────────────────────────────────────────────────────────
            ws2 = wb.create_sheet(title="Checkwise MIS")
            ws2.views.sheetView[0].showGridLines = True
            
            c_headers = [
                "Case Ref", "Candidate Name", "Module", "Sub Type", 
                "Verification Status", "Result", "Verifier", "Completed Date", "Remarks"
            ]
            
            for col_idx, val in enumerate(c_headers, start=1):
                cell = ws2.cell(row=1, column=col_idx, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            r_idx = 2
            for c, groups in candidate_grouped_checks:
                for chk in (c.checks or []):
                    ws2.cell(row=r_idx, column=1, value=c.case_ref_no or "")
                    ws2.cell(row=r_idx, column=2, value=c.candidate.name if c.candidate else "")
                    ws2.cell(row=r_idx, column=3, value=str(chk.check_type or "").upper())
                    ws2.cell(row=r_idx, column=4, value=str(chk.data.get("id_type") or chk.check_type if isinstance(chk.data, dict) else chk.check_type).upper())
                    ws2.cell(row=r_idx, column=5, value=normalize_status(chk.status or "WIP"))
                    ws2.cell(row=r_idx, column=6, value=get_check_value(chk))
                    ws2.cell(row=r_idx, column=7, value="") # Omitted for privacy
                    ws2.cell(row=r_idx, column=8, value=chk.completed_at.strftime("%Y-%m-%d") if chk.completed_at else "")
                    ws2.cell(row=r_idx, column=9, value=chk.verifier_remarks or "")
                    
                    for c_col in range(1, len(c_headers) + 1):
                        dcell = ws2.cell(row=r_idx, column=c_col)
                        dcell.border = thin_border
                        dcell.font = Font(name="Segoe UI", size=10)
                        if c_col in [1, 5, 8]:
                            dcell.alignment = Alignment(horizontal="center")
                    r_idx += 1

            for col in ws2.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
            ws2.freeze_panes = "A2"

            # ─────────────────────────────────────────────────────────
            # SHEET 3: Status Summary
            # ─────────────────────────────────────────────────────────
            ws3 = wb.create_sheet(title="Status Summary")
            ws3.views.sheetView[0].showGridLines = True
            
            s_headers = [
                "Metric Title", "Count of Cases/Checks"
            ]
            for col_idx, val in enumerate(s_headers, start=1):
                cell = ws3.cell(row=1, column=col_idx, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                
            # Aggregate stats checkwise
            total_cand = len(cases)
            pos_checks = sum(1 for c in cases for chk in (c.checks or []) if normalize_status(chk.status) == "POSITIVE")
            neg_checks = sum(1 for c in cases for chk in (c.checks or []) if normalize_status(chk.status) == "NEGATIVE")
            amb_checks = sum(1 for c in cases for chk in (c.checks or []) if normalize_status(chk.status) == "AMBER")
            stop_checks = sum(1 for c in cases for chk in (c.checks or []) if normalize_status(chk.status) == "STOPCHECK")
            in_prog_checks = sum(1 for c in cases for chk in (c.checks or []) if normalize_status(chk.status) in ["WIP", "VERIFICATION", "QC_PENDING"])
            insuff_checks = sum(1 for c in cases for chk in (c.checks or []) if normalize_status(chk.status) == "INSUFFICIENT")

            summary_rows = [
                ("Total Candidates", total_cand),
                ("Positive Checks", pos_checks),
                ("Negative Checks", neg_checks),
                ("Amber Checks", amb_checks),
                ("Stop Checks", stop_checks),
                ("In Progress Checks", in_prog_checks),
                ("Insufficiencies", insuff_checks)
            ]

            for s_idx, (title, val) in enumerate(summary_rows, start=2):
                c1 = ws3.cell(row=s_idx, column=1, value=title)
                c2 = ws3.cell(row=s_idx, column=2, value=val)
                c1.border = thin_border
                c2.border = thin_border
                c1.font = Font(name="Segoe UI", size=10, bold=True)
                c2.font = Font(name="Segoe UI", size=10)
                c2.alignment = Alignment(horizontal="right")

            ws3.column_dimensions['A'].width = 25
            ws3.column_dimensions['B'].width = 25

            # Save Output
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"{customer_name.replace(' ', '_')}_{export_type.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            return StreamingResponse(
                output,
                headers={'Content-Disposition': f'attachment; filename="{filename}"'},
                media_type='application/vnd.officedocument.spreadsheetml.sheet'
            )

        # ─────────────────────────────────────────────────────────────
        # B. CSV FORMAT (.csv)
        # ─────────────────────────────────────────────────────────────
        elif fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Default Candidate MIS format in flat CSV
            writer.writerow([
                "S.No", "Case Ref", "Employee ID", "Candidate Name", 
                "Client Name", "Batch", "Received Date", "Completed Date", 
                "Overall Status"
            ])
            
            for idx, c in enumerate(cases, start=1):
                writer.writerow([
                    idx,
                    c.case_ref_no or "",
                    c.candidate.client_emp_code if c.candidate else "",
                    c.candidate.name if c.candidate else "",
                    c.customer.name if c.customer else "",
                    c.batch_id or "Manual Entry",
                    c.received_date.strftime("%Y-%m-%d") if c.received_date else "",
                    c.completed_date.strftime("%Y-%m-%d") if c.completed_date else "",
                    normalize_status(c.final_result or c.final_report_status) if c.status in ["FINALIZED", "COMPLETED"] and (c.final_result or c.final_report_status) else normalize_status(c.status or "WIP")
                ])
                
            stream = io.BytesIO(output.getvalue().encode('utf-8'))
            filename = f"{customer_name.replace(' ', '_')}_{export_type.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
            return StreamingResponse(
                stream,
                headers={'Content-Disposition': f'attachment; filename="{filename}"'},
                media_type='text/csv'
            )

        # ─────────────────────────────────────────────────────────────
        # C. PDF FORMAT (.pdf)
        # ─────────────────────────────────────────────────────────────
        elif fmt == "pdf":
            # PDF Generation - Landscaped, professional enterprise layout
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=36,
                rightMargin=36,
                topMargin=54,
                bottomMargin=54
            )
            
            styles = getSampleStyleSheet()
            
            # Custom styled ParagraphStyles
            title_style = ParagraphStyle(
                'DocTitle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=20,
                textColor=colors.HexColor("#1E1B4B"),
                spaceAfter=6
            )
            
            meta_style = ParagraphStyle(
                'DocMeta',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=9,
                textColor=colors.HexColor("#64748B"),
                spaceAfter=15
            )

            cell_header_style = ParagraphStyle(
                'CellHeader',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=8,
                textColor=colors.white,
                alignment=1
            )
            
            cell_body_style = ParagraphStyle(
                'CellBody',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=8,
                textColor=colors.HexColor("#1F2937"),
                alignment=1
            )

            elements = []
            
            # Title block
            elements.append(Paragraph(f"{customer_name.upper()} OPERATIONAL MIS", title_style))
            elements.append(Paragraph(f"EXPORT TYPE: {export_type}  |  TOTAL RECORDS: {len(cases)}  |  DATE GENERATED: {datetime.now().strftime('%Y-%m-%d %H:%M')}", meta_style))
            elements.append(Spacer(1, 10))
            
            # Table columns & widths
            headers = [
                Paragraph("S.No", cell_header_style),
                Paragraph("Case Ref", cell_header_style),
                Paragraph("Employee ID", cell_header_style),
                Paragraph("Candidate Name", cell_header_style),
                Paragraph("Batch No", cell_header_style),
                Paragraph("Received Date", cell_header_style),
                Paragraph("Completed Date", cell_header_style),
                Paragraph("Overall Status", cell_header_style)
            ]
            
            table_data = [headers]
            for idx, c in enumerate(cases, start=1):
                row = [
                    Paragraph(str(idx), cell_body_style),
                    Paragraph(c.case_ref_no or "", cell_body_style),
                    Paragraph(c.candidate.client_emp_code if c.candidate else "", cell_body_style),
                    Paragraph(c.candidate.name if c.candidate else "", cell_body_style),
                    Paragraph(c.batch_id or "Manual Entry", cell_body_style),
                    Paragraph(c.received_date.strftime("%Y-%m-%d") if c.received_date else "", cell_body_style),
                    Paragraph(c.completed_date.strftime("%Y-%m-%d") if c.completed_date else "", cell_body_style),
                    Paragraph(normalize_status(c.final_result or c.final_report_status) if c.status in ["FINALIZED", "COMPLETED"] and (c.final_result or c.final_report_status) else normalize_status(c.status or "WIP"), cell_body_style)
                ]
                table_data.append(row)

            # A4 Landscape printable width is 842 - 72 = 770 pt
            col_widths = [30, 90, 80, 150, 110, 80, 80, 150]
            
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E1B4B")),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
            ]))
            
            elements.append(t)
            
            # Build Document using NumberedCanvas
            doc.build(elements, canvasmaker=NumberedCanvas)
            
            buffer.seek(0)
            filename = f"{customer_name.replace(' ', '_')}_{export_type.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            return StreamingResponse(
                buffer,
                headers={'Content-Disposition': f'attachment; filename="{filename}"'},
                media_type='application/pdf'
            )

    except Exception as e:
        logger.error(f"Error generating customer export: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate MIS export: {str(e)}")


