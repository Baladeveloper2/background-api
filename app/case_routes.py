from fastapi import APIRouter, Depends, HTTPException, status, Response, BackgroundTasks, Request
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, and_, text, update, delete
from sqlalchemy.orm import joinedload, selectinload
from typing import List, Optional, Dict, Any
from . import models, schemas, enums, notification_utils, email_utils
from .database import get_async_db, get_read_db, SessionLocal
import uuid
import asyncio
import secrets
import os
from datetime import datetime, timedelta
from fastapi.responses import StreamingResponse
import requests
import httpx
from pypdf import PdfWriter, PdfReader
from io import BytesIO
import io
from .ocr_utils import get_scanner
from .ws import manager
from .cache import delete_cache, get_cache, set_cache, invalidate_dashboard_cache
from .auth_routes import check_module_permission, limiter, get_current_user, create_audit_log
from . import tat_utils, risk_utils
from .logging_config import logger

router = APIRouter(
    prefix="/cases",
    tags=["cases"]
)

def validate_case_completion(case_obj: models.Case):
    """
    Enterprise Validation Routine:
    Ensures a case cannot transition to COMPLETED until ALL associated verification 
    modules are completed by the assigned verifier.
    """
    if not case_obj.checks:
        return True
        
    incomplete_checks = []
    
    for chk in case_obj.checks:
        status_val = str(chk.status).upper()
        is_completed = status_val in ["POSITIVE", "NEGATIVE", "DISCREPANCY", "GREEN", "RED", "AMBER", "STOP", "COMPLETED", "QC_VERIFIED", "QC_PENDING"]
        
        if not is_completed:
            incomplete_checks.append(f"{chk.check_type} (Status: {chk.status})")
    
    if incomplete_checks:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot finalize case. Incomplete verification checks present: {', '.join(incomplete_checks)}. Complete these actions first."
        )
    return True

@router.get("/insufficient")
async def get_insufficient_cases(
    resolved: bool = False,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Fetch cases marked as insufficient for the current customer with filtering."""
    # 1. Fetch from new Insufficiency table
    stmt_new = (
        select(models.Insufficiency)
        .options(
            joinedload(models.Insufficiency.case).joinedload(models.Case.candidate),
            joinedload(models.Insufficiency.case).joinedload(models.Case.customer),
            joinedload(models.Insufficiency.check)
        )
        .filter(models.Insufficiency.is_resolved == resolved)
        .order_by(models.Insufficiency.created_at.desc())
    )
    
    if current_user.role == enums.UserRole.CUSTOMER:
        stmt_new = stmt_new.filter(models.Insufficiency.case.has(customer_id=current_user.customer_id))
        
    if from_date:
        try:
            f_dt = datetime.strptime(from_date, "%Y-%m-%d")
            stmt_new = stmt_new.filter(models.Insufficiency.created_at >= f_dt)
        except: pass
            
    if to_date:
        try:
            t_dt = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            stmt_new = stmt_new.filter(models.Insufficiency.created_at <= t_dt)
        except: pass

    res_new = await db.execute(stmt_new)
    insufficiencies = res_new.unique().scalars().all()
    
    results = []
    seen_check_ids = set()

    for i in insufficiencies:
        results.append({
            "id": i.case_id,
            "insufficiency_id": i.id,
            "case_ref_no": i.case.case_ref_no if i.case else "N/A",
            "candidate_name": i.case.candidate.name if i.case and i.case.candidate else "N/A",
            "customer_name": i.case.customer.name if i.case and i.case.customer else "N/A",
            "check_id": i.check_id,
            "check_name": i.check.check_type if i.check else "General",
            "marked_at": i.created_at.strftime("%Y-%m-%d %H:%M") if i.created_at else "N/A",
            "remarks": i.message or "No remarks found",
            "status": i.status,
            "documents": i.documents or [],
            "is_resolved": i.is_resolved,
            "resolved_at": i.updated_at.strftime("%Y-%m-%d %H:%M") if (i.is_resolved and i.updated_at) else None
        })
        seen_check_ids.add((i.case_id, i.check_id))

    # Legacy tracking doesn't fully apply to 'Completed' state, so exit early if requesting resolved items
    if resolved:
        return sorted(results, key=lambda x: x["marked_at"], reverse=True)

    # 2. Fetch from legacy Case status for backward compatibility
    stmt_legacy = select(models.Case).filter(
        models.Case.status == enums.CaseStatus.INSUFFICIENT
    ).options(
        joinedload(models.Case.candidate),
        joinedload(models.Case.customer),
        selectinload(models.Case.insufficiency_logs).options(
            joinedload(models.InsufficiencyLog.check)
        )
    )
    
    if current_user.role == enums.UserRole.CUSTOMER:
        stmt_legacy = stmt_legacy.filter(models.Case.customer_id == current_user.customer_id)
        
    if from_date:
        try:
            f_dt = datetime.strptime(from_date, "%Y-%m-%d")
            stmt_legacy = stmt_legacy.filter(models.Case.received_date >= f_dt)
        except: pass
            
    if to_date:
        try:
            t_dt = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            stmt_legacy = stmt_legacy.filter(models.Case.received_date <= t_dt)
        except: pass

    res_legacy = await db.execute(stmt_legacy)
    legacy_cases = res_legacy.unique().scalars().all()
    
    for c in legacy_cases:
        # Get latest check from logs
        last_log = c.insufficiency_logs[-1] if c.insufficiency_logs else None
        check_id = last_log.check_id if last_log else None
        
        if not check_id:
            ck_stmt = select(models.VerificationCheck.id).filter(models.VerificationCheck.case_id == c.id).limit(1)
            ck_res = await db.execute(ck_stmt)
            check_id = ck_res.scalar_one_or_none()
        
        if check_id and (c.id, check_id) not in seen_check_ids:
            results.append({
                "id": c.id,
                "case_ref_no": c.case_ref_no,
                "candidate_name": c.candidate.name if c.candidate else "N/A",
                "customer_name": c.customer.name if c.customer else "N/A",
                "check_id": check_id,
                "check_name": last_log.check.check_type if last_log and last_log.check else "General",
                "marked_at": last_log.marked_at.strftime("%Y-%m-%d %H:%M") if last_log else (c.received_date.strftime("%Y-%m-%d %H:%M") if c.received_date else "N/A"),
                "remarks": last_log.notes if last_log else "Legacy insufficiency",
                "status": "INSUFFICIENT",
                "is_resolved": False
            })
            seen_check_ids.add((c.id, check_id))
            
    # Final Sort (Latest First)
    return sorted(results, key=lambda x: x["marked_at"] if x["marked_at"] != "N/A" else "", reverse=True)

@router.post("/{case_id}/resolve-insufficiency")
async def resolve_insufficiency(
    case_id: str,
    data: schemas.ResolveInsufficiencyRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """Resolve an insufficiency by providing remarks and documents."""
    stmt = select(models.Case).filter(models.Case.id == case_id).options(joinedload(models.Case.candidate))
    res = await db.execute(stmt)
    case = res.scalar_one_or_none()
    
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    remarks = data.remarks
    if not remarks:
        raise HTTPException(status_code=400, detail="Remarks are mandatory for clearing insufficiency")
    
    # Add comment
    comment = models.CaseComment(
        case_id=case.id,
        user_id=current_user.id,
        content=f"INSUFFICIENCY CLEARED: {remarks}" + (f" (Check: {data.check_id})" if data.check_id else "")
    )
    db.add(comment)
    
    # Resolve the specific check if check_id provided
    if data.check_id:
        check_stmt = select(models.VerificationCheck).filter(models.VerificationCheck.id == data.check_id)
        check_res = await db.execute(check_stmt)
        check = check_res.scalar_one_or_none()
        if check:
            check.status = enums.CheckStatus.VERIFICATION
    
    # Update latest insufficiency log (Legacy)
    log_filter = [models.InsufficiencyLog.case_id == case_id, models.InsufficiencyLog.resolved_at == None]
    if data.check_id:
        log_filter.append(models.InsufficiencyLog.check_id == data.check_id)
    
    log_stmt = select(models.InsufficiencyLog).filter(*log_filter).order_by(models.InsufficiencyLog.marked_at.desc()).limit(1)
    log_res = await db.execute(log_stmt)
    log = log_res.scalar_one_or_none()
    if log:
        log.resolved_at = datetime.utcnow()

    # Update New Insufficiency Record (Source of Truth)
    new_insuff_stmt = select(models.Insufficiency).filter(
        models.Insufficiency.case_id == case_id,
        models.Insufficiency.is_resolved == False
    )
    if data.check_id:
        new_insuff_stmt = new_insuff_stmt.filter(models.Insufficiency.check_id == data.check_id)
    
    ni_res = await db.execute(new_insuff_stmt)
    new_insuff = ni_res.scalars().first()
    if new_insuff:
        new_insuff.status = "RESOLVED"
        new_insuff.is_resolved = True
        new_insuff.resolved_at = datetime.utcnow()
        new_insuff.resolved_by = current_user.id
        new_insuff.documents = data.documents or []
        new_insuff.updated_at = datetime.utcnow()
        new_insuff.updated_by = current_user.id

    # Check if any REMAINING insufficiencies are still in 'INSUFFICIENT' state (not uploaded)
    rem_stmt = select(func.count(models.Insufficiency.id)).filter(
        models.Insufficiency.case_id == case_id,
        models.Insufficiency.status == "INSUFFICIENT",
        models.Insufficiency.is_resolved == False
    )
    rem_res = await db.execute(rem_stmt)
    remaining_count = rem_res.scalar() or 0
    
    if remaining_count == 0:
        # Move case back to VERIFICATION so verifiers see it in their queue for review
        case.status = enums.CaseStatus.VERIFICATION
    
    # Notify internal team
    if case.assigned_to:
        await notification_utils.create_notification(
            db, case.assigned_to,
            "Insufficiency Response Received",
            f"Candidate {case.candidate.name} has submitted evidence for insufficiency. Remarks: {remarks}",
            enums.NotificationCategory.INSUFFICIENT_DOCS,
            case_id=case.id,
            background_tasks=background_tasks
        )
    
    await db.commit()
    await manager.broadcast({"type": "CASE_STATUS_UPDATE", "case_id": case.id, "status": case.status})
    
    return {"message": "Insufficiency response submitted successfully"}

@router.get("/invitations")
async def get_candidate_invitations(
    response: Response,
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncSession = Depends(get_read_db),
    current_user: models.User = Depends(get_current_user)
):
    """Fetch candidates in the invitation/self-service flow with optimized pagination and search."""
    # Base status conditions for invitations
    allowed_statuses = [enums.CaseStatus.PENDING, enums.CaseStatus.LINK_SHARED, enums.CaseStatus.DOCUMENTS_SUBMITTED]
    
    # Initialize base conditions
    base_conditions = []
    
    # If a specific status is requested, use it (narrowing down)
    if status and status != 'ALL':
        base_conditions.append(models.Case.status == status)
    else:
        # Default: show all invitation statuses
        base_conditions.append(models.Case.status.in_(allowed_statuses))
    
    if current_user.role == enums.UserRole.CUSTOMER:
        base_conditions.append(models.Case.customer_id == current_user.customer_id)
        
    if from_date:
        try:
            f_dt = datetime.strptime(from_date, "%Y-%m-%d")
            base_conditions.append(models.Case.received_date >= f_dt)
        except Exception as e:
            logger.warning(f"Invalid from_date format: {from_date}")
        
    if to_date:
        try:
            t_dt = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            base_conditions.append(models.Case.received_date <= t_dt)
        except Exception as e:
            logger.warning(f"Invalid to_date format: {to_date}")
            
    logger.info(f"Fetching invitations with status={status}, search={search}, from={from_date}, to={to_date}")
        
    count_stmt = select(func.count(models.Case.id)).filter(*base_conditions)
    stmt = select(models.Case).filter(*base_conditions)
    
    # OuterJoin Candidate to permit searching and seamless display even if missing
    count_stmt = count_stmt.outerjoin(models.Case.candidate).outerjoin(models.Case.customer)
    stmt = stmt.outerjoin(models.Case.candidate).outerjoin(models.Case.customer)
    
    # Prepare distinct loading strategy optimized for paginated results
    stmt = stmt.options(
        selectinload(models.Case.candidate),
        selectinload(models.Case.customer)
    )
    
    if search:
        search_filter = or_(
            models.Candidate.name.ilike(f"%{search}%"),
            models.Candidate.email.ilike(f"%{search}%"),
            models.Candidate.client_emp_code.ilike(f"%{search}%"),
            models.Customer.name.ilike(f"%{search}%")
        )
        count_stmt = count_stmt.filter(search_filter)
        stmt = stmt.filter(search_filter)
        
    # Execute distinct counting
    cnt_res = await db.execute(count_stmt)
    total_count = cnt_res.scalar() or 0
    
    # Sort by received_date (Latest First) and Execute Main Data Stream
    stmt = stmt.order_by(models.Case.received_date.desc(), models.Case.id.desc()).offset(skip).limit(limit)
    res = await db.execute(stmt)
    cases = res.unique().scalars().all()
    
    # Finalize Headers compliant with UI expectation
    response.headers["X-Total-Count"] = str(total_count)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Count"
    
    return [
        {
            "id": c.id,
            "candidate_name": c.candidate.name if c.candidate else "N/A",
            "email": c.candidate.email if c.candidate else "N/A",
            "phone": c.candidate.phone if c.candidate else "N/A",
            "emp_id": c.candidate.client_emp_code if c.candidate else "N/A",
            "status": c.status,
            "client_name": c.customer.name if c.customer else "N/A",
            "created_at": c.received_date.strftime("%Y-%m-%d %H:%M") if c.received_date else None,
            "link_shared_at": c.link_shared_at.strftime("%Y-%m-%d %H:%M") if c.link_shared_at else (c.received_date.strftime("%Y-%m-%d %H:%M") if c.status == enums.CaseStatus.LINK_SHARED else None),
            "submitted_at": c.submitted_at.strftime("%Y-%m-%d %H:%M") if c.submitted_at else None
        }
        for c in cases
    ]

@router.post("/invite-candidate")
async def invite_candidate(
    data: schemas.InviteCandidateRequest,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """Create a new candidate and case in PENDING status for invitation."""
    # Create Candidate
    candidate = models.Candidate(
        name=data.name,
        email=data.email,
        phone=data.phone,
        client_emp_code=data.emp_id
    )
    db.add(candidate)
    await db.flush()
    
    # Ensure customer_id is never None - fallback to current_user or provided data
    final_customer_id = current_user.customer_id if (hasattr(current_user, 'customer_id') and current_user.customer_id) else data.customer_id
    
    if not final_customer_id:
        raise HTTPException(status_code=400, detail="Customer association missing. Cannot create invitation.")

    # Retrieve Customer for shortcode
    customer_res = await db.execute(select(models.Customer).filter(models.Customer.id == final_customer_id))
    customer = customer_res.scalar_one_or_none()

    # Standardized prefix: CL-{SHORTCODE}
    sc = customer.short_code if customer and customer.short_code else (customer.name[:3].upper() if customer else "BGV")
    prefix = f"CL-{sc}-"

    # Get the current total count to start with
    count_res = await db.execute(select(func.count(models.Case.id)).filter(models.Case.customer_id == final_customer_id))
    count = count_res.scalar() or 0

    # Collision loop: Ensure unique reference number
    suffix_num = count + 1
    while True:
        case_ref = f"{prefix}{str(suffix_num).zfill(3)}"
        exists_res = await db.execute(select(models.Case.id).filter(models.Case.case_ref_no == case_ref))
        if not exists_res.scalar_one_or_none():
            break
        suffix_num += 1

    # Find or create a daily batch for this customer
    from datetime import datetime
    date_str = datetime.utcnow().strftime('%Y-%m-%d')
    batch_name = f"Batch-{sc}-{date_str}"
    
    batch_stmt = select(models.Batch).filter(
        models.Batch.customer_id == final_customer_id, 
        models.Batch.batch_no == batch_name
    ).limit(1)
    batch_res = await db.execute(batch_stmt)
    inv_batch = batch_res.scalar_one_or_none()

    if not inv_batch:
        batch_count_res = await db.execute(select(func.count(models.Batch.id)).filter(models.Batch.customer_id == final_customer_id))
        batch_count = batch_count_res.scalar() or 0
        inv_batch = models.Batch(
            customer_id=final_customer_id,
            batch_no=batch_name,
            cl_ref_no=f"CL-{sc}-{(batch_count + 1):03d}",
            cases_count=0,
            upload_date=datetime.utcnow()
        )
        db.add(inv_batch)
        await db.flush()
        
    # Increment cases count for the batch
    inv_batch.cases_count = (inv_batch.cases_count or 0) + 1

    new_case = models.Case(
        case_ref_no=case_ref,
        customer_id=final_customer_id,
        candidate_id=candidate.id,
        batch_id=inv_batch.id,
        status=enums.CaseStatus.PENDING
    )
    db.add(new_case)
    await db.commit()
    
    return {"message": "Candidate added to invitation list", "case_id": new_case.id}

@router.post("/{case_id}/send-bgv-link")
async def send_bgv_link(
    case_id: str,
    data: schemas.SendBgvLinkRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """Update case status to LINK_SHARED and trigger BGV form email."""
    stmt = select(models.Case).filter(models.Case.id == case_id).options(joinedload(models.Case.candidate))
    res = await db.execute(stmt)
    case = res.scalar_one_or_none()
    
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # 1. Get existing check types for this case to avoid duplicates
    existing_checks_res = await db.execute(select(models.VerificationCheck.check_type).filter(models.VerificationCheck.case_id == case.id))
    existing_check_types = {c[0] for c in existing_checks_res.all()}

    checks = data.checks
    # Only add checks that don't already exist
    for check_type in checks:
        if check_type not in existing_check_types:
            new_check = models.VerificationCheck(
                case_id=case.id,
                check_type=check_type,
                status=enums.CheckStatus.VERIFICATION
            )
            db.add(new_check)
    
    case.status = enums.CaseStatus.LINK_SHARED
    case.link_shared_at = datetime.utcnow()
    
    # Trigger email
    # Generate the form link using frontend URL (you might want to configure this in .env later)
    # For now, assuming the frontend runs on localhost:5173 or the domain
    import os
    frontend_url = os.getenv("FRONTEND_URL", "https://background-verification-91d11.web.app")
    form_link = f"{frontend_url}/candidate/form/{case.id}"
    
    if case.candidate and case.candidate.email:
        background_tasks.add_task(
            email_utils.send_bgv_invitation_email,
            to_email=case.candidate.email,
            candidate_name=case.candidate.name,
            form_link=form_link
        )
    
    await notification_utils.create_notification(
        db, current_user.id, # Internal notification
        "BGV Link Shared",
        f"BGV Link has been shared with {case.candidate.name} ({case.candidate.email})",
        enums.NotificationCategory.INSUFFICIENT_DOCS,
        case_id=case.id,
        background_tasks=background_tasks
    )
    
    await db.commit()
    return {"message": "BGV Link shared successfully"}

@router.get("/{case_id}/public")
async def get_case_public(case_id: str, db: AsyncSession = Depends(get_read_db)):
    """Fetch case details publicly for candidate self-service."""
    stmt = select(models.Case).filter(models.Case.id == case_id).options(
        joinedload(models.Case.candidate),
        joinedload(models.Case.customer),
        selectinload(models.Case.checks)
    )
    res = await db.execute(stmt)
    case = res.scalar_one_or_none()
    
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    return {
        "id": case.id,
        "candidate": {
            "name": case.candidate.name if case.candidate else "",
            "email": case.candidate.email if case.candidate else "",
            "phone": case.candidate.phone if case.candidate else "",
            "emp_id": case.candidate.client_emp_code if case.candidate else "",
            "dob": case.candidate.dob.isoformat() if case.candidate and case.candidate.dob else "",
            "pan_no": case.candidate.pan_no if case.candidate else "",
            "passport_no": case.candidate.passport_no if case.candidate else "",
            "nationality": case.candidate.nationality if case.candidate else "",
            "gender": case.candidate.gender if case.candidate else "",
            "address": case.candidate.address if case.candidate else ""
        },
        "customer_name": case.customer.name if case.customer else "",
        "status": case.status,
        "checks": [chk.check_type for chk in case.checks]
    }

@router.post("/{case_id}/submit-documents")
async def submit_candidate_documents(
    case_id: str,
    payload: dict,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Called when a candidate finishes submitting their documents via the BGV form.
    - Transitions status: LINK_SHARED → DOCUMENTS_SUBMITTED
    - Sends notifications to the client contact and the internal team.
    """
    stmt = (
        select(models.Case)
        .filter(models.Case.id == case_id)
        .options(
            joinedload(models.Case.candidate),
            joinedload(models.Case.customer),
            selectinload(models.Case.checks)
        )
    )
    res = await db.execute(stmt)
    case = res.scalar_one_or_none()

    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    # Mark as documents submitted
    case.status = enums.CaseStatus.DOCUMENTS_SUBMITTED
    case.submitted_at = datetime.utcnow()

    candidate_data = payload.get("candidate_data", {})

    # 1. Update Candidate Profile with fresh details
    if case.candidate:
        case.candidate.name = f"{candidate_data.get('first_name', '')} {candidate_data.get('last_name', '')}".strip()
        case.candidate.email = candidate_data.get('email', case.candidate.email)
        case.candidate.phone = candidate_data.get('contact_no', case.candidate.phone)
        case.candidate.gender = candidate_data.get('gender', case.candidate.gender)
        case.candidate.address = candidate_data.get('address', case.candidate.address)

        dob_str = candidate_data.get('dob')
        if dob_str:
            try:
                case.candidate.dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
            except:
                pass

    # 2. Map sections to check types and store data
    section_mapping = {
        'identities': ['identity', 'aadhaar', 'pan'],
        'educations': ['education'],
        'employments': ['employment'],
        'criminal': ['criminal'],
        'credit': ['credit', 'cibil'],
        'global': ['global'],
        'database': ['database'],
        'drug_tests': ['drug'],
        'references': ['reference'],
        'social_media': ['social'],
        'addresses': ['address', 'resident']
    }

    for check in case.checks:
        check_type = check.check_type.lower()
        target_section = None
        for section, types in section_mapping.items():
            if any(t in check_type for t in types):
                target_section = section
                break

        if target_section and target_section in candidate_data:
            # Store the specific records for this module
            section_data = candidate_data[target_section]
            check.data = section_data

    # 3. Synchronize data to Candidate model for unified view (Registry/Detail)
    if case.candidate:
        details = case.candidate.address_details or {}
        all_docs = list(case.candidate.documents or [])
        
        # Mapping for unified registry structure
        registry_map = {
            'identities': 'identities',
            'educations': 'educations',
            'employments': 'employments',
            'criminal': 'criminal_records',
            'credit': 'cibil_checks',
            'global': 'global_database_checks',
            'database': 'global_database_checks',
            'drug_tests': 'drug_tests',
            'references': 'references',
            'social_media': 'social_media_details',
            'addresses': 'addresses'
        }

        for section, target_key in registry_map.items():
            if section in candidate_data:
                details[target_key] = candidate_data[section]
                
                # Extract any uploaded documents from this section
                if isinstance(candidate_data[section], list):
                    for record in candidate_data[section]:
                        if isinstance(record, dict) and 'files' in record:
                            for f in record['files']:
                                # Add check_type to the file metadata for filtering in UI
                                f_copy = dict(f)
                                f_copy['check_type'] = section.capitalize()
                                if section == 'educations': f_copy['check_type'] = 'Educational'
                                elif section == 'criminal': f_copy['check_type'] = 'Criminal'
                                elif section == 'drug_tests': f_copy['check_type'] = 'Drug'
                                elif section == 'addresses': f_copy['check_type'] = 'Address'
                                elif section == 'employments': f_copy['check_type'] = 'Employment'
                                elif section == 'identities': f_copy['check_type'] = 'Identity'
                                elif section == 'references': f_copy['check_type'] = 'Reference'
                                elif section == 'credit': f_copy['check_type'] = 'CIBIL'
                                elif section == 'global': f_copy['check_type'] = 'Global Database'
                                elif section == 'social_media': f_copy['check_type'] = 'Social'
                                else: f_copy['check_type'] = section.capitalize()
                                
                                # Avoid duplicates
                                if not any(existing.get('public_id') == f_copy.get('public_id') for existing in all_docs):
                                    all_docs.append(f_copy)
        
        case.candidate.address_details = details
        case.candidate.documents = all_docs
        # Force session refresh if needed
        db.add(case.candidate)

    # Find all active customer users to notify
    customer_users = []
    if case.customer_id:
        cu_stmt = select(models.User).filter(
            models.User.customer_id == case.customer_id,
            models.User.role == enums.UserRole.CUSTOMER,
            models.User.status == enums.Status.ACTIVE
        )
        cu_res = await db.execute(cu_stmt)
        customer_users = cu_res.scalars().all()
        logger.info(f"Notification Targets (Customer): found {len(customer_users)} users")

    logger.info(f"Triggering notifications for case {case.id}")

    # Fire dual notifications (client + internal team)
    await notification_utils.notify_documents_submitted(
        db=db,
        case_id=case.id,
        case_ref=case.case_ref_no or case.id,
        candidate_name=case.candidate.name if case.candidate else "Candidate",
        customer_users=customer_users,
        background_tasks=background_tasks
    )

    await db.commit()

    # WebSocket broadcast for real-time listing update
    await manager.broadcast({
        "type": "CASE_STATUS_UPDATE",
        "case_id": case.id,
        "status": enums.CaseStatus.DOCUMENTS_SUBMITTED
    })

    return {"message": "Documents submitted successfully. Notifications sent to the client and internal team."}

@router.post("/bulk-mark-insufficient")
async def bulk_mark_insufficient(data: schemas.BulkInsufficientRequest, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    case_ids = data.case_ids
    reason = data.reason or "Incomplete documentation"
    
    if not case_ids:
        raise HTTPException(status_code=400, detail="No cases selected")

    # Update cases
    await db.execute(
        update(models.Case)
        .where(models.Case.id.in_(case_ids))
        .values(status=enums.CaseStatus.INSUFFICIENT, comments=reason)
    )

    # Fetch assigned users to notify them their work is on hold
    stmt = select(models.Case).filter(models.Case.id.in_(case_ids))
    res = await db.execute(stmt)
    cases = res.scalars().all()

    notified_users = set()
    for c in cases:
        if c.assigned_to:
            notified_users.add(c.assigned_to)
            await notification_utils.create_notification(
                db, c.assigned_to,
                "Case Marked Insufficient",
                f"Protocol {c.case_ref_no} has been moved to Insufficiency: {reason}",
                enums.NotificationCategory.INSUFFICIENT_DOCS,
                case_id=c.id,
                background_tasks=background_tasks
            )
        
        # Audit Log
        log = models.AuditLog(
            user_id=current_user.id,
            action="BULK_INSUFFICIENT",
            details=f"Case {c.case_ref_no} flagged as insufficient: {reason}",
            resource_id=c.id
        )
        db.add(log)

    await db.commit()
    await invalidate_dashboard_cache()
    
    # Global workforce update signal
    await manager.broadcast({"type": "WORKFORCE_UPDATE", "source": "bulk_insufficient"})
    
    return {"message": f"Successfully moved {len(case_ids)} cases to Insufficiency.", "notified_user_count": len(notified_users)}

@router.post("/{case_id}/checks/{check_id}/raise-insufficiency")
async def raise_check_insufficiency(
    case_id: str,
    check_id: str,
    data: schemas.RaiseInsufficiencyRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Raises an insufficiency at the check level.
    - Updates check status to INSUFFICIENT
    - Updates case status to INSUFFICIENT
    - Creates a secure insufficiency record with a token
    - Sends email to candidate with evidence upload link
    - Notifies internal stakeholders and customer users
    """
    # 1. Fetch Case and Check with Candidate/Customer info
    stmt = select(models.Case).filter(models.Case.id == case_id).options(
        joinedload(models.Case.candidate),
        joinedload(models.Case.customer),
        selectinload(models.Case.checks)
    )
    res = await db.execute(stmt)
    case = res.scalar_one_or_none()
    
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
        
    check = next((c for c in case.checks if c.id == check_id), None)
    if not check:
        raise HTTPException(status_code=404, detail="Check not found in this case")
        
    # 2. Update Statuses
    check.status = enums.CheckStatus.INSUFFICIENT
    case.status = enums.CaseStatus.INSUFFICIENT
    
    # Check for existing open insufficiency for this check to avoid duplicates
    exist_stmt = select(models.Insufficiency).filter(
        models.Insufficiency.case_id == case.id,
        models.Insufficiency.check_id == check.id,
        models.Insufficiency.is_resolved == False
    )
    exist_res = await db.execute(exist_stmt)
    new_insuff = exist_res.scalars().first()

    import uuid
    token = uuid.uuid4().hex

    if new_insuff:
        # Update existing record (Re-notification)
        new_insuff.message = data.message
        new_insuff.token = token
        new_insuff.status = "INSUFFICIENT"
        new_insuff.updated_at = datetime.utcnow()
    else:
        # Create fresh record
        case.insufficiency_count = (case.insufficiency_count or 0) + 1
        new_insuff = models.Insufficiency(
            case_id=case.id,
            check_id=check.id,
            raised_by=current_user.id,
            role=current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role),
            message=data.message,
            status="INSUFFICIENT",
            token=token,
            documents=data.documents or []
        )
        db.add(new_insuff)
    
    db.add(check)
    db.add(case)
    
    # 4. Audit Log
    audit = models.AuditLog(
        user_id=current_user.id,
        action="RAISE_INSUFFICIENCY",
        resource_id=case.id,
        details=f"Insufficiency raised for check {check.check_type}. Message: {data.message}"
    )
    db.add(audit)
    
    # 5. Find Customer Users to notify
    customer_users = []
    if case.customer_id:
        cu_stmt = select(models.User).filter(
            models.User.customer_id == case.customer_id,
            models.User.role == enums.UserRole.CUSTOMER,
            models.User.status == "ACTIVE"
        )
        cu_res = await db.execute(cu_stmt)
        customer_users = cu_res.scalars().all()
        
    # 6. Notify Stakeholders
    try:
        await notification_utils.notify_insufficiency_raised(
            db=db,
            case_id=case.id,
            case_ref=case.case_ref_no or case.id,
            candidate_name=case.candidate.name if case.candidate else "Candidate",
            check_id=check.id,
            check_name=check.check_type,
            raised_by_name=current_user.full_name or current_user.email,
            raised_by_role=current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role),
            message=data.message,
            customer_user_ids=[u.id for u in customer_users],
            background_tasks=background_tasks
        )
    except Exception as e:
        logger.error(f"Stakeholder notification failed but continuing: {e}")

    # 7. Send Email to Candidate
    if case.candidate and case.candidate.email:
        import os
        frontend_url = os.getenv("FRONTEND_URL", "http://localhost:5173")
        upload_link = f"{frontend_url}/candidate/insufficiency/{token}"
        
        background_tasks.add_task(
            email_utils.send_insufficiency_email,
            to_email=case.candidate.email,
            candidate_name=case.candidate.name,
            case_ref=case.case_ref_no or case.id, # Fixed keyword mismatch
            check_name=check.check_type,
            custom_message=data.message,
            upload_link=upload_link
        )
    
    await db.commit()
    
    # WebSocket broadcast for real-time status update in dashboards
    await manager.broadcast({
        "type": "CASE_STATUS_UPDATE", 
        "case_id": case.id, 
        "status": case.status
    })
    
    return {
        "status": "success", 
        "insufficiency_id": new_insuff.id, 
        "token": token
    }

@router.get("/export")
async def export_mis_data(
    status: Optional[str] = None,
    customer_id: Optional[str] = None,
    customer_name: Optional[str] = None,
    search: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """Generates structured MIS Excel report for cases using a professionally styled template."""
    try:
        stmt = select(models.Case).options(
            joinedload(models.Case.candidate),
            joinedload(models.Case.customer),
            joinedload(models.Case.assigned_user),
            selectinload(models.Case.checks),
            selectinload(models.Case.verification_logs),
            selectinload(models.Case.insufficiencies)
        )
        
        # Join Customer if client_name or search is used
        if customer_name or search:
            stmt = stmt.join(models.Customer, models.Case.customer_id == models.Customer.id)

        # RBAC for Customers
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        role_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
        if user_role == "CUSTOMER" or role_name == "CUSTOMER":
            stmt = stmt.filter(
                models.Case.customer_id == current_user.customer_id,
                ~models.Case.status.in_([enums.CaseStatus.PENDING, enums.CaseStatus.LINK_SHARED, enums.CaseStatus.DOCUMENTS_SUBMITTED])
            )
        else:
            if customer_id:
                stmt = stmt.filter(models.Case.customer_id == customer_id)
            if customer_name:
                stmt = stmt.filter(models.Customer.name == customer_name)

        if status and status.strip().upper() not in ('ALL', ''):
            _FINAL = ['FINALIZED','COMPLETED','POSITIVE','NEGATIVE','DISCREPANCY','UNABLE TO VERIFY','HOLD','INSUFFICIENT','QC_VERIFIED','CLOSED']
            if status.strip().upper() in ('COMPLETED', 'FINALIZED'):
                stmt = stmt.filter(models.Case.status.in_(_FINAL))
            else:
                stmt = stmt.filter(models.Case.status == status)
        
        if search:
            stmt = stmt.join(models.Candidate).filter(or_(
                models.Case.case_ref_no.ilike(f"%{search}%"),
                models.Candidate.name.ilike(f"%{search}%")
            ))

        if from_date:
            try:
                f_date = datetime.strptime(from_date, "%Y-%m-%d")
                stmt = stmt.filter(models.Case.received_date >= f_date)
            except: pass
        if to_date:
            try:
                t_date = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                stmt = stmt.filter(models.Case.received_date <= t_date)
            except: pass

        stmt = stmt.order_by(models.Case.received_date.desc())
        res = await db.execute(stmt)
        cases_gen = res.unique().scalars().all()

        # Load the MIS Format.xlsx template file
        import openpyxl
        from copy import copy
        
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "frontend", "src", "assets", "MIS Format.xlsx"
        )
        
        if not os.path.exists(template_path):
            # Graceful fallback to working directory or local asset folder
            template_path = "MIS Format.xlsx"

        workbook = openpyxl.load_workbook(template_path)

        def copy_style(src_cell, dest_cell):
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.border = copy(src_cell.border)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.number_format = src_cell.number_format

        def apply_row_styles(ws, ref_row_idx, target_row_idx):
            if target_row_idx == ref_row_idx:
                return
            for col_idx in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=ref_row_idx, column=col_idx)
                dest_cell = ws.cell(row=target_row_idx, column=col_idx)
                copy_style(src_cell, dest_cell)

        # -------------------------------------------------------------
        # SHEET 1: Case wise MIS
        # -------------------------------------------------------------
        ws1 = workbook["Case wise MIS"]
        
        # Clear existing mock/old data starting from row 2
        for r in range(2, ws1.max_row + 1):
            for col in range(1, ws1.max_column + 1):
                ws1.cell(row=r, column=col).value = None

        row_idx1 = 2
        for idx, c in enumerate(cases_gen, start=1):
            apply_row_styles(ws1, 2, row_idx1)
            
            # Extract insufficiency timelines
            raised_date = None
            raised_remarks = []
            cleared_date = None
            cleared_remarks = []
            for ins in (c.insufficiencies or []):
                if ins.created_at:
                    if not raised_date or ins.created_at < raised_date:
                        raised_date = ins.created_at
                    raised_remarks.append(ins.message or "")
                if ins.is_resolved and ins.resolved_at:
                    if not cleared_date or ins.resolved_at > cleared_date:
                        cleared_date = ins.resolved_at
                    if ins.resolved_remarks:
                        cleared_remarks.append(ins.resolved_remarks)

            # Extract log-based event timelines
            interim_date = None
            interim_remarks = []
            stop_date = None
            stop_remarks = []
            amber_date = None
            amber_remarks = []
            negative_remarks = []
            reinitiate_date = None
            reinitiate_remarks = []
            reinitiate_completed_date = None
            reinitiate_completed_remarks = []

            for log in (c.verification_logs or []):
                action_lower = str(log.action or "").lower()
                remarks_lower = str(log.remarks or "").lower()
                
                # Interim
                if log.new_status == "INTERIM" or "interim" in action_lower or "interim" in remarks_lower:
                    if not interim_date or log.created_at < interim_date:
                        interim_date = log.created_at
                    if log.remarks:
                        interim_remarks.append(log.remarks)
                
                # Stop Check
                if log.new_status == "STOP" or "stop" in action_lower or "stop" in remarks_lower:
                    if not stop_date or log.created_at < stop_date:
                        stop_date = log.created_at
                    if log.remarks:
                        stop_remarks.append(log.remarks)

                # Amber
                if log.new_status == "AMBER" or "amber" in action_lower or "amber" in remarks_lower:
                    if not amber_date or log.created_at < amber_date:
                        amber_date = log.created_at
                    if log.remarks:
                        amber_remarks.append(log.remarks)

                # Negative
                if log.new_status == "NEGATIVE" or "negative" in action_lower or "negative" in remarks_lower:
                    if log.remarks:
                        negative_remarks.append(log.remarks)

                # Re-initiated
                if "reinitiate" in action_lower or "re-initiate" in action_lower or "reinitiated" in remarks_lower or "re-initiated" in remarks_lower:
                    if not reinitiate_date or log.created_at < reinitiate_date:
                        reinitiate_date = log.created_at
                    if log.remarks:
                        reinitiate_remarks.append(log.remarks)

                # Re-initiation Completed
                if ("reinitiate" in remarks_lower or "re-initiate" in remarks_lower) and log.new_status == "COMPLETED":
                    if not reinitiate_completed_date or log.created_at < reinitiate_completed_date:
                        reinitiate_completed_date = log.created_at
                    if log.remarks:
                        reinitiate_completed_remarks.append(log.remarks)

            if not negative_remarks and c.final_report_status == 'NEGATIVE' and c.qc_remarks:
                negative_remarks.append(c.qc_remarks)

            ws1.cell(row=row_idx1, column=1, value=idx)
            ws1.cell(row=row_idx1, column=2, value=c.received_date.replace(tzinfo=None) if c.received_date else None)
            ws1.cell(row=row_idx1, column=3, value=c.case_ref_no or "")
            ws1.cell(row=row_idx1, column=4, value=c.candidate.client_emp_code if c.candidate else "")
            ws1.cell(row=row_idx1, column=5, value=c.candidate.name if c.candidate else "")
            ws1.cell(row=row_idx1, column=6, value=c.status or "")
            ws1.cell(row=row_idx1, column=7, value=c.completed_date.replace(tzinfo=None) if c.completed_date else None)
            ws1.cell(row=row_idx1, column=8, value=raised_date.replace(tzinfo=None) if raised_date else None)
            ws1.cell(row=row_idx1, column=9, value="; ".join(raised_remarks) if raised_remarks else "")
            ws1.cell(row=row_idx1, column=10, value=cleared_date.replace(tzinfo=None) if cleared_date else None)
            ws1.cell(row=row_idx1, column=11, value="; ".join(cleared_remarks) if cleared_remarks else "")
            ws1.cell(row=row_idx1, column=12, value=interim_date.replace(tzinfo=None) if interim_date else None)
            ws1.cell(row=row_idx1, column=13, value="; ".join(interim_remarks) if interim_remarks else "")
            ws1.cell(row=row_idx1, column=14, value=stop_date.replace(tzinfo=None) if stop_date else None)
            ws1.cell(row=row_idx1, column=15, value="; ".join(stop_remarks) if stop_remarks else "")
            ws1.cell(row=row_idx1, column=16, value=amber_date.replace(tzinfo=None) if amber_date else None)
            ws1.cell(row=row_idx1, column=17, value="; ".join(amber_remarks) if amber_remarks else "")
            ws1.cell(row=row_idx1, column=18, value="; ".join(negative_remarks) if negative_remarks else "")
            ws1.cell(row=row_idx1, column=19, value=reinitiate_date.replace(tzinfo=None) if reinitiate_date else None)
            ws1.cell(row=row_idx1, column=20, value="; ".join(reinitiate_remarks) if reinitiate_remarks else "")
            ws1.cell(row=row_idx1, column=21, value=reinitiate_completed_date.replace(tzinfo=None) if reinitiate_completed_date else None)
            ws1.cell(row=row_idx1, column=22, value="; ".join(reinitiate_completed_remarks) if reinitiate_completed_remarks else "")

            row_idx1 += 1

        # -------------------------------------------------------------
        # SHEET 2: Check wise MIS
        # -------------------------------------------------------------
        ws2 = workbook["Check wise MIS"]
        
        # Clear existing mock/old data starting from row 4
        for r in range(4, ws2.max_row + 1):
            for col in range(1, ws2.max_column + 1):
                ws2.cell(row=r, column=col).value = None

        row_idx2 = 4
        for idx, c in enumerate(cases_gen, start=1):
            apply_row_styles(ws2, 4, row_idx2)

            # Columns 1 to 20 are the same as Case wise MIS
            raised_date = None
            raised_remarks = []
            cleared_date = None
            cleared_remarks = []
            for ins in (c.insufficiencies or []):
                if ins.created_at:
                    if not raised_date or ins.created_at < raised_date:
                        raised_date = ins.created_at
                    raised_remarks.append(ins.message or "")
                if ins.is_resolved and ins.resolved_at:
                    if not cleared_date or ins.resolved_at > cleared_date:
                        cleared_date = ins.resolved_at
                    if ins.resolved_remarks:
                        cleared_remarks.append(ins.resolved_remarks)

            interim_date = None
            interim_remarks = []
            stop_date = None
            stop_remarks = []
            amber_date = None
            amber_remarks = []
            negative_remarks = []
            reinitiate_date = None
            reinitiate_remarks = []
            reinitiate_completed_date = None
            reinitiate_completed_remarks = []

            for log in (c.verification_logs or []):
                action_lower = str(log.action or "").lower()
                remarks_lower = str(log.remarks or "").lower()
                
                if log.new_status == "INTERIM" or "interim" in action_lower or "interim" in remarks_lower:
                    if not interim_date or log.created_at < interim_date:
                        interim_date = log.created_at
                    if log.remarks:
                        interim_remarks.append(log.remarks)
                
                if log.new_status == "STOP" or "stop" in action_lower or "stop" in remarks_lower:
                    if not stop_date or log.created_at < stop_date:
                        stop_date = log.created_at
                    if log.remarks:
                        stop_remarks.append(log.remarks)

                if log.new_status == "AMBER" or "amber" in action_lower or "amber" in remarks_lower:
                    if not amber_date or log.created_at < amber_date:
                        amber_date = log.created_at
                    if log.remarks:
                        amber_remarks.append(log.remarks)

                if log.new_status == "NEGATIVE" or "negative" in action_lower or "negative" in remarks_lower:
                    if log.remarks:
                        negative_remarks.append(log.remarks)

                if "reinitiate" in action_lower or "re-initiate" in action_lower or "reinitiated" in remarks_lower or "re-initiated" in remarks_lower:
                    if not reinitiate_date or log.created_at < reinitiate_date:
                        reinitiate_date = log.created_at
                    if log.remarks:
                        reinitiate_remarks.append(log.remarks)

                if ("reinitiate" in remarks_lower or "re-initiate" in remarks_lower) and log.new_status == "COMPLETED":
                    if not reinitiate_completed_date or log.created_at < reinitiate_completed_date:
                        reinitiate_completed_date = log.created_at
                    if log.remarks:
                        reinitiate_completed_remarks.append(log.remarks)

            if not negative_remarks and c.final_report_status == 'NEGATIVE' and c.qc_remarks:
                negative_remarks.append(c.qc_remarks)

            ws2.cell(row=row_idx2, column=1, value=idx)
            ws2.cell(row=row_idx2, column=2, value=c.received_date.replace(tzinfo=None) if c.received_date else None)
            ws2.cell(row=row_idx2, column=3, value=c.case_ref_no or "")
            ws2.cell(row=row_idx2, column=4, value=c.candidate.client_emp_code if c.candidate else "")
            ws2.cell(row=row_idx2, column=5, value=c.candidate.name if c.candidate else "")
            ws2.cell(row=row_idx2, column=6, value=c.status or "")
            ws2.cell(row=row_idx2, column=7, value=c.completed_date.replace(tzinfo=None) if c.completed_date else None)
            ws2.cell(row=row_idx2, column=8, value=raised_date.replace(tzinfo=None) if raised_date else None)
            ws2.cell(row=row_idx2, column=9, value="; ".join(raised_remarks) if raised_remarks else "")
            ws2.cell(row=row_idx2, column=10, value=cleared_date.replace(tzinfo=None) if cleared_date else None)
            ws2.cell(row=row_idx2, column=11, value="; ".join(cleared_remarks) if cleared_remarks else "")
            ws2.cell(row=row_idx2, column=12, value=interim_date.replace(tzinfo=None) if interim_date else None)
            ws2.cell(row=row_idx2, column=13, value="; ".join(interim_remarks) if interim_remarks else "")
            ws2.cell(row=row_idx2, column=14, value=stop_date.replace(tzinfo=None) if stop_date else None)
            ws2.cell(row=row_idx2, column=15, value="; ".join(stop_remarks) if stop_remarks else "")
            ws2.cell(row=row_idx2, column=16, value=amber_date.replace(tzinfo=None) if amber_date else None)
            ws2.cell(row=row_idx2, column=17, value="; ".join(amber_remarks) if amber_remarks else "")
            ws2.cell(row=row_idx2, column=18, value="; ".join(negative_remarks) if negative_remarks else "")
            ws2.cell(row=row_idx2, column=19, value=reinitiate_date.replace(tzinfo=None) if reinitiate_date else None)
            ws2.cell(row=row_idx2, column=20, value="; ".join(reinitiate_remarks) if reinitiate_remarks else "")

            # Specific verification check mapping
            ad = c.candidate.address_details if c.candidate and c.candidate.address_details else {}
            
            # Address Check (Col 21, Col 22)
            addr_check = next((chk for chk in (c.checks or []) if 'address' in str(chk.check_type).lower()), None)
            ws2.cell(row=row_idx2, column=21, value="Permanent")
            ws2.cell(row=row_idx2, column=22, value=addr_check.status if addr_check else "N/A")

            # Criminal Check (Col 23, Col 24)
            crim_check = next((chk for chk in (c.checks or []) if 'criminal' in str(chk.check_type).lower()), None)
            ws2.cell(row=row_idx2, column=23, value="")
            ws2.cell(row=row_idx2, column=24, value=crim_check.status if crim_check else "N/A")

            # Employments (Col 25 to 33)
            employments = ad.get("employments", []) if isinstance(ad.get("employments"), list) else []
            emp_checks = [chk for chk in (c.checks or []) if 'employment' in str(chk.check_type).lower()]
            
            # Employment 1
            emp1_name = employments[0].get("employer_name") if len(employments) > 0 else ""
            emp1_chk = emp_checks[0] if len(emp_checks) > 0 else None
            ws2.cell(row=row_idx2, column=25, value=emp1_name)
            ws2.cell(row=row_idx2, column=26, value=emp1_name)
            ws2.cell(row=row_idx2, column=27, value=emp1_chk.status if emp1_chk else "N/A")

            # Employment 2
            emp2_name = employments[1].get("employer_name") if len(employments) > 1 else ""
            emp2_chk = emp_checks[1] if len(emp_checks) > 1 else None
            ws2.cell(row=row_idx2, column=28, value=emp2_name)
            ws2.cell(row=row_idx2, column=29, value=emp2_name)
            ws2.cell(row=row_idx2, column=30, value=emp2_chk.status if emp2_chk else "N/A")

            # Employment 3
            emp3_name = employments[2].get("employer_name") if len(employments) > 2 else ""
            emp3_chk = emp_checks[2] if len(emp_checks) > 2 else None
            ws2.cell(row=row_idx2, column=31, value=emp3_name)
            ws2.cell(row=row_idx2, column=32, value=emp3_name)
            ws2.cell(row=row_idx2, column=33, value=emp3_chk.status if emp3_chk else "N/A")

            # Educations (Col 34 to 39)
            educations = ad.get("educations", []) if isinstance(ad.get("educations"), list) else []
            edu_checks = [chk for chk in (c.checks or []) if 'education' in str(chk.check_type).lower()]

            # Education 1
            edu1_course = educations[0].get("course") if len(educations) > 0 else ""
            edu1_chk = edu_checks[0] if len(edu_checks) > 0 else None
            ws2.cell(row=row_idx2, column=34, value=edu1_course)
            ws2.cell(row=row_idx2, column=35, value=edu1_course)
            ws2.cell(row=row_idx2, column=36, value=edu1_chk.status if edu1_chk else "N/A")

            # Education 2
            edu2_course = educations[1].get("course") if len(educations) > 1 else ""
            edu2_chk = edu_checks[1] if len(edu_checks) > 1 else None
            ws2.cell(row=row_idx2, column=37, value=edu2_course)
            ws2.cell(row=row_idx2, column=38, value=edu2_course)
            ws2.cell(row=row_idx2, column=39, value=edu2_chk.status if edu2_chk else "N/A")

            # Identity ID1 (Col 40 to 42)
            identities = ad.get("identities", []) if isinstance(ad.get("identities"), list) else []
            id_checks = [chk for chk in (c.checks or []) if 'identity' in str(chk.check_type).lower()]
            id1_type = "PAN"
            if len(identities) > 0:
                id1_type = identities[0].get("id_type") or "PAN"
            elif c.candidate and c.candidate.identity_type:
                id1_type = c.candidate.identity_type
            id1_chk = id_checks[0] if len(id_checks) > 0 else None
            ws2.cell(row=row_idx2, column=40, value=id1_type)
            ws2.cell(row=row_idx2, column=41, value=id1_type)
            ws2.cell(row=row_idx2, column=42, value=id1_chk.status if id1_chk else "N/A")

            row_idx2 += 1

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f"MIS_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output, 
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}, 
            media_type='application/vnd.officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting MIS data: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/strategic-mis")
async def get_strategic_mis(
    customer_id: Optional[str] = None,
    customer_name: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """Returns detailed candidate list with check-wise status for Strategic MIS."""
    try:
        stmt = select(models.Case).options(
            selectinload(models.Case.candidate),
            selectinload(models.Case.customer),
            selectinload(models.Case.checks),
            selectinload(models.Case.assigned_user)
        )

        # RBAC Isolation
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        if "CUSTOMER" in user_role or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER"):
            stmt = stmt.filter(
                models.Case.customer_id == current_user.customer_id,
                ~models.Case.status.in_([enums.CaseStatus.PENDING, enums.CaseStatus.LINK_SHARED, enums.CaseStatus.DOCUMENTS_SUBMITTED])
            )
        elif customer_id:
            stmt = stmt.filter(models.Case.customer_id == customer_id)
        elif customer_name:
            stmt = stmt.join(models.Customer).filter(models.Customer.name == customer_name)

        if from_date:
            stmt = stmt.filter(models.Case.received_date >= datetime.strptime(from_date, "%Y-%m-%d"))
        if to_date:
            stmt = stmt.filter(models.Case.received_date <= datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59))

        stmt = stmt.order_by(models.Case.received_date.desc())
        res = await db.execute(stmt)
        cases = res.unique().scalars().all()

        report_keys = ['identity', 'resident', 'education', 'employment', 'criminal', 'reference', 'drug', 'credit', 'global', 'social']

        report = []
        for c in cases:
            # Map check statuses with technical-to-report translation
            check_map = {k: "NA" for k in report_keys}
            detailed_checks = []

            # Deduplicate checks by type to prevent counting redundancies (e.g. 12 cards for 10 checks)
            distinct_checks = {}
            for chk in c.checks:
                status_label = "In Progress"
                s = str(chk.status).upper()
                if s in ["GREEN", "POSITIVE", "COMPLETED", "VERIFIED", "QC_VERIFIED"]: status_label = "Positive"
                elif s in ["RED", "NEGATIVE"]: status_label = "Negative"
                elif s == "AMBER": status_label = "Amber"
                elif s == "INTERIM": status_label = "Interim"
                
                # Normalize technical type to report key
                tech_type = chk.check_type.lower()
                report_key = tech_type # Default
                
                if "address" in tech_type or "resident" in tech_type: report_key = "resident"
                elif "identity" in tech_type: report_key = "identity"
                elif "education" in tech_type or "academic" in tech_type: report_key = "education"
                elif "employment" in tech_type: report_key = "employment"
                elif "criminal" in tech_type: report_key = "criminal"
                elif "reference" in tech_type: report_key = "reference"
                elif "drug" in tech_type: report_key = "drug"
                elif "cibil" in tech_type or "credit" in tech_type: report_key = "credit"
                elif "global" in tech_type: report_key = "global"
                elif "social" in tech_type: report_key = "social"
                
                if report_key in report_keys:
                    check_map[report_key] = status_label
                
                # Take the most recent one if duplicates exist
                chk_data = {
                    "id": chk.id,
                    "type": chk.check_type,
                    "status": status_label,
                    "raw_status": chk.status,
                    "data": chk.data or {},
                    "remarks": chk.verifier_remarks or "",
                    "updated_at": chk.verified_date.strftime("%Y-%m-%d %H:%M") if chk.verified_date else None,
                    "verified_date": chk.verified_date.strftime("%Y-%m-%d") if chk.verified_date else None,
                    "created_at_raw": chk.verified_date or datetime.min # For sorting
                }
                
                if chk.check_type not in distinct_checks:
                    distinct_checks[chk.check_type] = chk_data
                else:
                    # If duplicate, keep the one with more data or more recent
                    if chk.verified_date and (not distinct_checks[chk.check_type]["created_at_raw"] or chk.verified_date > distinct_checks[chk.check_type]["created_at_raw"]):
                         distinct_checks[chk.check_type] = chk_data

            detailed_checks = list(distinct_checks.values())



            report.append({
                "id": c.id,
                "case_ref_no": c.case_ref_no,
                "emp_code": c.candidate.client_emp_code if c.candidate else "N/A",
                "candidate_name": c.candidate.name if c.candidate else "N/A",
                "received_date": c.received_date.strftime("%Y-%m-%d") if c.received_date else "N/A",
                "completed_date": c.completed_date.strftime("%Y-%m-%d") if c.completed_date else "N/A",
                "customer_name": c.customer.name if c.customer else "N/A",
                "status": c.status,
                "assigned_to_name": c.assigned_user.full_name if c.assigned_user else "Unallocated",
                "tat_days": c.tat_days or 10,
                "checks": check_map,
                "detailed_checks": detailed_checks
            })
        return report
    except Exception as e:
        logger.error(f"Error generating strategic MIS: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/strategic-mis/{case_id}")
async def get_case_strategic_details(
    case_id: str,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """Returns strategic MIS details for a single specific case."""
    try:
        stmt = select(models.Case).options(
            joinedload(models.Case.candidate),
            joinedload(models.Case.customer),
            joinedload(models.Case.assigned_user),
            selectinload(models.Case.checks).selectinload(models.VerificationCheck.documents).selectinload(models.VerificationDocument.uploader),
            selectinload(models.Case.checks).selectinload(models.VerificationCheck.logs).selectinload(models.VerificationLog.performer),
            selectinload(models.Case.checks).selectinload(models.VerificationCheck.assigned_verifier),
            joinedload(models.Case.verification_logs).joinedload(models.VerificationLog.performer)
        ).filter(models.Case.id == case_id)


        # RBAC Isolation
        user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
        if "CUSTOMER" in user_role or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER"):
            stmt = stmt.filter(models.Case.customer_id == current_user.customer_id)

        res = await db.execute(stmt)
        c = res.unique().scalar_one_or_none()

        if not c:
            raise HTTPException(status_code=404, detail="Case not found or access denied")

        report_keys = ['identity', 'resident', 'education', 'employment', 'criminal', 'reference', 'drug', 'credit', 'global', 'social']
        check_map = {k: "NA" for k in report_keys}
        detailed_checks = []

        for chk in c.checks:
            status_label = "In Progress"
            s = str(chk.status).upper()
            if s in ["GREEN", "POSITIVE", "COMPLETED", "VERIFIED", "QC_VERIFIED"]: status_label = "Positive"
            elif s in ["RED", "NEGATIVE"]: status_label = "Negative"
            elif s == "AMBER": status_label = "Amber"
            elif s == "INTERIM": status_label = "Interim"
            
            tech_type = chk.check_type.lower()
            report_key = tech_type
            if "address" in tech_type or "resident" in tech_type: report_key = "resident"
            elif "identity" in tech_type: report_key = "identity"
            elif "education" in tech_type or "academic" in tech_type: report_key = "education"
            elif "employment" in tech_type: report_key = "employment"
            elif "criminal" in tech_type: report_key = "criminal"
            elif "reference" in tech_type: report_key = "reference"
            elif "drug" in tech_type: report_key = "drug"
            elif "cibil" in tech_type or "credit" in tech_type: report_key = "credit"
            elif "global" in tech_type: report_key = "global"
            elif "social" in tech_type: report_key = "social"
            
            if report_key in report_keys:
                check_map[report_key] = status_label
            
            detailed_checks.append({
                "id": chk.id,
                "type": chk.check_type,
                "status": status_label,
                "raw_status": chk.status,
                "data": chk.data or {},
                "remarks": chk.verifier_remarks or "",
                "confidence_score": chk.confidence_score or 0.0,
                "api_sync_status": chk.api_sync_status or "NOT_SYNCED",
                "assigned_verifier_name": chk.assigned_verifier.full_name if chk.assigned_verifier else "Unallocated",
                "updated_at": chk.verified_date.strftime("%Y-%m-%d %H:%M") if chk.verified_date else None,
                "verified_date": chk.verified_date.strftime("%Y-%m-%d") if chk.verified_date else None,
                "documents": [
                    {
                        "id": d.id,
                        "file_name": d.file_name,
                        "file_url": d.file_url,
                        "file_type": d.file_type,
                        "uploaded_at": d.uploaded_at.isoformat() if d.uploaded_at else None,
                        "uploaded_by_name": d.uploader.full_name if d.uploader else "System"
                    } for d in chk.documents
                ],
                "logs": [
                    {
                        "id": l.id,
                        "action": l.action,
                        "remarks": l.remarks,
                        "old_status": l.old_status,
                        "new_status": l.new_status,
                        "created_at": l.created_at.isoformat() if l.created_at else None,
                        "performer_name": l.performer.full_name if l.performer else "System"
                    } for l in chk.logs
                ]
            })


        # Auto-detect QC Verifier from logs if not explicitly set in qc_id
        qc_name = "Not Assigned"
        if c.qc_user:
            qc_name = c.qc_user.full_name if c.qc_user.full_name else c.qc_user.email
        else:
            # Fallback: Find the person who performed a finalization action
            final_keywords = ["COMPLETE", "AUTHORIZE", "QC", "APPROVE", "FINAL", "CLOSE", "DONE", "REPORT", "ISSUE"]
            for l in reversed(c.verification_logs):
                action_upper = str(l.action).upper()
                new_status_upper = str(l.new_status).upper()
                if new_status_upper in ["COMPLETED", "QC_VERIFIED", "APPROVED", "FINALIZED"] or \
                   any(kw in action_upper for kw in final_keywords):
                    qc_name = l.performer.full_name if (l.performer and l.performer.full_name) else (l.performer.email if l.performer else "System")
                    break
            
            # Last Resort: If still Not Assigned but case is COMPLETED, take the very last log performer
            if qc_name == "Not Assigned" and str(c.status).upper() in ["COMPLETED", "QC_VERIFIED"]:
                if c.verification_logs:
                    # Get the most recent log that has a performer
                    for l in reversed(c.verification_logs):
                        if l.performer:
                            qc_name = l.performer.full_name if l.performer.full_name else l.performer.email
                            break

        # Find Verifier Assignment Date
        verifier_assigned_at = c.assigned_at.strftime("%Y-%m-%d %H:%M") if c.assigned_at else "N/A"
        
        # Find QC Assignment Date (from logs)
        qc_assigned_at = "N/A"
        for l in c.verification_logs:
            if str(l.new_status).upper() == "QC":
                qc_assigned_at = l.created_at.strftime("%Y-%m-%d %H:%M")
                break

        return {
            "id": c.id,
            "case_ref_no": c.case_ref_no,
            "emp_code": c.candidate.client_emp_code if c.candidate else "N/A",
            "candidate_name": c.candidate.name if c.candidate else "N/A",
            "candidate_email": c.candidate.email if c.candidate else "N/A",
            "candidate_phone": c.candidate.phone if c.candidate else "N/A",
            "received_date": c.received_date.strftime("%Y-%m-%d") if c.received_date else "N/A",
            "completed_date": c.completed_date.strftime("%Y-%m-%d") if c.completed_date else "N/A",
            "customer_name": c.customer.name if c.customer else "N/A",
            "status": c.status,
            "current_stage": c.status.replace('_', ' ').capitalize() if c.status else "Pending",
            "final_decision": "Positive" if str(c.status).upper() in ["COMPLETED", "QC_VERIFIED", "GREEN", "POSITIVE"] else ("Negative" if str(c.status).upper() in ["RED", "NEGATIVE"] else "Pending"),
            "assigned_to_name": c.assigned_user.full_name if (c.assigned_user and c.assigned_user.full_name) else (c.assigned_user.email if c.assigned_user else "Unallocated"),
            "verifier_assigned_at": verifier_assigned_at,
            "qc_verifier_name": qc_name,
            "qc_assigned_at": qc_assigned_at,
            "ops_manager": "Aravind (Ops Head)", # Placeholder or map from account manager
            "tat_days": c.tat_days or 10,
            "risk_score": c.risk_score or 15,
            "checks": check_map,
            "detailed_checks": detailed_checks or [],
            "verification_logs": [
                {
                    "id": l.id,
                    "action": l.action,
                    "remarks": l.remarks,
                    "old_status": l.old_status,
                    "new_status": l.new_status,
                    "created_at": l.created_at.isoformat() if l.created_at else None,
                    "performer_name": (l.performer.full_name if l.performer else "System")
                } for l in (c.verification_logs or [])
            ]
        }


    except Exception as e:
        logger.error(f"Error fetching case strategic details: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{case_id}/history", dependencies=[Depends(get_current_user)])
async def get_case_history(case_id: str, db: AsyncSession = Depends(get_read_db)):
    # Use per-case cache key to avoid cross-case cache collisions
    cache_key = f"case_history:{case_id}"
    from .cache import get_cache, set_cache
    cached = await get_cache(cache_key)
    if cached is not None:
        return cached

    stmt = select(models.AuditLog, models.User.full_name).join(models.User, models.AuditLog.user_id == models.User.id).filter(models.AuditLog.resource_id == case_id).order_by(models.AuditLog.timestamp.desc())
    res = await db.execute(stmt)
    history = []
    for log, name in res.all():
        history.append({
            "id": log.id,
            "action": log.action,
            "details": log.details,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
            "user_name": name
        })
    await set_cache(cache_key, history, ttl=300)
    return history

@router.post("", response_model=schemas.Case, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
async def create_case(case: schemas.CaseCreate, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    if not case.case_ref_no:
        customer_res = await db.execute(select(models.Customer).filter(models.Customer.id == case.customer_id))
        customer = customer_res.scalar_one_or_none()
        
        # Standardized prefix: CL-{SHORTCODE}
        sc = customer.short_code if customer and customer.short_code else (customer.name[:3].upper() if customer else "BGV")
        prefix = f"CL-{sc}-"
        
        count_res = await db.execute(select(func.count(models.Case.id)).filter(models.Case.customer_id == case.customer_id))
        count = count_res.scalar() or 0
        case.case_ref_no = f"{prefix}{str(count + 1).zfill(3)}"
        
    # For Customer role, force their own customer_id
    user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    if user_role == "CUSTOMER" or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER"):
        case.customer_id = current_user.customer_id

    # Resolve batch_id (it might be batch_no from frontend)
    if case.batch_id and case.batch_id.strip():
        batch_id_val = case.batch_id.strip()
        b_stmt = select(models.Batch.id).filter(or_(models.Batch.id == batch_id_val, models.Batch.batch_no == batch_id_val))
        b_res = await db.execute(b_stmt)
        resolved_batch_id = b_res.scalar_one_or_none()
        if resolved_batch_id:
            case.batch_id = resolved_batch_id

    db_case = models.Case(**case.dict())
    db.add(db_case)
    await db.commit()
    res = await db.execute(
        select(models.Case).options(
            joinedload(models.Case.candidate),
            joinedload(models.Case.customer),
            selectinload(models.Case.checks).options(
                selectinload(models.VerificationCheck.documents).joinedload(models.VerificationDocument.uploader),
                selectinload(models.VerificationCheck.logs).joinedload(models.VerificationLog.performer),
                selectinload(models.VerificationCheck.assigned_verifier),
                selectinload(models.VerificationCheck.finalized_user)
            )
        ).filter(models.Case.id == db_case.id)
    )
    db_case = res.unique().scalar_one()
    
    # Trigger Background Summary Refresh
    from .stats_service import refresh_dashboard_summary
    background_tasks.add_task(refresh_dashboard_summary, db, db_case.customer_id)

    return db_case

@router.post("/create-full", response_model=schemas.Case, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
async def create_case_full(case_data: schemas.CaseCreateExtended, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    # 1. Create/Get Candidate
    candidate_dict = case_data.candidate.dict()
    db_candidate = models.Candidate(**candidate_dict)
    db.add(db_candidate)
    await db.flush() # Get candidate ID

    # 2. Create Case with Collision-Aware Reference Generation
    if not case_data.case_ref_no:
        customer_res = await db.execute(select(models.Customer).filter(models.Customer.id == case_data.customer_id))
        customer = customer_res.scalar_one_or_none()
        
        # Standardized prefix: CL-{SHORTCODE}
        sc = customer.short_code if customer and customer.short_code else (customer.name[:3].upper() if customer else "BGV")
        prefix = f"CL-{sc}-"
        
        # Get the current total count to start with
        count_res = await db.execute(select(func.count(models.Case.id)).filter(models.Case.customer_id == case_data.customer_id))
        count = count_res.scalar() or 0
        
        # Collision loop: Ensure unique reference number
        suffix_num = count + 1
        while True:
            case_ref = f"{prefix}{str(suffix_num).zfill(3)}"
            exists_res = await db.execute(select(models.Case.id).filter(models.Case.case_ref_no == case_ref))
            if not exists_res.scalar_one_or_none():
                break
            suffix_num += 1
    else:
        case_ref = case_data.case_ref_no

    # For Customer role, force their own customer_id
    user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    target_customer_id = current_user.customer_id if (user_role == "CUSTOMER" or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER")) else case_data.customer_id

    # Resolve batch_id (it might be batch_no from frontend)
    print(f"DEBUG: create_case_full incoming batch_id: {case_data.batch_id}")
    resolved_batch_id = None
    if case_data.batch_id and case_data.batch_id.strip():
        batch_id_val = case_data.batch_id.strip()
        # Try finding by ID or batch_no
        b_stmt = select(models.Batch.id).filter(or_(models.Batch.id == batch_id_val, models.Batch.batch_no == batch_id_val))
        b_res = await db.execute(b_stmt)
        resolved_batch_id = b_res.scalar_one_or_none()
    
    print(f"DEBUG: create_case_full resolved_batch_id: {resolved_batch_id}")

    db_case = models.Case(
        case_ref_no=case_ref,
        customer_id=target_customer_id,
        candidate_id=db_candidate.id,
        batch_id=resolved_batch_id,
        status=models.CaseStatus.PENDING,
        received_date=datetime.utcnow(),
        file_no=case_data.file_no
    )
    db.add(db_case)
    await db.flush()

    # 3. Create Verification Checks
    for service in case_data.services:
        rate = case_data.check_rates.get(service, 0.0) if case_data.check_rates else 0.0
        # Get scope from map or global field
        check_scope = case_data.scope_of_work
        if case_data.check_scopes and service in case_data.check_scopes:
            check_scope = case_data.check_scopes[service]
            
        db_check = models.VerificationCheck(
            case_id=db_case.id,
            check_type=service,
            status=models.CheckStatus.VERIFICATION,
            rate=rate,
            data={"scope_of_work": check_scope} if check_scope else {}
        )
        db.add(db_check)
    
    await db.commit()
    
    # Reload with relationships for response validation
    stmt = select(models.Case).options(
        joinedload(models.Case.candidate),
        joinedload(models.Case.customer),
        selectinload(models.Case.checks).options(
            selectinload(models.VerificationCheck.documents).joinedload(models.VerificationDocument.uploader),
            selectinload(models.VerificationCheck.logs).joinedload(models.VerificationLog.performer),
            selectinload(models.VerificationCheck.assigned_verifier),
            selectinload(models.VerificationCheck.finalized_user)
        )
    ).filter(models.Case.id == db_case.id)
    res = await db.execute(stmt)
    db_case = res.unique().scalar_one()

    # Trigger Background Summary Refresh
    from .stats_service import refresh_dashboard_summary
    background_tasks.add_task(refresh_dashboard_summary, db, db_case.customer_id)

    return db_case

@router.get("/allocation-stats", dependencies=[Depends(check_module_permission("bvs", "verification", action="read"))])
async def get_allocation_stats(db: AsyncSession = Depends(get_read_db)):
    FINAL_STATUSES = [
        'FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE',
        'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT',
        'QC_VERIFIED', 'CLOSED'
    ]

    # Cases not yet assigned to anyone — covers all pre-assignment statuses
    unallocated_stmt = select(func.count(models.Case.id)).filter(
        models.Case.assigned_to == None,
        models.Case.status.notin_(FINAL_STATUSES)
    )

    # Cases actively being worked on by a verifier
    allocated_stmt = select(func.count(models.Case.id)).filter(
        models.Case.assigned_to != None,
        models.Case.status.notin_(FINAL_STATUSES)
    )

    # All terminally finalized cases
    completed_stmt = select(func.count(models.Case.id)).filter(
        models.Case.status.in_(FINAL_STATUSES)
    )

    unallocated_res = await db.execute(unallocated_stmt)
    allocated_res   = await db.execute(allocated_stmt)
    completed_res   = await db.execute(completed_stmt)

    return {
        "unallocated":      unallocated_res.scalar() or 0,
        "allocated":        allocated_res.scalar() or 0,
        "completed":        completed_res.scalar() or 0,
        "active_verifiers": 0
    }

@router.get("/recommend-allocation")
async def recommend_allocation(db: AsyncSession = Depends(get_async_db)):
    """
    Analyzes workforce capacity and suggests the best verifiers for new assignments
    based on current load and completion trends.
    Optimized: Uses 3 aggregate queries instead of 2N individual queries.
    """
    seven_days_ago = datetime.utcnow() - timedelta(days=7)

    # 1. Fetch all active verifiers
    v_stmt = select(models.User.id, models.User.full_name).filter(
        models.User.role == enums.UserRole.VERIFIER,
        models.User.status == enums.Status.ACTIVE
    )
    v_res = await db.execute(v_stmt)
    verifiers = v_res.all()

    # 2. Aggregate active load per verifier (single query)
    load_stmt = select(
        models.Case.assigned_to,
        func.count(models.Case.id)
    ).filter(
        models.Case.status.in_([enums.CaseStatus.VERIFICATION]),
        models.Case.assigned_to.isnot(None)
    ).group_by(models.Case.assigned_to)
    load_res = await db.execute(load_stmt)
    load_map = {row[0]: row[1] for row in load_res.all()}

    # 3. Aggregate 7-day velocity per verifier (single query)
    vel_stmt = select(
        models.Case.assigned_to,
        func.count(models.Case.id)
    ).filter(
        models.Case.completed_date >= seven_days_ago,
        models.Case.assigned_to.isnot(None)
    ).group_by(models.Case.assigned_to)
    vel_res = await db.execute(vel_stmt)
    vel_map = {row[0]: row[1] for row in vel_res.all()}

    # 4. Build recommendations from pre-fetched data
    recommendations = []
    for v_id, full_name in verifiers:
        active_load = load_map.get(v_id, 0)
        velocity = vel_map.get(v_id, 0)
        score = active_load / (velocity + 1)

        recommendations.append({
            "user_id": v_id,
            "full_name": full_name,
            "active_load": active_load,
            "velocity_7d": velocity,
            "efficiency_score": round(float(score), 2),
            "recommend_rank": 0
        })

    # Sort by score ascending (lowest = best candidate for assignment)
    recommendations.sort(key=lambda x: x["efficiency_score"])
    for i, rec in enumerate(recommendations):
        rec["recommend_rank"] = i + 1

    return recommendations

@router.post("/scan-escalations")
async def scan_escalations(background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db)):
    """
    Scans for cases that have crossed the Risk threshold and dispatches alerts.
    """
    # 1. Fetch all cases in transition (Verification or QC)
    stmt = select(models.Case).options(selectinload(models.Case.checks)).filter(models.Case.status.in_([enums.CaseStatus.VERIFICATION, enums.CaseStatus.QC]))
    res = await db.execute(stmt)
    cases = res.scalars().all()
    
    escalation_count: int = 0
    assigned_user_ids = set()
    
    for c in cases:
        if not c.received_date or not c.assigned_to:
            continue
            
        # Calculate Predictive TAT
        check_types = [chk.check_type for chk in c.checks]
        p_tat = tat_utils.calculate_predictive_tat(check_types)
        
        # Check if At Risk
        if tat_utils.check_is_at_risk(c.received_date, p_tat):
            # Dispatch Alert
            await notification_utils.notify_at_risk(
                db, c.assigned_to, 
                c.case_ref_no, 
                c.id, 
                p_tat,
                background_tasks=background_tasks
            )
            escalation_count = escalation_count + 1
            assigned_user_ids.add(c.assigned_to)
            
    await db.commit()
    return {"message": f"Scan complete. Dispatched {escalation_count} risk alerts to {len(assigned_user_ids)} team members."}

@router.post("/case/ping/{case_id}")
async def ping_case(case_id: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    """
    Manually ping a verifier for a specific case.
    """
    stmt = select(models.Case).filter(models.Case.id == case_id)
    res = await db.execute(stmt)
    case_obj = res.scalar_one_or_none()
    
    if not case_obj:
        raise HTTPException(404, detail="Case not found")
    
    if not case_obj.assigned_to:
        raise HTTPException(400, detail="Case is not assigned to any verifier")
        
    await notification_utils.notify_ping(
        db, 
        case_obj.assigned_to, 
        case_obj.case_ref_no, 
        case_obj.id, 
        current_user.full_name or current_user.email,
        background_tasks=background_tasks
    )
    
    await db.commit()
    return {"message": f"Urgent ping dispatched to verifier for Case {case_obj.case_ref_no}"}

@router.get("", response_model=List[schemas.CaseRead], dependencies=[Depends(check_module_permission("bvs", "verification", action="read"))])
async def read_cases(
    response: Response,
    status: Optional[str] = None, 
    batch_id: Optional[str] = None,
    customer_id: Optional[str] = None,
    customer_name: Optional[str] = None,
    search: Optional[str] = None,
    search_name: Optional[str] = None,
    search_ref: Optional[str] = None,
    assigned: Optional[bool] = None,
    assigned_to: Optional[str] = None,
    exclude_completed: Optional[bool] = None,
    skip: int = 0, 
    limit: int = 200, 
    sort: str = "received_date",
    order: str = "desc",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    at_risk: Optional[bool] = None,
    filter: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    # 1. Base query for cases with their relationships - Standardized Loading for Async
    stmt = select(models.Case).options(
        selectinload(models.Case.candidate),
        selectinload(models.Case.customer),
        selectinload(models.Case.batch),
        selectinload(models.Case.assigned_user).joinedload(models.User.role_rel),
        selectinload(models.Case.finalized_user),
        selectinload(models.Case.checks).options(
            selectinload(models.VerificationCheck.documents).joinedload(models.VerificationDocument.uploader),
            selectinload(models.VerificationCheck.logs).joinedload(models.VerificationLog.performer),
            selectinload(models.VerificationCheck.assigned_verifier),
            selectinload(models.VerificationCheck.finalized_user)
        ),
        selectinload(models.Case.verification_logs).joinedload(models.VerificationLog.performer),
        selectinload(models.Case.insufficiencies).options(joinedload(models.Insufficiency.check))
    )


    # Apply joins for filtering/sorting if needed (using selectinload for data, join for query)
    stmt = stmt.outerjoin(models.Case.candidate).outerjoin(models.Case.customer)


    # 2. Count for pagination - Optimized to avoid unnecessary joins
    base_count_stmt = select(func.count(models.Case.id))
    # Only join if we are filtering by fields in those tables
    if search or search_name or customer_name:
        base_count_stmt = base_count_stmt.outerjoin(models.Case.candidate).outerjoin(models.Case.customer)

    # Determine permission profile
    user_role_raw = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    # Normalize: strip enum class name if present (e.g., 'USERROLE.QA' -> 'QA')
    user_role_clean = user_role_raw.split('.')[-1]
    
    role_rel_name = (current_user.role_rel.name.upper() if current_user.role_rel else "").upper()
    
    oversight_keywords = ["SUPER_ADMIN", "ADMIN", "MANAGER", "QA", "QC", "SUPER ADMIN", "QC VERIFIER"]
    is_oversight = any(k in user_role_clean for k in oversight_keywords) or \
                   any(k in role_rel_name for k in oversight_keywords)
    
    is_customer = "CUSTOMER" in user_role_clean or "CUSTOMER" in role_rel_name

    # Combined filtering logic
    if is_customer:
        c_id = str(current_user.customer_id) if current_user.customer_id else None
        if c_id:
            stmt = stmt.filter(models.Case.customer_id == c_id)
            base_count_stmt = base_count_stmt.filter(models.Case.customer_id == c_id)
        else:
            # If a customer user has no associated customer_id, they should see no data
            stmt = stmt.filter(models.Case.id == "RESTRICTED")
            base_count_stmt = base_count_stmt.filter(models.Case.id == "RESTRICTED")
    elif not is_oversight:
        # For restricted users (Verifiers), show cases where they are involved in ANY capacity
        personal_filter = or_(
            models.Case.assigned_to == current_user.id,
            models.Case.qa_id == current_user.id,
            models.Case.qc_id == current_user.id
        )
        stmt = stmt.filter(personal_filter)
        base_count_stmt = base_count_stmt.filter(personal_filter)
    
    # 3. Dynamic Filtering
    ALL_FINAL_STATUSES = [
        'FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE',
        'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT',
        'QC_VERIFIED', 'CLOSED'
    ]
    if status and str(status).strip().upper() not in ['ALL', '']:
        status_up = str(status).strip().upper()
        if status_up == 'VERIFICATION':
            # WIP: strictly only active verification stage, not pending intake
            s_filter = models.Case.status == 'VERIFICATION'
        elif status_up == 'QC':
            # QC Active Pipeline: include all stages including verified but not yet finalized COMPLETED
            s_filter = models.Case.status.in_(['QC', 'QC_PENDING', 'QA_PENDING', 'QC_VERIFIED'])
        elif status_up in ('COMPLETED', 'FINALIZED'):
            # Treat COMPLETED/FINALIZED as an alias for ALL terminal verdict statuses
            s_filter = models.Case.status.in_(ALL_FINAL_STATUSES)
        else:
            s_filter = models.Case.status == status
            
        stmt = stmt.filter(s_filter)
        base_count_stmt = base_count_stmt.filter(s_filter)
    else:
        # Exclude old self-onboarding-only statuses (literal strings, not enum aliases,
        # because the enum aliases now resolve to ASSIGNED/IN_PROGRESS which would
        # incorrectly exclude active cases from all queries).
        exclude_filter = ~models.Case.status.in_([
            'LINK_SHARED', 'DOCUMENTS_SUBMITTED'
        ])
        stmt = stmt.filter(exclude_filter)
        base_count_stmt = base_count_stmt.filter(exclude_filter)
    if batch_id:
        stmt = stmt.filter(models.Case.batch_id == batch_id)
        base_count_stmt = base_count_stmt.filter(models.Case.batch_id == batch_id)
    if customer_id:
        stmt = stmt.filter(models.Case.customer_id == customer_id)
        base_count_stmt = base_count_stmt.filter(models.Case.customer_id == customer_id)
    if customer_name:
        stmt = stmt.filter(models.Customer.name == customer_name)
        base_count_stmt = base_count_stmt.filter(models.Customer.name == customer_name)
    if assigned_to:
        involvement_filter = or_(
            models.Case.assigned_to == assigned_to,
            models.Case.qa_id == assigned_to,
            models.Case.qc_id == assigned_to
        )
        stmt = stmt.filter(involvement_filter)
        base_count_stmt = base_count_stmt.filter(involvement_filter)
    if search:
        f = or_(models.Case.case_ref_no.ilike(f"%{search}%"), models.Candidate.name.ilike(f"{search}%"))
        stmt = stmt.filter(f)
        base_count_stmt = base_count_stmt.filter(f)
    if search_name:
        stmt = stmt.filter(models.Candidate.name.ilike(f"{search_name}%"))
        base_count_stmt = base_count_stmt.filter(models.Candidate.name.ilike(f"{search_name}%"))
    if search_ref:
        stmt = stmt.filter(models.Case.case_ref_no.ilike(f"%{search_ref}%"))
        base_count_stmt = base_count_stmt.filter(models.Case.case_ref_no.ilike(f"%{search_ref}%"))
    if assigned is not None:
        _FINAL_S = ['FINALIZED','COMPLETED','POSITIVE','NEGATIVE','DISCREPANCY','UNABLE TO VERIFY','HOLD','INSUFFICIENT','QC_VERIFIED','CLOSED']
        if assigned:
            # Active Allocations: cases that have an assigned verifier and are not yet finalized
            active_filter = and_(
                models.Case.assigned_to != None,
                models.Case.status.notin_(_FINAL_S)
            )
            stmt = stmt.filter(active_filter)
            base_count_stmt = base_count_stmt.filter(active_filter)
        else:
            # Unassigned: cases with no verifier and not yet finalized
            unassigned_filter = and_(
                models.Case.assigned_to == None,
                models.Case.status.notin_(_FINAL_S)
            )
            stmt = stmt.filter(unassigned_filter)
            base_count_stmt = base_count_stmt.filter(unassigned_filter)
    
    if exclude_completed:
        _EXCL = ['FINALIZED','COMPLETED','POSITIVE','NEGATIVE','DISCREPANCY','UNABLE TO VERIFY','HOLD','INSUFFICIENT','QC_VERIFIED','CLOSED']
        comp_filter = models.Case.status.notin_(_EXCL)
        stmt = stmt.filter(comp_filter)
        base_count_stmt = base_count_stmt.filter(comp_filter)

    if from_date:
        try:
            f_date = datetime.strptime(from_date, "%Y-%m-%d")
            stmt = stmt.filter(models.Case.received_date >= f_date)
            base_count_stmt = base_count_stmt.filter(models.Case.received_date >= f_date)
        except: pass
    if to_date:
        try:
            t_date = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            stmt = stmt.filter(models.Case.received_date <= t_date)
            base_count_stmt = base_count_stmt.filter(models.Case.received_date <= t_date)
        except: pass

    if at_risk is not None and at_risk:
        # Include both actual breaches and cases approaching breach
        # BUT exclude finalized/archived cases
        risk_threshold = datetime.utcnow() - timedelta(days=7)
        risk_filter = and_(
            models.Case.status.notin_(['COMPLETED', 'QC_VERIFIED']),
            or_(
                models.Case.is_in_tat == 0,
                models.Case.received_date < risk_threshold
            )
        )
        stmt = stmt.filter(risk_filter)
        base_count_stmt = base_count_stmt.filter(risk_filter)
    elif at_risk is not None and not at_risk:
        stmt = stmt.filter(models.Case.is_in_tat == 1)
        base_count_stmt = base_count_stmt.filter(models.Case.is_in_tat == 1)
    
    if filter == 'in_tat':
        stmt = stmt.filter(models.Case.is_in_tat == 1)
        base_count_stmt = base_count_stmt.filter(models.Case.is_in_tat == 1)
    elif filter == 'out_tat':
        stmt = stmt.filter(models.Case.is_in_tat == 0)
        base_count_stmt = base_count_stmt.filter(models.Case.is_in_tat == 0)
    


    total_count_res = await db.execute(base_count_stmt)
    total_count = total_count_res.scalar() or 0
    response.headers["X-Total-Count"] = str(total_count)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Count"

    # 4. Sorting logic
    sort_attr = getattr(models.Case, sort, None)
    if sort_attr is None:
        sort_attr = models.Case.received_date
        
    if order == "desc":
        stmt = stmt.order_by(sort_attr.desc())
    else:
        stmt = stmt.order_by(sort_attr.asc())
        
    stmt = stmt.offset(skip).limit(limit)
    res = await db.execute(stmt)
    cases_models = res.unique().scalars().all()
    
    # 3. Transform to CaseRead format
    cases_read = []
    for case in cases_models:
        case_data = schemas.CaseRead.model_validate(case)
        if case.candidate: case_data.candidate_name = case.candidate.name
        if case.customer: case_data.customer_name = case.customer.name
        if case.batch:
            if not case_data.tat_days: case_data.tat_days = case.batch.tat_days
            case_data.batch_date = case.batch.upload_date
            case_data.batch_no = case.batch.batch_no
        if case.assigned_user: 
            case_data.assigned_user_name = case.assigned_user.full_name
            r_enum_val = str(case.assigned_user.role.value if hasattr(case.assigned_user.role, 'value') else case.assigned_user.role).upper()
            role_name = case.assigned_user.role_rel.name if case.assigned_user.role_rel else ("QC Verifier" if r_enum_val in ["QA", "QC"] else r_enum_val)
            case_data.assigned_user_role = role_name.upper()
        if case.qa_user: case_data.qa_user_name = case.qa_user.full_name
        if case.qc_user: case_data.qc_user_name = case.qc_user.full_name
        
        # Calculate queue age
        if case.received_date:
            age = datetime.utcnow() - case.received_date
            case_data.queue_age = f"{int(age.total_seconds() // 3600)}h"
        else:
            case_data.queue_age = "0h"
        
        # Add check_name to insufficiencies and populate nested log names
        if hasattr(case, 'insufficiencies'):
            for i, insuff in enumerate(case.insufficiencies):
                if insuff.check:
                    case_data.insufficiencies[i].check_name = insuff.check.check_type
        
        # Populate nested names for logs and documents
        if case_data.verification_logs:
            for i, log in enumerate(case.verification_logs):
                if log.performer:
                    case_data.verification_logs[i].performer_name = log.performer.full_name

        for i, check_model in enumerate(case.checks):
            # Populate logs in check
            if check_model.logs:
                for j, log_model in enumerate(check_model.logs):
                    if log_model.performer:
                        case_data.checks[i].logs[j].performer_name = log_model.performer.full_name
            # Populate documents in check
            if check_model.documents:
                for j, doc_model in enumerate(check_model.documents):
                    if doc_model.uploader:
                        case_data.checks[i].documents[j].uploaded_by_name = doc_model.uploader.full_name

        
        # Calculate In-TAT/Out-TAT
        if case.received_date:
            from datetime import timezone
            now_dt = datetime.now(timezone.utc)
            
            r_date = case.received_date
            if r_date.tzinfo is None:
                r_date = r_date.replace(tzinfo=timezone.utc)
                
            e_date = case.completed_date or now_dt
            if e_date.tzinfo is None:
                e_date = e_date.replace(tzinfo=timezone.utc)
                
            diff_seconds = (e_date - r_date).total_seconds()
            # Use inclusive calendar days for TAT
            total_days = (e_date.date() - r_date.date()).days + 1
            total_days = max(1, total_days)
            
            allowed = case_data.tat_days or 10
            if total_days <= allowed:
                case_data.in_tat = total_days
                case_data.out_tat = 0
            else:
                case_data.in_tat = allowed
                case_data.out_tat = total_days - allowed
            
            # Predictive TAT Analysis
            check_types = [c.check_type for c in case.checks]
            p_tat = tat_utils.calculate_predictive_tat(check_types)
            case_data.predicted_tat = p_tat
            # Suppress risk alerts for archived protocols
            if str(case.status).upper() in ["COMPLETED", "QC_VERIFIED"]:
                case_data.is_at_risk = False
            else:
                case_data.is_at_risk = tat_utils.check_is_at_risk(case.received_date, p_tat)
            
        
        cases_read.append(case_data)
    
    return cases_read

@router.get("/clients", response_model=List[str], dependencies=[Depends(check_module_permission("bvs", "verification", action="read"))])
async def read_case_clients(db: AsyncSession = Depends(get_async_db)):
    stmt = select(models.Customer.name).distinct().join(models.Case)
    res = await db.execute(stmt)
    return [r for r in res.scalars().all() if r]

@router.get("/report-stats", dependencies=[Depends(check_module_permission("bvs", "verification", action="read"))])
async def get_report_stats(customer_id: Optional[str] = None, db: AsyncSession = Depends(get_async_db)):
    # 1. Pie Data: Status distribution
    stmt = select(models.Case.status, func.count(models.Case.id)).group_by(models.Case.status)
    if customer_id: stmt = stmt.filter(models.Case.customer_id == customer_id)
    res = await db.execute(stmt)
    pie_data = [{"name": str(s), "value": count} for s, count in res.all()]

    # 2. Aggregates
    base_stmt = select(models.Case)
    if customer_id: base_stmt = base_stmt.filter(models.Case.customer_id == customer_id)
    
    total_res = await db.execute(select(func.count(models.Case.id)).select_from(base_stmt.subquery()))
    total = total_res.scalar() or 0
    
    comp_res = await db.execute(select(func.count(models.Case.id)).filter(models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT'])).select_from(base_stmt.subquery()))
    completed = comp_res.scalar() or 0
    
    tat_res = await db.execute(select(func.avg(models.Case.tat_days)).filter(models.Case.status.in_(['FINALIZED', 'COMPLETED', 'POSITIVE', 'NEGATIVE', 'DISCREPANCY', 'UNABLE TO VERIFY', 'HOLD', 'INSUFFICIENT'])).select_from(base_stmt.subquery()))
    avg_tat = tat_res.scalar() or 0

    return {
        "pie_data": pie_data,
        "total_cases": total,
        "completion_rate": float(round(float(completed / total * 100), 1)) if total > 0 else 0.0,
        "avg_tat": float(round(float(avg_tat or 0), 1))
    }

@router.get("/{case_id}", response_model=schemas.CaseRead, dependencies=[Depends(check_module_permission("bvs", "verification", action="read"))])
async def read_case(case_id: str, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    stmt = select(models.Case).options(
        joinedload(models.Case.candidate),
        joinedload(models.Case.customer),
        selectinload(models.Case.checks).options(
            selectinload(models.VerificationCheck.documents).joinedload(models.VerificationDocument.uploader),
            selectinload(models.VerificationCheck.logs).joinedload(models.VerificationLog.performer),
            selectinload(models.VerificationCheck.assigned_verifier),
            selectinload(models.VerificationCheck.finalized_user)
        ),
        selectinload(models.Case.verification_logs).joinedload(models.VerificationLog.performer),
        joinedload(models.Case.batch),
        joinedload(models.Case.assigned_user).joinedload(models.User.role_rel),
        joinedload(models.Case.finalized_user),
        selectinload(models.Case.insufficiencies).options(joinedload(models.Insufficiency.check))
    ).filter(models.Case.id == case_id)
    res = await db.execute(stmt)
    db_case = res.unique().scalar_one_or_none()
    
    if db_case is None:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Tenancy/Isolation check
    user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    if (user_role == "CUSTOMER" or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER")) and db_case.customer_id != current_user.customer_id:
        raise HTTPException(status_code=403, detail="Unauthorized access to this case")
    if current_user.role == models.UserRole.VERIFIER and db_case.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="Unauthorized access to this case")
        
    # Aggregated document collection from address_details and direct uploads
    if db_case.candidate and db_case.candidate.address_details:
        current_docs = db_case.candidate.documents or []
        existing_urls = {d.get('url') for d in current_docs if isinstance(d, dict) and d.get('url')}
        
        all_docs = list(current_docs)
        doc_mapping = {
            'addresses': 'Address',
            'employments': 'Employment',
            'educations': 'Education',
            'identities': 'Identity',
            'references': 'Reference',
            'criminal_records': 'Criminal',
            'drug_tests': 'Drug Test',
            'cibil_checks': 'CIBIL',
            'global_database_checks': 'Global Database'
        }
        
        addr_details = db_case.candidate.address_details
        for key, check_label in doc_mapping.items():
            section_items = addr_details.get(key, [])
            if isinstance(section_items, list):
                for item in section_items:
                    if isinstance(item, dict) and 'files' in item and isinstance(item['files'], list):
                        for f in item['files']:
                            if isinstance(f, dict) and f.get('url') and f.get('url') not in existing_urls:
                                all_docs.append({
                                    'url': f['url'],
                                    'original_filename': f.get('original_filename') or f.get('file_name') or f.get('name', 'Supporting Document'),
                                    'check_type': check_label,
                                    'uploaded_at': datetime.utcnow().isoformat(),
                                    'is_primary': True,
                                    'public_id': f.get('public_id') or f.get('path')
                                })
                                existing_urls.add(f['url'])
        db_case.candidate.documents = all_docs

    # Convert to Pydantic and populate metadata
    case_data = schemas.CaseRead.model_validate(db_case)
    
    # Populate metadata
    if db_case.candidate: case_data.candidate_name = db_case.candidate.name
    if db_case.customer: case_data.customer_name = db_case.customer.name
    if db_case.batch:
        case_data.batch_no = db_case.batch.batch_no
        case_data.batch_date = db_case.batch.upload_date
        if not case_data.tat_days:
            case_data.tat_days = db_case.batch.tat_days
    
    if db_case.assigned_user: 
        case_data.assigned_user_name = db_case.assigned_user.full_name
        r_enum_val = str(db_case.assigned_user.role.value if hasattr(db_case.assigned_user.role, 'value') else db_case.assigned_user.role).upper()
        role_name = db_case.assigned_user.role_rel.name if db_case.assigned_user.role_rel else ("QC Verifier" if r_enum_val in ["QA", "QC"] else r_enum_val)
        case_data.assigned_user_role = role_name.upper()
    
    if db_case.qc_user: case_data.qc_user_name = db_case.qc_user.full_name

    # Add check_name to insufficiencies
    if hasattr(db_case, 'insufficiencies'):
        for i, insuff in enumerate(db_case.insufficiencies):
            if insuff.check:
                case_data.insufficiencies[i].check_name = insuff.check.check_type
    
    # Predict TAT if not provided
    check_types = [chk.check_type for chk in db_case.checks]
    case_data.predicted_tat = tat_utils.calculate_predictive_tat(check_types)
    if not case_data.tat_days:
        case_data.tat_days = case_data.predicted_tat or 10

    # Calculate In-TAT/Out-TAT for single view
    if db_case.received_date:
        from datetime import timezone
        now_dt = datetime.now(timezone.utc)
        
        r_date = db_case.received_date
        if r_date.tzinfo is None:
            r_date = r_date.replace(tzinfo=timezone.utc)
            
        e_date = db_case.completed_date or now_dt
        if e_date.tzinfo is None:
            e_date = e_date.replace(tzinfo=timezone.utc)
            
        diff_seconds = (e_date - r_date).total_seconds()
        # Use inclusive calendar days for TAT
        total_days = (e_date.date() - r_date.date()).days + 1
        total_days = max(1, total_days)
        allowed = case_data.tat_days or 10
        
        if total_days <= allowed:
            case_data.in_tat = total_days
            case_data.out_tat = 0
        else:
            case_data.in_tat = allowed
            case_data.out_tat = total_days - allowed

    return case_data

@router.post("/face-match")
async def face_match(req: schemas.FaceMatchRequest, current_user: models.User = Depends(get_current_user)):
    url1 = req.url1  # ID Photo
    url2 = req.url2  # Profile/Selfie Photo
    
    if not url1 or not url2:
        return {"success": False, "message": "Missing URLs"}
        
    try:
        from .ocr_utils import get_scanner
        
        scanner = get_scanner()
        
        # Download images asynchronously
        async with httpx.AsyncClient() as client:
            r1, r2 = await asyncio.gather(
                client.get(url1, timeout=10.0),
                client.get(url2, timeout=10.0)
            )
        
        if r1.status_code != 200 or r2.status_code != 200:
            return {"success": False, "message": "Failed to download images"}
            
        face1 = scanner.get_face(r1.content)
        face2 = scanner.get_face(r2.content)
        
        if face1 is None: return {"success": False, "message": "No face detected in Image 1"}
        if face2 is None: return {"success": False, "message": "No face detected in Image 2"}
        
        score = scanner.match_faces(face1, face2)
        
        return {
            "success": True,
            "match_score": round(score, 2),
            "is_match": score > 60, # Threshold for match
            "message": "Match successful" if score > 60 else "Potential mismatch"
        }
    except Exception as e:
        return {"success": False, "message": str(e)}

@router.post("/bulk-action", dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
async def bulk_action(req: schemas.BulkActionRequest, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    if not req.case_ids:
        return {"msg": "No cases provided"}
        
    update_data: Dict[str, Any] = {}
    if req.action == "assign":
        update_data["assigned_to"] = req.target_value
        update_data["assigned_at"] = datetime.utcnow()
    elif req.action == "status":
        update_data["status"] = req.target_value
        if req.target_value == models.CaseStatus.PENDING:
            # When moving back to Pending (e.g. Cross-Check), clear current assignment
            update_data["assigned_to"] = None
            update_data["assigned_at"] = None
        if req.target_value == models.CaseStatus.COMPLETED:
            update_data["completed_date"] = datetime.utcnow()
            # Attribute audit/completion to the actor
            if current_user.role == models.UserRole.QC:
                update_data["qc_id"] = current_user.id
            if current_user.role == models.UserRole.QA:
                update_data["qa_id"] = current_user.id
            
    if not update_data:
        return {"msg": "Invalid action"}
        
    from .ws import manager
    # Revoke Logic for Bulk
    if req.action == "status":
        for cid in req.case_ids:
            res_c = await db.execute(
                select(models.Case)
                .options(selectinload(models.Case.checks))
                .filter(models.Case.id == cid)
            )
            case_obj = res_c.scalar_one_or_none()
            if not case_obj: continue

            # Validation for completion
            if req.target_value == models.CaseStatus.COMPLETED:
                validate_case_completion(case_obj)

            
            # QC Revoke: Moving back from QC or Completed
            if (case_obj.status in [models.CaseStatus.QC, models.CaseStatus.COMPLETED]) and req.target_value in [models.CaseStatus.VERIFICATION, models.CaseStatus.PENDING]:
                case_obj.qc_revoke_count += 1
                db.add(models.RevokeLog(
                    case_id=cid,
                    user_id=current_user.id,
                    revoke_type='QC',
                    from_status=case_obj.status,
                    to_status=req.target_value
                ))
            
            # QA/Lead Revoke: Moving back from QA to QC or Verifier
            if case_obj.status == models.CaseStatus.QA_PENDING and req.target_value in [models.CaseStatus.QC, models.CaseStatus.VERIFICATION, models.CaseStatus.PENDING]:
                # We'll treat QA-initiated revokes as QC revokes for the purpose of the current tracking dashboard
                case_obj.qc_revoke_count += 1
                db.add(models.RevokeLog(
                    case_id=cid,
                    user_id=current_user.id,
                    revoke_type='QC',
                    from_status=case_obj.status,
                    to_status=req.target_value
                ))
                
            # Insufficiency Logic: Marking as insufficient
            if req.target_value == models.CaseStatus.INSUFFICIENT and case_obj.status != models.CaseStatus.INSUFFICIENT:
                case_obj.insufficiency_count += 1
                db.add(models.InsufficiencyLog(
                    case_id=cid,
                    user_id=current_user.id,
                    from_status=case_obj.status,
                    notes="Bulk marked as insufficient"
                ))
            
            # Resolution Logic: Moving out of insufficient
            if case_obj.status == models.CaseStatus.INSUFFICIENT and req.target_value != models.CaseStatus.INSUFFICIENT:
                # Update latest log with resolution time
                res_log_stmt = select(models.InsufficiencyLog).filter(models.InsufficiencyLog.case_id == cid).order_by(models.InsufficiencyLog.marked_at.desc())
                res_log_exec = await db.execute(res_log_stmt)
                latest_log = res_log_exec.scalars().first()
                if latest_log and not latest_log.resolved_at:
                    latest_log.resolved_at = datetime.utcnow()

            # Update status and dates
            case_obj.status = req.target_value
            if req.target_value == models.CaseStatus.COMPLETED:
                case_obj.completed_date = datetime.utcnow()
            else:
                case_obj.completed_date = None
            
            if req.target_value == models.CaseStatus.COMPLETED:
                # TAT Logic
                if case_obj.batch_id:
                    res_batch = await db.execute(select(models.Batch).filter(models.Batch.id == case_obj.batch_id))
                    batch = res_batch.scalar_one_or_none()
                    if batch and case_obj.received_date:
                        diff = (datetime.utcnow() - case_obj.received_date).days
                        case_obj.is_in_tat = 1 if diff <= (batch.tat_days or 10) else 0

    await db.commit()

    verifier_name = "Verifier"
    if req.action == "assign" and req.target_value:
        v_res = await db.execute(select(models.User).filter(models.User.id == req.target_value))
        v_obj = v_res.scalar_one_or_none()
        if v_obj: verifier_name = v_obj.full_name

    # Trigger Notifications for Bulk
    bulk_case_info = []
    for cid in req.case_ids:
        # Re-fetch for reference info with candidate
        ref_res = await db.execute(select(models.Case).options(joinedload(models.Case.candidate)).filter(models.Case.id == cid))
        cbd = ref_res.scalar_one_or_none()
        if not cbd: continue
        candidate_name = cbd.candidate.name if cbd.candidate else "Candidate"
        
        # Track for admin summary
        bulk_case_info.append({"id": cid, "ref": cbd.case_ref_no, "candidate": candidate_name})

        if req.action == "assign" and req.target_value:
             # Standard assignment logic...
             update_stmt = update(models.Case).where(models.Case.id == cid).values(assigned_to=req.target_value, assigned_at=datetime.utcnow() if not cbd.assigned_at else cbd.assigned_at)
             await db.execute(update_stmt)
             await create_audit_log(db, current_user.id, "CASE_ASSIGNMENT", f"Case bulk-assigned to verifier via system operator", resource_id=cid)
             await manager.broadcast({"type": "CASE_UPDATED", "case_id": cid, "action": "bulk-assignment", "assigned_to": req.target_value})
             await notification_utils.notify_new_assignment(db, req.target_value, cbd.case_ref_no, cid, candidate_name)
        elif req.action == "status":
            old_s = cbd.status.value if hasattr(cbd.status, 'value') else cbd.status
            new_s = req.target_value.value if hasattr(req.target_value, 'value') else req.target_value
            await create_audit_log(db, current_user.id, "STATUS_CHANGE", f"Case status bulk-updated from {old_s} to {new_s}", resource_id=cid)
            if req.target_value == models.CaseStatus.INSUFFICIENT:
                await notification_utils.notify_insufficient(db, cbd.assigned_to, cbd.case_ref_no, cid)
            elif req.target_value == models.CaseStatus.COMPLETED:
                await notification_utils.notify_case_closed(db, cid, cbd.case_ref_no, candidate_name)
            elif req.target_value == models.CaseStatus.QC:
                await notification_utils.notify_verification_completed(db, cid, cbd.case_ref_no, candidate_name, current_user.full_name)

    # Notify Admin ONE summary instead of many small ones
    if req.action == "assign" and req.target_value:
        await notification_utils.create_notification(
            db, current_user.id,
            "Bulk Allocation Finalized",
            f"Successfully deployed {len(req.case_ids)} cases to {verifier_name}. All mission protocols are now active in the verifier queue.",
            enums.NotificationCategory.SYSTEM_ALERT,
            extra_data={"type": "BULK_ASSIGN", "verifier": verifier_name, "cases": bulk_case_info}
        )
    
    await db.commit()
    await invalidate_dashboard_cache()
    
    # Trigger Real-time Workforce Update
    await manager.broadcast({"type": "WORKFORCE_UPDATE", "source": "bulk_allocation"})
    
    return {"message": "Bulk allocation successful"}
    
    # Audit log and broadcast
@router.post("/auto-allocate")
async def auto_allocate(req: schemas.BulkActionRequest, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    if not req.case_ids:
        return {"msg": "No cases provided"}
        
    # 1. Fetch available verifiers
    res_users = await db.execute(select(models.User).filter(models.User.role == models.UserRole.VERIFIER, models.User.status == "ACTIVE"))
    verifiers = res_users.scalars().all()
    
    if not verifiers:
        return {"msg": "No active verifiers found", "success": False}
        
    # 2. Get current workloads — ONE aggregate query instead of N
    load_stmt = select(
        models.Case.assigned_to,
        func.count(models.Case.id)
    ).filter(
        models.Case.status != models.CaseStatus.COMPLETED,
        models.Case.assigned_to.isnot(None)
    ).group_by(models.Case.assigned_to)
    load_res = await db.execute(load_stmt)
    load_map = {row[0]: row[1] for row in load_res.all()}
    
    workloads = {v.id: load_map.get(v.id, 0) for v in verifiers}
    v_name_map = {v.id: v.full_name for v in verifiers}

    # 3. Pre-fetch all case data in ONE query instead of per-case
    cases_stmt = select(models.Case).options(
        joinedload(models.Case.candidate)
    ).filter(models.Case.id.in_(req.case_ids))
    cases_res = await db.execute(cases_stmt)
    cases_map = {c.id: c for c in cases_res.scalars().unique().all()}

    # 4. Assign cases greedily to verifier with least cases
    assigned_count = 0
    auto_assigned_info = []
    
    for cid in req.case_ids:
        # Find verifier with minimum workload
        target_v_id = min(workloads, key=workloads.get)
        
        # Update case
        await db.execute(
            update(models.Case).where(models.Case.id == cid).values(
                assigned_to=target_v_id,
                assigned_at=datetime.utcnow(),
                status=models.CaseStatus.VERIFICATION
            )
        )
        
        # Increment workload for next iteration
        workloads[target_v_id] += 1
        assigned_count += 1
        
        # Audit log and broadcast
        await create_audit_log(db, current_user.id, "AUTO_ALLOCATION", f"Case automatically assigned to verifier", resource_id=cid)
        await manager.broadcast({"type": "CASE_UPDATED", "case_id": cid, "action": "auto-assignment"})
        
        # Use pre-fetched case data for notification
        cbd = cases_map.get(cid)
        candidate_name = cbd.candidate.name if cbd and cbd.candidate else "Candidate"
        v_name = v_name_map.get(target_v_id, "Verifier")
            
        auto_assigned_info.append({"id": cid, "ref": cbd.case_ref_no if cbd else cid, "candidate": candidate_name, "verifier": v_name})
        await notification_utils.notify_new_assignment(db, target_v_id, cbd.case_ref_no if cbd else cid, cid, candidate_name)

    # Notify Admin ONE summary
    await notification_utils.create_notification(
        db, current_user.id,
        "Auto-Allocation Protocol Executed",
        f"Distributed {assigned_count} cases across active operational units. System load balancing complete.",
        enums.NotificationCategory.SYSTEM_ALERT,
        extra_data={"type": "AUTO_ALLOCATE", "cases": auto_assigned_info}
    )
    
    await db.commit()
    await invalidate_dashboard_cache()
    return {"msg": f"Successfully auto-allocated {assigned_count} cases", "success": True}

@router.patch("/{case_id}", response_model=schemas.Case, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
async def update_case(case_id: str, case_update: schemas.CaseUpdate, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    stmt = select(models.Case).options(
        selectinload(models.Case.checks),
        joinedload(models.Case.candidate)
    ).filter(models.Case.id == case_id)
    res = await db.execute(stmt)
    db_case = res.scalar_one_or_none()
    if db_case is None:
        raise HTTPException(status_code=404, detail="Case not found")
    
    old_status = db_case.status
    
    update_data = case_update.dict(exclude_unset=True)
    candidate_update_data = update_data.pop("candidate", None)
    services_update = update_data.pop("services", None)
    rates_update = update_data.pop("check_rates", {})
    scope_of_work = update_data.pop("scope_of_work", None)
    check_scopes = update_data.pop("check_scopes", None) or {}

    manual_received_date = update_data.get("received_date")
    manual_completed_date = update_data.get("completed_date")

    if update_data.get("status") == models.CaseStatus.COMPLETED and db_case.status != models.CaseStatus.COMPLETED:
        validate_case_completion(db_case)
        db_case.completed_date = manual_completed_date or datetime.utcnow()
        # Propagate completion to checks
        for chk in db_case.checks:
            if str(chk.status).upper() in ["QC_PENDING", "VERIFICATION", "INTERIM"]:
                if chk.final_result:
                    res_val = str(chk.final_result).upper()
                    if res_val == "POSITIVE":
                        chk.status = "GREEN"
                    elif res_val == "NEGATIVE":
                        chk.status = "RED"
                    elif res_val == "DISCREPANCY":
                        chk.status = "AMBER"
                    else:
                        chk.status = res_val
                else:
                    chk.status = "GREEN"
        # TAT Performance tracking
        if db_case.batch_id and db_case.received_date:
            res_batch = await db.execute(select(models.Batch).filter(models.Batch.id == db_case.batch_id))
            batch = res_batch.scalar_one_or_none()
            if batch:
                diff = (datetime.utcnow() - db_case.received_date).days
                db_case.is_in_tat = 1 if diff <= (batch.tat_days or 10) else 0
        
        # Attribute completion credit...
        if current_user.role == models.UserRole.QC:
            db_case.qc_id = current_user.id
        elif current_user.role == models.UserRole.QA:
            db_case.qa_id = current_user.id
    elif update_data.get("status") in [models.CaseStatus.QA_PENDING, models.CaseStatus.QC_PENDING] and db_case.status not in [models.CaseStatus.QA_PENDING, models.CaseStatus.QC_PENDING]:
         if current_user.role in [models.UserRole.QA, models.UserRole.QC]:
            db_case.qa_id = current_user.id
         # Auto-propagate to checks
         for chk in db_case.checks:
            if chk.status in [models.CheckStatus.VERIFICATION, models.CheckStatus.INTERIM]:
                chk.status = models.CheckStatus.QC_PENDING
    elif update_data.get("status") and update_data.get("status") != models.CaseStatus.COMPLETED:
        # Revoke Logic for Single Update
        # QC Revoke: from QC/QA/Completed to Verification/Pending
        if (db_case.status in [models.CaseStatus.QC, models.CaseStatus.QA_PENDING, models.CaseStatus.COMPLETED]) and update_data.get("status") in [models.CaseStatus.VERIFICATION, models.CaseStatus.PENDING]:
            db_case.qc_revoke_count += 1
            db.add(models.RevokeLog(
                case_id=case_id,
                user_id=current_user.id,
                revoke_type='QC',
                from_status=db_case.status,
                to_status=update_data.get('status')
            ))
            # Reset checks to WIP so verifier can re-verify
            for chk in db_case.checks:
                if chk.status == models.CheckStatus.QC_PENDING:
                    chk.status = models.CheckStatus.VERIFICATION
            
        # Transition Revoke: from Completed back to QC/QA
        if db_case.status == models.CaseStatus.COMPLETED and update_data.get("status") in [models.CaseStatus.QC, models.CaseStatus.QA_PENDING]:
            db_case.qc_revoke_count += 1
            db.add(models.RevokeLog(
                case_id=case_id,
                user_id=current_user.id,
                revoke_type='QC',
                from_status=db_case.status,
                to_status=update_data.get('status')
            ))

        # Insufficiency Logic: Marking as insufficient
        if update_data.get("status") == models.CaseStatus.INSUFFICIENT and db_case.status != models.CaseStatus.INSUFFICIENT:
            db_case.insufficiency_count += 1
            db.add(models.InsufficiencyLog(
                case_id=case_id,
                user_id=current_user.id,
                from_status=db_case.status,
                notes=update_data.get("notes", "Marked as insufficient")
            ))
        
        # Resolution Logic: Moving out of insufficient
        if db_case.status == models.CaseStatus.INSUFFICIENT and update_data.get("status") and update_data.get("status") != models.CaseStatus.INSUFFICIENT:
            # Update latest log with resolution time
            res_log_stmt = select(models.InsufficiencyLog).filter(models.InsufficiencyLog.case_id == case_id).order_by(models.InsufficiencyLog.marked_at.desc())
            res_log_exec = await db.execute(res_log_stmt)
            latest_log = res_log_exec.scalars().first()
            if latest_log and not latest_log.resolved_at:
                latest_log.resolved_at = datetime.utcnow()

        if manual_completed_date is None:
            db_case.completed_date = None
        else:
            db_case.completed_date = manual_completed_date

    if update_data.get("assigned_to") and not db_case.assigned_at:
        db_case.assigned_at = datetime.utcnow()

    for key, value in update_data.items():
        if getattr(db_case, key) != value:
            # Simple audit log for status changes
            if key == "status":
                old_val = db_case.status.value if hasattr(db_case.status, 'value') else db_case.status
                new_val = value.value if hasattr(value, 'value') else value
                await create_audit_log(db, current_user.id, "STATUS_CHANGE", f"Case status updated from {old_val} to {new_val}", resource_id=case_id)
            elif key == "assigned_to":
                await create_audit_log(db, current_user.id, "CASE_ASSIGNMENT", f"Case assigned to user ID {value}", resource_id=case_id)
        setattr(db_case, key, value)
    
    # Auto-transition from DOCS_SUBMITTED/LINK_SHARED to PENDING upon data entry complete
    if not update_data.get("status") and db_case.status in [models.CaseStatus.DOCUMENTS_SUBMITTED, models.CaseStatus.LINK_SHARED]:
        if candidate_update_data or services_update is not None:
            db_case.status = models.CaseStatus.PENDING
    
    # Update or Create Candidate
    if candidate_update_data:
        if db_case.candidate_id:
            res_cand = await db.execute(select(models.Candidate).filter(models.Candidate.id == db_case.candidate_id))
            db_candidate = res_cand.scalar_one_or_none()
            if db_candidate:
                for key, value in candidate_update_data.items():
                    if value is not None:
                        setattr(db_candidate, key, value)
        else:
            # Create new candidate if missing
            db_candidate = models.Candidate(**candidate_update_data)
            db.add(db_candidate)
            await db.flush()
            db_case.candidate_id = db_candidate.id

    # Sync Services/Checks & Scope of Work
    
    if services_update is not None:
        # 1. Get existing checks
        existing_checks_res = await db.execute(select(models.VerificationCheck).filter(models.VerificationCheck.case_id == case_id))
        existing_checks = {c.check_type: c for c in existing_checks_res.scalars().all()}
        
        # 2. Add or Update checks
        for svc in services_update:
            rate = rates_update.get(svc, 0.0)
            svc_scope = check_scopes.get(svc, scope_of_work)
            
            if svc in existing_checks:
                # Update rate if provided
                existing_checks[svc].rate = rate
                # Also update scope_of_work if provided (either per-check or global)
                if svc_scope is not None:
                    chk_data = existing_checks[svc].data
                    updated_data = dict(chk_data) if isinstance(chk_data, dict) else {}
                    updated_data["scope_of_work"] = svc_scope
                    existing_checks[svc].data = updated_data
            else:
                # Create new
                new_check = models.VerificationCheck(
                    case_id=case_id,
                    check_type=svc,
                    status=models.CheckStatus.INTERIM,
                    rate=rate,
                    data={"scope_of_work": svc_scope} if svc_scope else {}
                )
                db.add(new_check)
        
        # 3. Remove checks not in services_update
        for svc_type in list(existing_checks.keys()):
            if svc_type not in services_update:
                await db.delete(existing_checks[svc_type])
    elif check_scopes or scope_of_work:
        # Update scope for all existing checks if services list not provided
        existing_checks_res = await db.execute(select(models.VerificationCheck).filter(models.VerificationCheck.case_id == case_id))
        for chk in existing_checks_res.scalars().all():
            svc_scope = check_scopes.get(chk.check_type, scope_of_work)
            if svc_scope is not None:
                if chk.data is None: chk.data = {}
                chk.data["scope_of_work"] = svc_scope
    
    # 4. Trigger Notifications (Enhanced Stakeholder Flow)
    candidate_name = db_case.candidate.name if db_case.candidate else "Candidate"
    
    if "status" in update_data:
        new_status = update_data["status"]
        
        # Auto-set completion metadata
        if str(new_status).upper() == "COMPLETED":
            if not db_case.completed_date:
                db_case.completed_date = datetime.utcnow()
            if not db_case.qc_id:
                db_case.qc_id = current_user.id
        
        # Scenario 1: Verifier finishes -> Moves to QC
        if old_status == models.CaseStatus.VERIFICATION and new_status == models.CaseStatus.QC:
            await notification_utils.notify_verification_completed(db, case_id, db_case.case_ref_no, candidate_name, current_user.full_name)
        
        # Scenario 2: QC finishes -> Moves to Completed
        elif str(old_status).upper() != "COMPLETED" and str(new_status).upper() == "COMPLETED":
            # Capture final result for notification (checking both result and status fields)
            f_result = update_data.get("final_report_status") or update_data.get("final_result") or db_case.final_report_status or "POSITIVE"
            
            # Sync final result status to checks
            for chk in db_case.checks:
                if str(chk.status).upper() in ["QC_PENDING", "VERIFICATION", "INTERIM"]:
                    if chk.final_result:
                        res_val = str(chk.final_result).upper()
                        if res_val == "POSITIVE":
                            chk.status = "GREEN"
                        elif res_val == "NEGATIVE":
                            chk.status = "RED"
                        elif res_val == "DISCREPANCY":
                            chk.status = "AMBER"
                        else:
                            chk.status = res_val
                    else:
                        chk.status = "GREEN"

            # Ensure final_result column is also synced
            db_case.final_result = f_result
            db_case.final_report_status = f_result
            db_case.completed_date = datetime.utcnow() # Force set completion date here too for safety
            
            # Scenario 2a: If coming from QC Review
            if old_status == models.CaseStatus.QC:
                await notification_utils.notify_qc_completed(db, case_id, db_case.case_ref_no, candidate_name, current_user.full_name)
            
            # Final Closure Notification with Result
            await notification_utils.notify_case_closed(db, case_id, db_case.case_ref_no, candidate_name, final_result=f_result)

        # Scenario 4: Insufficient Flags
        elif new_status == models.CaseStatus.INSUFFICIENT:
            await notification_utils.notify_insufficient(db, db_case.assigned_to, db_case.case_ref_no, case_id)

    if "assigned_to" in update_data and update_data["assigned_to"]:
        await notification_utils.notify_new_assignment(db, update_data["assigned_to"], db_case.case_ref_no, case_id, candidate_name)

    await db.commit()
    res = await db.execute(
        select(models.Case).options(
            joinedload(models.Case.candidate),
            joinedload(models.Case.customer),
            selectinload(models.Case.checks).options(
                selectinload(models.VerificationCheck.documents).joinedload(models.VerificationDocument.uploader),
                selectinload(models.VerificationCheck.logs).joinedload(models.VerificationLog.performer),
                selectinload(models.VerificationCheck.assigned_verifier),
                selectinload(models.VerificationCheck.finalized_user)
            )
        ).filter(models.Case.id == db_case.id)
    )
    db_case = res.unique().scalar_one()

    # Trigger Background Summary Refresh
    from .stats_service import refresh_dashboard_summary
    background_tasks.add_task(refresh_dashboard_summary, db, db_case.customer_id)
    
    return db_case

@router.get("/{resource_id}/verification-logs", response_model=List[schemas.AuditLogRead], dependencies=[Depends(check_module_permission("bvs", "verification", action="read"))])
async def get_verification_logs(resource_id: str, db: AsyncSession = Depends(get_async_db)):
    # Note: I used DATABASE_get_async_db to avoid conflict with local variable if any, 
    # but the import is actually get_async_db in this file. Correcting to get_async_db.
    stmt = (
        select(models.AuditLog, models.User.full_name.label("user_full_name"))
        .outerjoin(models.User, models.AuditLog.user_id == models.User.id)
        .filter(models.AuditLog.resource_id == resource_id)
        .order_by(models.AuditLog.timestamp.desc())
    )
    res = await db.execute(stmt)
    results = []
    for log, full_name in res.all():
        log_data = schemas.AuditLogRead.model_validate(log)
        log_data.user_full_name = full_name
        results.append(log_data)
    return results

@router.delete("/{case_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
async def delete_case(case_id: str, db: AsyncSession = Depends(get_async_db)):
    res = await db.execute(select(models.Case).filter(models.Case.id == case_id))
    db_case = res.scalar_one_or_none()
    if db_case is None:
        raise HTTPException(status_code=404, detail="Case not found")
    
    candidate_id = db_case.candidate_id
    
    # Delete related checks first (cascade)
    await db.execute(delete(models.VerificationCheck).filter(models.VerificationCheck.case_id == case_id))
    
    # Delete the case
    await db.delete(db_case)
    
    # Check if candidate has other cases, if not, delete candidate
    if candidate_id:
        other_cases_res = await db.execute(select(func.count(models.Case.id)).filter(models.Case.candidate_id == candidate_id, models.Case.id != case_id))
        count = other_cases_res.scalar() or 0
        if count == 0:
            await db.execute(delete(models.Candidate).filter(models.Candidate.id == candidate_id))
            
    await db.commit()
    return None

from .aws_utils import s3_client, aws_bucket

async def _do_merge(case_id: str, docs: list, candidate_name: str, case_ref: str):
    """Async Background Task for PDF Merge with Concurrent Downloads."""
    import logging
    from fastapi.concurrency import run_in_threadpool
    
    merger = PdfWriter()
    urls = [doc.get('url') for doc in docs if doc.get('url')]
    if not urls:
        return
        
    # Download all documents concurrently
    async with httpx.AsyncClient(timeout=15.0) as client:
        try:
            tasks = [client.get(url) for url in urls]
            responses = await asyncio.gather(*tasks, return_exceptions=True)
        except Exception as e:
            logging.error(f"Failed concurrent download gather: {e}")
            return

    for url, res in zip(urls, responses):
        if isinstance(res, Exception) or res.status_code != 200:
            logging.error(f"Failed download for {url}: {res}")
            continue
        try:
            content = res.content
            if url.lower().endswith('.pdf'):
                # Append PDF in threadpool to keep event loop responsive
                await run_in_threadpool(merger.append, PdfReader(BytesIO(content)))
            else:
                from PIL import Image
                def convert_img():
                    img = Image.open(BytesIO(content)).convert('RGB')
                    buf = BytesIO()
                    img.save(buf, format='PDF')
                    buf.seek(0)
                    return buf
                buf = await run_in_threadpool(convert_img)
                await run_in_threadpool(merger.append, PdfReader(buf))
        except Exception as e:
            logging.error(f"Merge error for {url}: {e}")
            
    if len(merger.pages) > 0:
        out = BytesIO()
        # Non-blocking write
        await run_in_threadpool(merger.write, out)
        out.seek(0)
        filename = f"{candidate_name}_{case_ref}_merged.pdf"
        
        if s3_client and aws_bucket:
            s3_key = f"merged/{case_id}/{filename}"
            # Put object in thread pool
            await run_in_threadpool(s3_client.put_object, Bucket=aws_bucket, Key=s3_key, Body=out.getvalue(), ContentType='application/pdf')
            
            # Non-blocking SQLAlchemy session for DB update
            def update_db():
                db = SessionLocal()
                try:
                    c = db.query(models.Case).filter(models.Case.id == case_id).first()
                    if c:
                        c.merged_pdf_key = s3_key
                        db.commit()
                except Exception as db_err:
                    logging.error(f"DB update error: {db_err}")
                finally:
                    db.close()
            await run_in_threadpool(update_db)

@router.post("/{case_id}/merge-pdfs", status_code=202, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
@limiter.limit("10/minute")
async def merge_pdfs(case_id: str, request: Request, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(select(models.Case).options(joinedload(models.Case.candidate)).filter(models.Case.id == case_id))
    db_case = res.unique().scalar_one_or_none()
    if not db_case or not db_case.candidate:
        raise HTTPException(status_code=404, detail="Case/Candidate not found")
    
    # Tenancy check
    user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    if (user_role == "CUSTOMER" or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER")) and db_case.customer_id != current_user.customer_id:
        raise HTTPException(status_code=403, detail="Unauthorized access to this case")
    
    docs = db_case.candidate.documents or []
    background_tasks.add_task(_do_merge, case_id, docs, db_case.candidate.name, db_case.case_ref_no)
    return {"message": "PDF merge queued"}

@router.post("/{case_id}/generate-report", status_code=202, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
async def generate_report_async(case_id: str, request: Request, current_user: models.User = Depends(get_current_user)):
    """
    Fires a fully decoupled asynchronous rendering pipeline that employs headless Playwright 
    mechanics to output a pixel-perfect final verification vector report.
    """
    auth_header = request.headers.get("Authorization")
    token = None
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
    
    if not token:
        raise HTTPException(status_code=401, detail="Authorization bearer required for headless bootstrap.")
        
    # Import the Celery Task context-locally to ensure dynamic load chain integrity
    from .worker import generate_case_pdf
    
    # Extract client frontend context if passed as query to facilitate local runtime testing
    frontend_override = request.query_params.get("fe_url")
    
    # Hand-off to persistent cluster node
    task = generate_case_pdf.delay(case_id, token, frontend_override)
    
    return {
        "message": "Headless rendering cycle initiated in backend cluster.",
        "task_id": task.id,
        "status": "QUEUED"
    }

@router.post("/{case_id}/finalize-case", response_model=schemas.CaseRead, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
@router.post("/finalize-case", response_model=schemas.CaseRead, dependencies=[Depends(check_module_permission("bvs", "verification", action="write"))])
async def finalize_case(
    case_id: Optional[str] = None,
    data: Optional[schemas.FinalizeCaseRequest] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    target_case_id = case_id or (data.case_id if data else None)
    if not target_case_id:
        raise HTTPException(status_code=400, detail="Case ID is required.")
        
    stmt = select(models.Case).options(
        selectinload(models.Case.checks).selectinload(models.VerificationCheck.documents),
        joinedload(models.Case.candidate),
        joinedload(models.Case.customer)
    ).filter(models.Case.id == target_case_id)
    res = await db.execute(stmt)
    db_case = res.unique().scalar_one_or_none()
    if not db_case:
        raise HTTPException(status_code=404, detail="Case not found.")
        
    # Tenancy check
    user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    if (user_role == "CUSTOMER" or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER")) and db_case.customer_id != current_user.customer_id:
        raise HTTPException(status_code=403, detail="Unauthorized access to this case")
        
    # 1. Validate all checks are complete
    validate_case_completion(db_case)
    
    # 2. Determine Case-level final_result
    if data and data.final_result:
        overall_result = data.final_result.upper()
    else:
        check_results = []
        for chk in db_case.checks:
            res_val = (chk.final_result or chk.status or "").upper()
            check_results.append(res_val)
            
        if any(r in ["NEGATIVE", "RED", "REJECTED"] for r in check_results):
            overall_result = "NEGATIVE"
        elif any(r in ["DISCREPANCY", "AMBER", "YELLOW", "UNABLE_TO_VERIFY"] for r in check_results):
            overall_result = "DISCREPANCY"
        else:
            overall_result = "POSITIVE"
        
    # 3. Transition status and set metadata
    db_case.status = overall_result
    db_case.final_result = overall_result
    db_case.final_report_status = overall_result
    db_case.finalized_by = current_user.id
    db_case.finalized_at = datetime.utcnow()
    db_case.completed_date = datetime.utcnow()
    
    if data and data.remarks:
        db_case.final_remarks = data.remarks
        
    # Save & Log
    db.add(db_case)
    await db.commit()
    
    await create_audit_log(db, current_user.id, "CASE_FINALIZED", f"Case {db_case.case_ref_no} finalized with result: {overall_result}", resource_id=target_case_id)
    
    # Trigger notifications
    try:
        await notification_utils.notify_case_closed(
            db, target_case_id, db_case.case_ref_no, db_case.candidate.name if db_case.candidate else "", overall_result
        )
    except Exception as notif_err:
        logger.error(f"Error sending case closed notification: {notif_err}")
        
    # Broadcast to workforce dashboard
    try:
        await manager.broadcast({"type": "WORKFORCE_UPDATE", "source": "finalize_case"})
    except Exception as bc_err:
        logger.error(f"Error broadcasting workforce update: {bc_err}")
        
    # Invalidate caches
    try:
        await invalidate_dashboard_cache()
    except Exception as cache_err:
        logger.error(f"Error invalidating dashboard cache: {cache_err}")
        
    # Reload case for returning schemas.CaseRead
    stmt_reload = select(models.Case).options(
        selectinload(models.Case.checks).selectinload(models.VerificationCheck.documents),
        joinedload(models.Case.candidate),
        joinedload(models.Case.customer)
    ).filter(models.Case.id == target_case_id)
    res_reload = await db.execute(stmt_reload)
    db_case_reloaded = res_reload.unique().scalar_one()
    
    return db_case_reloaded

@router.post("/bulk-allocate")
async def bulk_allocate(data: schemas.BulkAllocateRequest, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    case_ids = data.case_ids
    user_id = data.user_id
    logger.info(f"Bulk allocate requested for user_id: {user_id}, case_count: {len(case_ids)}")
    
    update_vals = {
        "assigned_to": user_id,
        "assigned_at": datetime.utcnow() if user_id else None,
        "status": models.CaseStatus.VERIFICATION if user_id else models.CaseStatus.PENDING
    }
    
    await db.execute(
        update(models.Case)
        .where(models.Case.id.in_(case_ids))
        .values(**update_vals)
    )
    
    # Trigger Real-time Notifications & Broadcasts
    if user_id:
        v_res = await db.execute(select(models.User).filter(models.User.id == user_id))
        v_obj = v_res.scalar_one_or_none()
        verifier_name = v_obj.full_name if v_obj else "Verifier"

        bulk_info = []
        for cid in case_ids:
            ref_res = await db.execute(select(models.Case).options(joinedload(models.Case.candidate)).filter(models.Case.id == cid))
            cbd = ref_res.scalar_one_or_none()
            if not cbd: continue
            
            candidate_name = cbd.candidate.name if cbd.candidate else "Candidate"
            bulk_info.append({"id": cid, "ref": cbd.case_ref_no, "candidate": candidate_name})
            await notification_utils.notify_new_assignment(db, user_id, cbd.case_ref_no, cid, candidate_name, background_tasks=background_tasks)
            await manager.broadcast({"type": "CASE_UPDATED", "case_id": cid, "action": "allocation"})

        # Notify Admin ONE summary
        await notification_utils.create_notification(
            db, current_user.id,
            "Allocation Strategy Finalized",
            f"Successfully deployed {len(case_ids)} cases to {verifier_name}. Mission protocols are now active.",
            enums.NotificationCategory.SYSTEM_ALERT,
            extra_data={"type": "BULK_ALLOCATE", "verifier": verifier_name, "cases": bulk_info}
        )

    await db.commit()
    await invalidate_dashboard_cache()
    await manager.broadcast({"type": "WORKFORCE_UPDATE", "source": "bulk_allocation"})
    return {"message": "Success"}

@router.post("/ocr-extract")
async def ocr_extract(data: schemas.OcrExtractRequest, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_async_db)):
    url = data.url
    if not url: raise HTTPException(status_code=400, detail="Document URL required")
    
    try:
        import requests
        import logging
        from fastapi.concurrency import run_in_threadpool
        logger = logging.getLogger(__name__)
        
        # Offload blocking HTTP request to thread pool
        response = await run_in_threadpool(requests.get, url, timeout=10)
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail="Failed to fetch document")
            
        from .ocr_utils import get_scanner
        scanner = get_scanner()
        
        # Offload blocking EasyOCR model execution to thread pool
        text = await run_in_threadpool(scanner.reader.readtext, response.content, detail=0)
        full_text = " ".join(text)
        
        # Offload blocking ID parsing to thread pool
        extracted = await run_in_threadpool(scanner.parse_id, full_text)
        
        return {
            "success": True,
            "extracted_data": extracted,
            "raw_text_debug": str(full_text)[:500] 
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{case_id}/ai-summary")
async def generate_ai_summary(case_id: str, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    stmt = select(models.Case).options(
        joinedload(models.Case.candidate),
        selectinload(models.Case.checks)
    ).filter(models.Case.id == case_id)
    
    res = await db.execute(stmt)
    db_case = res.unique().scalar_one_or_none()
    if not db_case:
        raise HTTPException(status_code=404, detail="Case not found")
        
    # Tenancy Check
    user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    if (user_role == "CUSTOMER" or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER")) and db_case.customer_id != current_user.customer_id:
        raise HTTPException(status_code=403, detail="Unauthorized access to this case")
        
    checks = db_case.checks
    summary_parts = []
    
    # 1. Overall Status
    status_counts = {}
    for c in checks:
        status_counts[c.status] = status_counts.get(c.status, 0) + 1
    
    overall = "GREEN" if status_counts.get("GREEN") == len(checks) else "AMBER" if status_counts.get("RED") else "GREEN"
    if status_counts.get("RED"): overall = "RED"
    
    name = db_case.candidate.name if db_case.candidate else "The candidate"
    summary_parts.append(f"Verification Summary for {name}:")
    summary_parts.append(f"The overall verification status is {overall}.")
    
    # 2. Key Findings
    findings = []
    for c in checks:
        if c.status == "GREEN":
            findings.append(f"• {c.check_type}: Verified successfully with no discrepancies.")
        elif c.status == "RED":
            findings.append(f"• {c.check_type}: CRITICAL DISCREPANCY FOUND. {c.verifier_remarks or 'Verification failed.'}")
        elif c.status == "AMBER":
            findings.append(f"• {c.check_type}: Minor discrepancy noted. {c.verifier_remarks or 'Check results were clear with minor remarks.'}")
        else:
            findings.append(f"• {c.check_type}: Verification is currently {c.status}.")
            
    summary_parts.extend(findings)
    
    # 3. Final Conclusion
    if overall == "GREEN":
        summary_parts.append("\nConclusion: All provided credentials and background details have been verified as authentic. The candidate is cleared for further processing.")
    elif overall == "RED":
        summary_parts.append("\nConclusion: Due to critical discrepancies found in one or more verification modules, extreme caution is advised. Further internal review is recommended.")
    else:
        summary_parts.append("\nConclusion: The verification process revealed minor inconsistencies. Please review the specific notes for each module.")
        
    full_summary = "\n".join(summary_parts)
    
    # Save to DB
    db_case.ai_summary = full_summary
    await db.commit()
    
    return {"summary": full_summary}

@router.get("/{case_id}/comments", response_model=List[schemas.CaseComment])
async def read_case_comments(case_id: str, db: AsyncSession = Depends(get_async_db)):
    stmt = (
        select(models.CaseComment, models.User.full_name.label("user_full_name"))
        .outerjoin(models.User, models.CaseComment.user_id == models.User.id)
        .filter(models.CaseComment.case_id == case_id)
        .order_by(models.CaseComment.created_at.asc())
    )
    res = await db.execute(stmt)
    results = []
    for comment, full_name in res.all():
        data = schemas.CaseComment.model_validate(comment)
        data.user_full_name = full_name
        results.append(data)
    return results

@router.post("/{case_id}/comments", response_model=schemas.CaseComment)
async def create_case_comment(
    case_id: str, 
    comment: schemas.CaseCommentCreate, 
    db: AsyncSession = Depends(get_async_db), 
    current_user: models.User = Depends(get_current_user)
):
    db_comment = models.CaseComment(
        case_id=case_id,
        user_id=current_user.id,
        content=comment.content
    )
    db.add(db_comment)
    await db.commit()
    
    # Reload with user name
    res = await db.execute(
        select(models.CaseComment, models.User.full_name.label("user_full_name"))
        .outerjoin(models.User, models.CaseComment.user_id == models.User.id)
        .filter(models.CaseComment.id == db_comment.id)
    )
    comment_model, full_name = res.one()
    data = schemas.CaseComment.model_validate(comment_model)
    data.user_full_name = full_name
    
    # Broadcast to WebSockets
    from .ws import manager
    await manager.broadcast({
        "type": "NEW_COMMENT",
        "case_id": case_id,
        "user": full_name,
        "content": comment.content
    })
    
    return data

# ─── Revoke Log Endpoints ────────────────────────────────────────────────────

@router.get("/{case_id}/revoke-logs", dependencies=[Depends(get_current_user)])
async def get_case_revoke_logs(case_id: str, db: AsyncSession = Depends(get_async_db)):
    """Return all revoke events for a specific case."""
    stmt = (
        select(models.RevokeLog, models.User.full_name, models.Case.case_ref_no, models.User.role, models.Role.name.label("custom_role_name"))
        .outerjoin(models.User, models.RevokeLog.user_id == models.User.id)
        .outerjoin(models.Role, models.User.role_id == models.Role.id)
        .outerjoin(models.Case, models.RevokeLog.case_id == models.Case.id)
        .filter(models.RevokeLog.case_id == case_id)
        .order_by(models.RevokeLog.revoked_at.desc())
    )
    result = await db.execute(stmt)
    rows = result.all()
    return [
        {
            "id": r.id,
            "case_id": r.case_id,
            "case_ref_no": ref_no,
            "user_id": r.user_id,
            "user_name": full_name,
            "user_role": custom_role_name if custom_role_name else ("QC Verifier" if str(role.value if hasattr(role, 'value') else role).upper() in ["QA", "QC"] else str(role.value if hasattr(role, 'value') else role).replace('_', ' ').title()),
            "revoke_type": r.revoke_type,
            "from_status": r.from_status,
            "to_status": r.to_status,
            "notes": r.notes,
            "revoked_at": r.revoked_at.isoformat() if r.revoked_at else None,
        }
        for r, full_name, ref_no, role, custom_role_name in rows
    ]

@router.get("/revoke-logs/all", dependencies=[Depends(get_current_user)])
async def get_all_revoke_logs(
    revoke_type: Optional[str] = None,
    user_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db)
):
    """Global revoke tracking: list all revoke events with filters."""
    stmt = (
        select(models.RevokeLog, models.User.full_name, models.Case.case_ref_no, models.User.role, models.Role.name.label("custom_role_name"))
        .outerjoin(models.User, models.RevokeLog.user_id == models.User.id)
        .outerjoin(models.Role, models.User.role_id == models.Role.id)
        .outerjoin(models.Case, models.RevokeLog.case_id == models.Case.id)
        .order_by(models.RevokeLog.revoked_at.desc())
    )
    if revoke_type:
        stmt = stmt.filter(models.RevokeLog.revoke_type == revoke_type.upper())
    if user_id:
        stmt = stmt.filter(models.RevokeLog.user_id == user_id)
    if from_date:
        try:
            dt = datetime.fromisoformat(from_date)
            stmt = stmt.filter(models.RevokeLog.revoked_at >= dt)
        except ValueError:
            pass
    if to_date:
        try:
            dt = datetime.fromisoformat(to_date)
            stmt = stmt.filter(models.RevokeLog.revoked_at <= dt)
        except ValueError:
            pass

    result = await db.execute(stmt)
    rows = result.all()

    # Aggregate counts per user
    user_summary: Dict[str, Any] = {}
    all_logs = []
    for r, full_name, ref_no, role, custom_role_name in rows:
        u_role_val = str(role.value if hasattr(role, 'value') else role).upper()
        ur_string = custom_role_name if custom_role_name else ("QC Verifier" if u_role_val in ["QA", "QC"] else u_role_val.replace('_', ' ').title())
        log = {
            "id": r.id,
            "case_id": r.case_id,
            "case_ref_no": ref_no,
            "user_id": r.user_id,
            "user_name": full_name,
            "user_role": ur_string,
            "revoke_type": r.revoke_type,
            "from_status": r.from_status,
            "to_status": r.to_status,
            "notes": r.notes,
            "revoked_at": r.revoked_at.isoformat() if r.revoked_at else None,
        }
        all_logs.append(log)

        key = r.user_id
        if key not in user_summary:
            user_summary[key] = {
                "user_id": r.user_id,
                "user_name": full_name,
                "user_role": ur_string,
                "verifier_revoke_count": 0,
                "qc_revoke_count": 0,
                "total_revoke_count": 0,
                "cases": set()
            }
        if r.revoke_type == "VERIFIER":
            user_summary[key]["verifier_revoke_count"] += 1
        elif r.revoke_type == "QC":
            user_summary[key]["qc_revoke_count"] += 1
        user_summary[key]["total_revoke_count"] += 1
        user_summary[key]["cases"].add(ref_no)

    summary = []
    for v in user_summary.values():
        v["cases"] = list(v["cases"])
        summary.append(v)

    return {"logs": all_logs, "user_summary": summary}

# ─── Insufficiency Log Endpoints ───────────────────────────────────────────

@router.get("/{case_id}/insufficiency-logs", dependencies=[Depends(get_current_user)])
async def get_case_insufficiency_logs(case_id: str, db: AsyncSession = Depends(get_async_db)):
    """Return all insufficiency events for a specific case."""
    stmt = (
        select(models.InsufficiencyLog, models.User.full_name, models.Case.case_ref_no, models.User.role, models.Role.name.label("custom_role_name"))
        .outerjoin(models.User, models.InsufficiencyLog.user_id == models.User.id)
        .outerjoin(models.Role, models.User.role_id == models.Role.id)
        .outerjoin(models.Case, models.InsufficiencyLog.case_id == models.Case.id)
        .filter(models.InsufficiencyLog.case_id == case_id)
        .order_by(models.InsufficiencyLog.marked_at.desc())
    )
    result = await db.execute(stmt)
    rows = result.all()
    return [
        {
            "id": r.id,
            "case_id": r.case_id,
            "case_ref_no": ref_no,
            "user_id": r.user_id,
            "user_name": full_name,
            "user_role": custom_role_name if custom_role_name else str(role.value if hasattr(role, 'value') else role).replace('_', ' ').title(),
            "from_status": r.from_status,
            "notes": r.notes,
            "marked_at": r.marked_at.isoformat() if r.marked_at else None,
            "resolved_at": r.resolved_at.isoformat() if r.resolved_at else None,
        }
        for r, full_name, ref_no, role, custom_role_name in rows
    ]

@router.get("/insufficiency-logs/all", dependencies=[Depends(get_current_user)])
async def get_all_insufficiency_logs(
    user_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db)
):
    """Global insufficiency tracking: list all records from insufficiencies table."""
    stmt = (
        select(
            models.Insufficiency, 
            models.User.full_name.label("user_name"), 
            models.Case.case_ref_no, 
            models.Candidate.name.label("candidate_name"),
            models.Customer.name.label("customer_name"),
            models.VerificationCheck.check_type.label("check_name"),
            models.Role.name.label("custom_role_name"),
            models.User.role.label("user_role_enum")
        )
        .join(models.Case, models.Insufficiency.case_id == models.Case.id)
        .join(models.VerificationCheck, models.Insufficiency.check_id == models.VerificationCheck.id)
        .join(models.User, models.Insufficiency.raised_by == models.User.id)
        .outerjoin(models.Candidate, models.Case.candidate_id == models.Candidate.id)
        .outerjoin(models.Customer, models.Case.customer_id == models.Customer.id)
        .outerjoin(models.Role, models.User.role_id == models.Role.id)
        .order_by(models.Insufficiency.created_at.desc())
    )
    
    if user_id:
        stmt = stmt.filter(models.Insufficiency.raised_by == user_id)
    if from_date:
        try:
            dt = datetime.fromisoformat(from_date)
            stmt = stmt.filter(models.Insufficiency.created_at >= dt)
        except ValueError:
            pass
    if to_date:
        try:
            dt = datetime.fromisoformat(to_date)
            stmt = stmt.filter(models.Insufficiency.created_at <= dt)
        except ValueError:
            pass

    result = await db.execute(stmt)
    rows = result.all()

    all_logs = []
    user_summary = {}

    for row in rows:
        r = row[0] # models.Insufficiency
        role_enum = row.user_role_enum
        u_role_val = str(role_enum.value if hasattr(role_enum, 'value') else role_enum).upper()
        
        log = {
            "id": r.id,
            "case_id": r.case_id,
            "case_ref_no": row.case_ref_no,
            "candidate_name": row.candidate_name or "N/A",
            "customer_name": row.customer_name or "N/A",
            "check_name": row.check_name or "General",
            "user_id": r.raised_by,
            "user_name": row.user_name,
            "user_role": row.custom_role_name if row.custom_role_name else u_role_val.replace('_', ' ').title(),
            "marked_at": r.created_at.isoformat() if r.created_at else None,
            "resolved_at": r.resolved_at.isoformat() if r.resolved_at else None,
            "notes": r.message,
            "status": r.status or ("RESOLVED" if r.is_resolved else "INSUFFICIENT")
        }
        all_logs.append(log)

        key = r.raised_by
        if key not in user_summary:
            user_summary[key] = {
                "user_id": r.raised_by,
                "user_name": row.user_name,
                "user_role": log["user_role"],
                "total_marked": 0,
                "total_resolved": 0,
                "cases": set()
            }
        user_summary[key]["total_marked"] += 1
        if r.is_resolved or r.status == "RESOLVED":
            user_summary[key]["total_resolved"] += 1
        user_summary[key]["cases"].add(row.case_ref_no)

    summary = []
    for v in user_summary.values():
        v["cases"] = list(v["cases"])
        summary.append(v)

    return {"logs": all_logs, "user_summary": summary}

@router.get("/insufficiencies/{id}")
async def get_insufficiency_detail(id: str, db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    """Fetch detailed insufficiency record with unified timeline."""
    stmt = (
        select(models.Insufficiency)
        .options(
            joinedload(models.Insufficiency.case).joinedload(models.Case.candidate),
            joinedload(models.Insufficiency.check),
            joinedload(models.Insufficiency.user),
            joinedload(models.Insufficiency.resolver),
            joinedload(models.Insufficiency.updater)
        )
        .filter(models.Insufficiency.id == id)
    )
    res = await db.execute(stmt)
    insuff = res.unique().scalar_one_or_none()
    if not insuff:
        raise HTTPException(404, detail="Insufficiency record not found")
        
    # Tenancy Check
    user_role = str(current_user.role.value if hasattr(current_user.role, 'value') else current_user.role).upper()
    if (user_role == "CUSTOMER" or (current_user.role_rel and current_user.role_rel.name.upper() == "CUSTOMER")) and insuff.case.customer_id != current_user.customer_id:
        raise HTTPException(status_code=403, detail="Unauthorized access to this record")

    # Construct Timeline
    timeline = []
    
    # 1. Raised Event
    timeline.append({
        "event": "Insufficiency Raised",
        "time": insuff.created_at.isoformat() if insuff.created_at else None,
        "user": insuff.user.full_name if insuff.user else "System",
        "role": insuff.role or "Verifier",
        "note": insuff.message,
        "type": "RAISED"
    })
    
    # 2. Customer Upload Event (if documents present)
    if insuff.documents and len(insuff.documents) > 0:
        timeline.append({
            "event": "Customer Uploaded Evidence",
            "time": insuff.updated_at.isoformat() if insuff.updated_at else None,
            "user": "Customer Portal",
            "note": f"{len(insuff.documents)} document(s) submitted for review.",
            "documents": insuff.documents,
            "type": "UPLOADED"
        })

    # 3. Review / Resolution Event
    if insuff.is_resolved:
        timeline.append({
            "event": "Resolved",
            "time": insuff.resolved_at.isoformat() if insuff.resolved_at else None,
            "user": insuff.resolver.full_name if insuff.resolver else "System",
            "note": insuff.resolved_remarks,
            "type": "RESOLVED"
        })
    elif insuff.status == "IN_REVIEW":
        timeline.append({
            "event": "Verifier Review in Progress",
            "time": insuff.updated_at.isoformat() if insuff.updated_at else None,
            "user": insuff.updater.full_name if insuff.updater else "Verifier",
            "note": "Documentation is being reviewed by the operations team.",
            "type": "REVIEW"
        })

    return {
        "id": insuff.id,
        "status": insuff.status or ("RESOLVED" if insuff.is_resolved else "INSUFFICIENT"),
        "is_resolved": insuff.is_resolved,
        "message": insuff.message,
        "documents": insuff.documents or [],
        "case": {
            "id": insuff.case.id,
            "ref_no": insuff.case.case_ref_no,
            "candidate": insuff.case.candidate.name if insuff.case.candidate else "N/A"
        },
        "check": {
            "id": insuff.check.id,
            "name": insuff.check.check_type
        },
        "raised_by": insuff.user.full_name if insuff.user else "Verifier",
        "timeline": timeline
    }

@router.post("/insufficiencies/{id}/evidence")
async def update_insufficiency_evidence(id: str, data: Dict[str, Any], db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    """Customer uploads evidence to resolve insufficiency."""
    stmt = select(models.Insufficiency).filter(models.Insufficiency.id == id)
    res = await db.execute(stmt)
    insuff = res.scalar_one_or_none()
    if not insuff:
        raise HTTPException(404, detail="Insufficiency not found")
        
    insuff.documents = data.get("documents", [])
    insuff.status = "CUSTOMER_UPLOADED"
    insuff.updated_at = datetime.utcnow()
    # In a real app, we might set updated_by to current_user.id if it's a customer
    
    await db.commit()
    return {"status": "success", "message": "Evidence uploaded successfully"}

@router.post("/insufficiencies/{id}/resolve")
async def resolve_insufficiency(id: str, data: Dict[str, str], db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    """Mark an insufficiency as resolved."""
    stmt = select(models.Insufficiency).filter(models.Insufficiency.id == id)
    res = await db.execute(stmt)
    insuff = res.scalar_one_or_none()
    if not insuff:
        raise HTTPException(404, detail="Insufficiency not found")
        
    insuff.is_resolved = True
    insuff.status = "RESOLVED"
    insuff.resolved_at = datetime.utcnow()
    insuff.resolved_by = current_user.id
    insuff.resolved_remarks = data.get("remarks", "Resolved via Audit Terminal")
    
    # Log audit
    audit_log = models.AuditLog(
        user_id=current_user.id,
        action="INSUFFICIENCY_RESOLVED",
        resource_id=insuff.case_id,
        details=f"Insufficiency resolved for check {insuff.check_id}. Remarks: {insuff.resolved_remarks}"
    )
    db.add(audit_log)
    
    # Auto-resolve case status if no more open insufficiencies
    rem_stmt = select(func.count(models.Insufficiency.id)).filter(models.Insufficiency.case_id == insuff.case_id, models.Insufficiency.is_resolved == False)
    rem_res = await db.execute(rem_stmt)
    if rem_res.scalar() == 0:
        case_stmt = select(models.Case).filter(models.Case.id == insuff.case_id)
        c_res = await db.execute(case_stmt)
        db_case = c_res.scalar_one()
        if db_case.status == "INSUFFICIENT":
            db_case.status = "VERIFICATION"
            
    # Also update the check status
    check_stmt = select(models.VerificationCheck).filter(models.VerificationCheck.id == insuff.check_id)
    ck_res = await db.execute(check_stmt)
    check = ck_res.scalar_one_or_none()
    if check:
        check.status = "VERIFICATION"
            
    await db.commit()
    return {"status": "success"}

@router.post("/insufficiencies/{id}/reject")
async def reject_insufficiency(id: str, data: Dict[str, str], db: AsyncSession = Depends(get_async_db), current_user: models.User = Depends(get_current_user)):
    """Reject customer evidence and revert to INSUFFICIENT."""
    stmt = select(models.Insufficiency).filter(models.Insufficiency.id == id)
    res = await db.execute(stmt)
    insuff = res.scalar_one_or_none()
    if not insuff:
        raise HTTPException(404, detail="Insufficiency not found")
        
    insuff.status = "INSUFFICIENT"
    insuff.updated_at = datetime.utcnow()
    insuff.updated_by = current_user.id
    insuff.message = f"REJECTED: {data.get('remarks', 'Evidence rejected by verifier')}. Previous message: {insuff.message}"
    
    await db.commit()
    return {"status": "success"}

    return {"status": "success"}

@router.post("/sync-risk", dependencies=[Depends(get_current_user)])
async def sync_all_case_risks(db: AsyncSession = Depends(get_async_db)):
    """Trigger the predictive risk assessment for all active cases."""
    await risk_utils.update_all_case_risks(db)
    return {"message": "SLA Risk Analysis completed successfully"}
