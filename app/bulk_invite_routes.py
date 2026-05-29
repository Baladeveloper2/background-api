"""
Bulk Candidate Invitation Routes
Handles:
  - POST /bulk-invite/candidates  → bulk create candidates + cases
  - POST /bulk-invite/send-links  → bulk send BGV links
  - GET  /bulk-invite/template    → download Excel sample file
"""
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from sqlalchemy.orm import joinedload
from typing import List, Optional
from pydantic import BaseModel, EmailStr
from datetime import datetime
import io

from .database import get_async_db
from .auth_routes import get_current_user
from . import models, enums, email_utils, notification_utils

router = APIRouter(prefix="/bulk-invite", tags=["Bulk Invite"])


# ─── Schemas ─────────────────────────────────────────────────────────────────

class BulkCandidateItem(BaseModel):
    emp_id: Optional[str] = None
    name: str
    email: str
    phone: Optional[str] = None
    dob: Optional[str] = None
    gender: Optional[str] = None
    aadhar_number: Optional[str] = None
    pan_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    client_reference_id: Optional[str] = None
    customer_id: Optional[str] = None


class BulkCreateRequest(BaseModel):
    candidates: List[BulkCandidateItem]
    checks: List[str] = []
    send_links: bool = False
    customer_id: Optional[str] = None


class BulkSendLinksRequest(BaseModel):
    case_ids: List[str]
    checks: List[str]


# ─── Template Download ────────────────────────────────────────────────────────

@router.get("/template")
async def download_bulk_template():
    """Download the Excel sample template for bulk candidate import."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates"

        headers = [
            "EMPLOYEE_ID", "CANDIDATE_NAME", "EMAIL", "PHONE"
        ]

        required_cols = {"EMPLOYEE_ID", "CANDIDATE_NAME", "EMAIL"}

        # Header styles
        header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
        required_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        ws.row_dimensions[1].height = 32
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = required_fill if header in required_cols else header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Sample rows
        sample_rows = [
            ["EMP001", "John Smith", "john.smith@company.com", "+91 9876543210"],
            ["EMP002", "Priya Sharma", "priya.sharma@company.com", "+91 9123456789"],
        ]

        data_fill_odd = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
        data_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(name="Calibri", size=10)

        for row_idx, row_data in enumerate(sample_rows, start=2):
            ws.row_dimensions[row_idx].height = 22
            fill = data_fill_odd if row_idx % 2 else data_fill_even
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                cell.fill = fill

        # Column widths
        col_widths = [14, 22, 32, 18]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Candidate_Bulk_Import_Template.xlsx"}
        )

    except ImportError:
        # Fallback: plain CSV
        csv_content = (
            "EMPLOYEE_ID,CANDIDATE_NAME,EMAIL,PHONE\n"
            "EMP001,John Smith,john@company.com,9876543210\n"
        )
        return StreamingResponse(
            io.BytesIO(csv_content.encode()),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=Candidate_Bulk_Import_Template.csv"}
        )


# ─── Bulk Create Candidates ───────────────────────────────────────────────────

@router.post("/candidates")
async def bulk_create_candidates(
    payload: BulkCreateRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Bulk create candidates + cases. Optionally send BGV links immediately.
    Returns per-row results with status and case_id.
    """
    final_customer_id = (
        payload.customer_id
        or (current_user.customer_id if hasattr(current_user, 'customer_id') else None)
    )
    if not final_customer_id:
        raise HTTPException(status_code=400, detail="Customer association missing.")

    # Load customer for prefix
    cust_res = await db.execute(select(models.Customer).filter(models.Customer.id == final_customer_id))
    customer = cust_res.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found.")

    sc = customer.short_code if customer.short_code else customer.name[:3].upper()
    prefix = f"CL-{sc}-"

    # Current case count
    count_res = await db.execute(
        select(func.count(models.Case.id)).filter(models.Case.customer_id == final_customer_id)
    )
    base_count = count_res.scalar() or 0

    # Find/create daily batch
    date_str = datetime.utcnow().strftime('%Y-%m-%d')
    batch_name = f"Batch-{sc}-{date_str}"
    batch_res = await db.execute(
        select(models.Batch).filter(
            models.Batch.customer_id == final_customer_id,
            models.Batch.batch_no == batch_name
        ).limit(1)
    )
    inv_batch = batch_res.scalar_one_or_none()

    if not inv_batch:
        batch_count_res = await db.execute(
            select(func.count(models.Batch.id)).filter(models.Batch.customer_id == final_customer_id)
        )
        batch_count = batch_count_res.scalar() or 0
        inv_batch = models.Batch(
            customer_id=final_customer_id,
            batch_no=batch_name,
            cl_ref_no=f"CL-{sc}-{(batch_count + 1):03d}",
            cases_count=0,
            upload_date=datetime.utcnow()
        )
        db.add(inv_batch)
        await db.flush()

    results = []
    suffix_num = base_count + 1

    frontend_url = __import__("os").getenv("FRONTEND_URL", "https://background-verification-91d11.web.app")

    for item in payload.candidates:
        try:
            # Duplicate check — email
            dup_email_res = await db.execute(
                select(models.Candidate.id).filter(models.Candidate.email == item.email)
            )
            if dup_email_res.scalar_one_or_none():
                results.append({
                    "name": item.name, "email": item.email, "emp_id": item.emp_id,
                    "status": "DUPLICATE_EMAIL", "case_id": None,
                    "error": "Email already exists in the system"
                })
                continue

            # Create candidate
            candidate = models.Candidate(
                name=item.name,
                email=item.email,
                phone=item.phone,
                client_emp_code=item.emp_id
            )
            db.add(candidate)
            await db.flush()

            # Unique case_ref
            while True:
                case_ref = f"{prefix}{str(suffix_num).zfill(3)}"
                exists_res = await db.execute(
                    select(models.Case.id).filter(models.Case.case_ref_no == case_ref)
                )
                if not exists_res.scalar_one_or_none():
                    break
                suffix_num += 1

            new_case = models.Case(
                case_ref_no=case_ref,
                customer_id=final_customer_id,
                candidate_id=candidate.id,
                batch_id=inv_batch.id,
                status=enums.CaseStatus.PENDING
            )
            db.add(new_case)
            await db.flush()
            inv_batch.cases_count = (inv_batch.cases_count or 0) + 1
            suffix_num += 1

            # If send_links requested — add checks and send email
            if payload.send_links and payload.checks:
                for check_type in payload.checks:
                    db.add(models.VerificationCheck(
                        case_id=new_case.id,
                        check_type=check_type,
                        status=enums.CheckStatus.VERIFICATION
                    ))
                new_case.status = enums.CaseStatus.LINK_SHARED
                new_case.link_shared_at = datetime.utcnow()
                form_link = f"{frontend_url}/candidate/form/{new_case.id}"
                background_tasks.add_task(
                    email_utils.send_bgv_invitation_email,
                    to_email=item.email,
                    candidate_name=item.name,
                    form_link=form_link
                )

            results.append({
                "name": item.name,
                "email": item.email,
                "emp_id": item.emp_id,
                "status": "LINK_SENT" if payload.send_links else "CREATED",
                "case_id": new_case.id,
                "case_ref": case_ref,
                "error": None
            })

        except Exception as e:
            results.append({
                "name": item.name, "email": item.email, "emp_id": item.emp_id,
                "status": "ERROR", "case_id": None, "error": str(e)
            })

    await db.commit()

    success = [r for r in results if r["status"] in ("CREATED", "LINK_SENT")]
    failed = [r for r in results if r["status"] not in ("CREATED", "LINK_SENT")]

    return {
        "total": len(results),
        "success": len(success),
        "failed": len(failed),
        "results": results
    }


# ─── Bulk Send Links ──────────────────────────────────────────────────────────

@router.post("/send-links")
async def bulk_send_links(
    payload: BulkSendLinksRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db),
    current_user: models.User = Depends(get_current_user)
):
    """Send BGV verification links to multiple existing cases at once."""
    if not payload.case_ids:
        raise HTTPException(status_code=400, detail="No case IDs provided.")

    frontend_url = __import__("os").getenv("FRONTEND_URL", "https://background-verification-91d11.web.app")
    results = []

    for case_id in payload.case_ids:
        stmt = select(models.Case).filter(models.Case.id == case_id).options(joinedload(models.Case.candidate))
        res = await db.execute(stmt)
        case = res.scalar_one_or_none()

        if not case:
            results.append({"case_id": case_id, "status": "NOT_FOUND", "error": "Case not found"})
            continue

        try:
            # Add checks (avoid duplicates)
            existing_res = await db.execute(
                select(models.VerificationCheck.check_type).filter(models.VerificationCheck.case_id == case.id)
            )
            existing_types = {c[0] for c in existing_res.all()}
            for check_type in payload.checks:
                if check_type not in existing_types:
                    db.add(models.VerificationCheck(
                        case_id=case.id,
                        check_type=check_type,
                        status=enums.CheckStatus.VERIFICATION
                    ))

            case.status = enums.CaseStatus.LINK_SHARED
            case.link_shared_at = datetime.utcnow()

            form_link = f"{frontend_url}/candidate/form/{case.id}"
            if case.candidate and case.candidate.email:
                background_tasks.add_task(
                    email_utils.send_bgv_invitation_email,
                    to_email=case.candidate.email,
                    candidate_name=case.candidate.name,
                    form_link=form_link
                )

            results.append({
                "case_id": case_id,
                "status": "SENT",
                "candidate": case.candidate.name if case.candidate else "Unknown"
            })
        except Exception as e:
            results.append({"case_id": case_id, "status": "ERROR", "error": str(e)})

    await db.commit()
    success = sum(1 for r in results if r["status"] == "SENT")
    return {"total": len(results), "success": success, "failed": len(results) - success, "results": results}
